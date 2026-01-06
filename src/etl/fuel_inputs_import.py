"""
Fuel model inputs import from Excel.

Imports use factors, heat rates, and coal contracts from Excel workbooks
into the database. Supports the OVEC fuel cost model Excel structure.

Functions:
    import_use_factors_from_excel: Import monthly use factors
    import_heat_rates_from_excel: Import monthly heat rates
    import_coal_contracts_from_excel: Import coal contract data
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from decimal import Decimal
from datetime import date

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


def import_use_factors_from_excel(
    db: Session,
    file_path: str,
    plant_id: int,
    year: int,
    sheet_name: str = "Use Factor Input",
) -> Dict:
    """Import use factors from Excel workbook.
    
    Expected Excel structure:
    - Row with plant name and monthly values (Jan-Dec in columns)
    - Separate rows for base use factor and ozone/non-SCR factors
    
    Args:
        db: Database session
        file_path: Path to Excel file
        plant_id: Plant ID to import for
        year: Year to import
        sheet_name: Name of the sheet with use factors
        
    Returns:
        Dictionary with import results
    """
    import openpyxl
    from src.models.use_factor import upsert_use_factor
    
    results = {
        "file": file_path,
        "sheet": sheet_name,
        "plant_id": plant_id,
        "year": year,
        "months_imported": 0,
        "errors": [],
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            results["errors"].append(f"Sheet '{sheet_name}' not found")
            return results
        
        ws = wb[sheet_name]
        
        # Find the plant row
        plant_name = "Kyger Creek" if plant_id == 1 else "Clifty Creek"
        plant_row = None
        ozone_row = None
        
        for row_idx in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and isinstance(cell_value, str):
                if plant_name.lower() in cell_value.lower():
                    if "ozone" in cell_value.lower() or "unit 6" in cell_value.lower():
                        ozone_row = row_idx
                    else:
                        plant_row = row_idx
        
        if not plant_row:
            results["errors"].append(f"Could not find row for {plant_name}")
            return results
        
        # Find month columns (usually B-M for Jan-Dec)
        # Look for month headers in first few rows
        month_cols = {}
        month_names = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
            "january": 1, "february": 2, "march": 3, "april": 4, 
            "june": 6, "july": 7, "august": 8, "september": 9,
            "october": 10, "november": 11, "december": 12,
        }
        
        for row_idx in range(1, 5):
            for col_idx in range(2, 14):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    for name, month_num in month_names.items():
                        if name in cell_lower:
                            month_cols[month_num] = col_idx
                            break
        
        # If no headers found, assume columns B-M are Jan-Dec
        if not month_cols:
            for i, col in enumerate(range(2, 14), start=1):
                month_cols[i] = col
        
        # Extract use factor values
        for month, col in month_cols.items():
            try:
                base_value = ws.cell(row=plant_row, column=col).value
                
                if base_value is not None:
                    # Convert to decimal (0-1 scale)
                    if isinstance(base_value, (int, float)):
                        if base_value > 1:  # Assume percentage
                            base_value = base_value / 100
                        base_uf = float(base_value)
                    else:
                        continue
                    
                    # Get ozone factor if available
                    ozone_uf = base_uf  # Default to same as base
                    if ozone_row:
                        ozone_value = ws.cell(row=ozone_row, column=col).value
                        if ozone_value is not None and isinstance(ozone_value, (int, float)):
                            if ozone_value > 1:
                                ozone_value = ozone_value / 100
                            ozone_uf = float(ozone_value)
                    
                    # For non-ozone months, ozone factor doesn't apply
                    if month not in {5, 6, 7, 8, 9}:
                        ozone_uf = base_uf
                    
                    upsert_use_factor(
                        db,
                        plant_id=plant_id,
                        year=year,
                        month=month,
                        use_factor_base=base_uf,
                        use_factor_ozone_non_scr=ozone_uf,
                        notes=f"Imported from Excel: {Path(file_path).name}",
                    )
                    results["months_imported"] += 1
                    
            except Exception as e:
                results["errors"].append(f"Month {month}: {str(e)}")
        
        wb.close()
        
    except Exception as e:
        results["errors"].append(str(e))
        logger.error(f"Error importing use factors: {e}")
    
    return results


def import_heat_rates_from_excel(
    db: Session,
    file_path: str,
    plant_id: int,
    year: int,
    sheet_name: str = "Heat Rate",
) -> Dict:
    """Import heat rates from Excel workbook.
    
    Expected Excel structure:
    - Row with plant name and monthly heat rate values
    - Optional rows for min load heat rate, SUF correction
    
    Args:
        db: Database session
        file_path: Path to Excel file
        plant_id: Plant ID to import for
        year: Year to import
        sheet_name: Name of the sheet with heat rates
        
    Returns:
        Dictionary with import results
    """
    import openpyxl
    from src.models.heat_rate import upsert_heat_rate
    
    results = {
        "file": file_path,
        "sheet": sheet_name,
        "plant_id": plant_id,
        "year": year,
        "months_imported": 0,
        "errors": [],
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            # Try alternative names
            alt_names = ["Heat Rates", "HeatRate", "HR"]
            sheet_found = None
            for alt in alt_names:
                if alt in wb.sheetnames:
                    sheet_found = alt
                    break
            
            if not sheet_found:
                results["errors"].append(f"Sheet '{sheet_name}' not found")
                return results
            sheet_name = sheet_found
        
        ws = wb[sheet_name]
        
        # Find the plant row
        plant_name = "Kyger Creek" if plant_id == 1 else "Clifty Creek"
        baseline_row = None
        suf_row = None
        min_load_row = None
        
        for row_idx in range(1, min(50, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower()
                if plant_name.lower() in cell_lower:
                    if "baseline" in cell_lower or "full load" in cell_lower:
                        baseline_row = row_idx
                    elif "suf" in cell_lower or "startup" in cell_lower:
                        suf_row = row_idx
                    elif "min" in cell_lower:
                        min_load_row = row_idx
                    elif baseline_row is None:  # First match is baseline
                        baseline_row = row_idx
        
        if not baseline_row:
            results["errors"].append(f"Could not find row for {plant_name}")
            return results
        
        # Find month columns
        month_cols = {}
        for col_idx in range(2, 14):
            month_cols[col_idx - 1] = col_idx  # Assume B-M = Jan-Dec
        
        # Get SUF correction (usually a single value, not monthly)
        suf_correction = 0
        if suf_row:
            for col_idx in range(2, 14):
                val = ws.cell(row=suf_row, column=col_idx).value
                if val and isinstance(val, (int, float)):
                    suf_correction = float(val)
                    break
        
        # Extract heat rate values
        for month, col in month_cols.items():
            if month > 12:
                break
            try:
                baseline_value = ws.cell(row=baseline_row, column=col).value
                
                if baseline_value is not None and isinstance(baseline_value, (int, float)):
                    baseline_hr = float(baseline_value)
                    
                    # Get min load heat rate if available
                    min_load_hr = None
                    if min_load_row:
                        min_val = ws.cell(row=min_load_row, column=col).value
                        if min_val and isinstance(min_val, (int, float)):
                            min_load_hr = float(min_val)
                    
                    upsert_heat_rate(
                        db,
                        plant_id=plant_id,
                        year=year,
                        month=month,
                        baseline_heat_rate=baseline_hr,
                        min_load_heat_rate=min_load_hr,
                        suf_correction=suf_correction,
                        prb_blend_adjustment=0,
                        notes=f"Imported from Excel: {Path(file_path).name}",
                    )
                    results["months_imported"] += 1
                    
            except Exception as e:
                results["errors"].append(f"Month {month}: {str(e)}")
        
        wb.close()
        
    except Exception as e:
        results["errors"].append(str(e))
        logger.error(f"Error importing heat rates: {e}")
    
    return results


def import_coal_contracts_from_excel(
    db: Session,
    file_path: str,
    sheet_name: str = "Coal Contracts Annual View",
) -> Dict:
    """Import coal contracts from Excel workbook.
    
    Expected Excel structure:
    - Headers: Contract ID, Supplier, Plant, Start Date, End Date, 
               Annual Tons, BTU/lb, Coal Price, Barge Price, Region
    - One row per contract
    
    Args:
        db: Database session
        file_path: Path to Excel file
        sheet_name: Name of the sheet with contracts
        
    Returns:
        Dictionary with import results
    """
    import openpyxl
    from src.models.coal_contract import CoalContract
    from src.models.plant import Plant
    
    results = {
        "file": file_path,
        "sheet": sheet_name,
        "contracts_imported": 0,
        "contracts_updated": 0,
        "errors": [],
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            # Try alternative names
            alt_names = ["Contracts", "Coal Contracts", "Contract"]
            sheet_found = None
            for alt in alt_names:
                if alt in wb.sheetnames:
                    sheet_found = alt
                    break
            
            if not sheet_found:
                results["errors"].append(f"Sheet '{sheet_name}' not found")
                return results
            sheet_name = sheet_found
        
        ws = wb[sheet_name]
        
        # Find header row and column indices
        headers = {}
        header_row = None
        
        for row_idx in range(1, 10):
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    val_lower = val.lower().strip()
                    if "contract" in val_lower and "id" in val_lower:
                        headers["contract_id"] = col_idx
                        header_row = row_idx
                    elif "supplier" in val_lower:
                        headers["supplier"] = col_idx
                    elif "plant" in val_lower:
                        headers["plant"] = col_idx
                    elif "start" in val_lower and "date" in val_lower:
                        headers["start_date"] = col_idx
                    elif "end" in val_lower and "date" in val_lower:
                        headers["end_date"] = col_idx
                    elif "annual" in val_lower and "ton" in val_lower:
                        headers["annual_tons"] = col_idx
                    elif "btu" in val_lower:
                        headers["btu_per_lb"] = col_idx
                    elif "coal" in val_lower and "price" in val_lower:
                        headers["coal_price"] = col_idx
                    elif "barge" in val_lower:
                        headers["barge_price"] = col_idx
                    elif "region" in val_lower:
                        headers["region"] = col_idx
        
        if not header_row or "contract_id" not in headers:
            results["errors"].append("Could not find contract header row")
            return results
        
        # Get plant mappings
        plant_map = {}
        for plant in db.query(Plant).all():
            plant_map[plant.name.lower()] = plant.id
            if "kyger" in plant.name.lower():
                plant_map["kc"] = plant.id
                plant_map["kyger"] = plant.id
            if "clifty" in plant.name.lower():
                plant_map["cc"] = plant.id
                plant_map["clifty"] = plant.id
        
        # Process data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            try:
                contract_id = ws.cell(row=row_idx, column=headers["contract_id"]).value
                if not contract_id:
                    continue
                
                contract_id = str(contract_id).strip()
                
                # Get supplier
                supplier = ws.cell(row=row_idx, column=headers.get("supplier", 2)).value
                if not supplier:
                    supplier = "Unknown"
                supplier = str(supplier).strip()
                
                # Get plant
                plant_cell = ws.cell(row=row_idx, column=headers.get("plant", 3)).value
                plant_id = None
                if plant_cell:
                    plant_lower = str(plant_cell).lower().strip()
                    plant_id = plant_map.get(plant_lower)
                
                if not plant_id:
                    results["errors"].append(f"Unknown plant for contract {contract_id}")
                    continue
                
                # Get dates
                start_date = ws.cell(row=row_idx, column=headers.get("start_date", 4)).value
                end_date = ws.cell(row=row_idx, column=headers.get("end_date", 5)).value
                
                if not isinstance(start_date, date):
                    start_date = date.today()
                if not isinstance(end_date, date):
                    end_date = date(date.today().year + 1, 12, 31)
                
                # Get numeric values
                annual_tons = ws.cell(row=row_idx, column=headers.get("annual_tons", 6)).value or 0
                btu_per_lb = ws.cell(row=row_idx, column=headers.get("btu_per_lb", 7)).value or 12500
                coal_price = ws.cell(row=row_idx, column=headers.get("coal_price", 8)).value or 0
                barge_price = ws.cell(row=row_idx, column=headers.get("barge_price", 9)).value or 0
                region = ws.cell(row=row_idx, column=headers.get("region", 10)).value or "NAPP"
                
                # Check if contract exists
                existing = db.query(CoalContract).filter(
                    CoalContract.contract_id == contract_id
                ).first()
                
                if existing:
                    # Update existing
                    existing.supplier = supplier
                    existing.plant_id = plant_id
                    existing.start_date = start_date
                    existing.end_date = end_date
                    existing.annual_tons = Decimal(str(annual_tons))
                    existing.btu_per_lb = Decimal(str(btu_per_lb))
                    existing.coal_price_per_ton = Decimal(str(coal_price))
                    existing.barge_price_per_ton = Decimal(str(barge_price))
                    existing.coal_region = str(region)
                    results["contracts_updated"] += 1
                else:
                    # Create new
                    contract = CoalContract(
                        contract_id=contract_id,
                        supplier=supplier,
                        plant_id=plant_id,
                        start_date=start_date,
                        end_date=end_date,
                        is_active=True,
                        annual_tons=Decimal(str(annual_tons)),
                        btu_per_lb=Decimal(str(btu_per_lb)),
                        coal_price_per_ton=Decimal(str(coal_price)),
                        barge_price_per_ton=Decimal(str(barge_price)),
                        coal_region=str(region),
                    )
                    db.add(contract)
                    results["contracts_imported"] += 1
                
            except Exception as e:
                results["errors"].append(f"Row {row_idx}: {str(e)}")
        
        db.commit()
        wb.close()
        
    except Exception as e:
        results["errors"].append(str(e))
        logger.error(f"Error importing coal contracts: {e}")
    
    return results


def import_all_fuel_inputs(
    db: Session,
    file_path: str,
    year: int,
) -> Dict:
    """Import all fuel inputs from a single Excel workbook.
    
    Attempts to find and import:
    - Use factors for both plants
    - Heat rates for both plants
    - Coal contracts
    
    Args:
        db: Database session
        file_path: Path to Excel file
        year: Year to import for
        
    Returns:
        Combined results from all imports
    """
    results = {
        "file": file_path,
        "year": year,
        "use_factors": {},
        "heat_rates": {},
        "contracts": {},
    }
    
    # Import for Kyger Creek (plant_id=1)
    results["use_factors"]["kyger"] = import_use_factors_from_excel(
        db, file_path, plant_id=1, year=year
    )
    results["heat_rates"]["kyger"] = import_heat_rates_from_excel(
        db, file_path, plant_id=1, year=year
    )
    
    # Import for Clifty Creek (plant_id=2)
    results["use_factors"]["clifty"] = import_use_factors_from_excel(
        db, file_path, plant_id=2, year=year
    )
    results["heat_rates"]["clifty"] = import_heat_rates_from_excel(
        db, file_path, plant_id=2, year=year
    )
    
    # Import contracts
    results["contracts"] = import_coal_contracts_from_excel(db, file_path)
    
    # Summary
    results["summary"] = {
        "use_factors_imported": (
            results["use_factors"]["kyger"].get("months_imported", 0) +
            results["use_factors"]["clifty"].get("months_imported", 0)
        ),
        "heat_rates_imported": (
            results["heat_rates"]["kyger"].get("months_imported", 0) +
            results["heat_rates"]["clifty"].get("months_imported", 0)
        ),
        "contracts_imported": results["contracts"].get("contracts_imported", 0),
        "contracts_updated": results["contracts"].get("contracts_updated", 0),
    }
    
    return results

