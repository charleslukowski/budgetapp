"""Sponsor report generation module.

Generates monthly Excel reports for sponsors including:
- Budget vs Actuals
- Variance analysis
- Cost breakdown by department
"""

from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
from calendar import month_name
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import text

from src.reports.variance_report import get_ytd_variance_summary

logger = logging.getLogger(__name__)


# Style definitions
HEADER_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, size=11)
CURRENCY_FORMAT = '"$"#,##0'
CURRENCY_DECIMAL_FORMAT = '"$"#,##0.00'
PERCENT_FORMAT = '0.0%'
NUMBER_FORMAT = '#,##0'

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_sponsor_workbook() -> Workbook:
    """Create a new workbook with styles."""
    wb = Workbook()
    return wb


def add_header_row(ws, row: int, headers: List[str]):
    """Add a header row with styling."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def add_section_row(ws, row: int, label: str, num_cols: int):
    """Add a section header row."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER


def add_data_row(
    ws,
    row: int,
    label: str,
    values: List,
    format_type: str = "currency",
    is_total: bool = False,
):
    """Add a data row with values.

    Args:
        ws: Worksheet
        row: Row number
        label: Row label
        values: List of values (one per column)
        format_type: currency, number, percent
        is_total: Whether this is a total row
    """
    cell = ws.cell(row=row, column=1, value=label)
    if is_total:
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    for col, value in enumerate(values, start=2):
        cell = ws.cell(row=row, column=col, value=value if value else 0)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')

        if is_total:
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)

        if format_type == "currency":
            cell.number_format = CURRENCY_FORMAT
        elif format_type == "currency_decimal":
            cell.number_format = CURRENCY_DECIMAL_FORMAT
        elif format_type == "number":
            cell.number_format = NUMBER_FORMAT
        elif format_type == "percent":
            cell.number_format = PERCENT_FORMAT


def generate_budget_summary_sheet(
    ws,
    db: Session,
    year: int,
    plant_code: str,
):
    """Generate budget summary sheet.

    Args:
        ws: Worksheet
        db: Database session
        year: Year
        plant_code: Plant code (KC or CC)
    """
    ws.title = "Budget Summary"

    # Headers
    headers = ["Department"] + [month_name[m][:3] for m in range(1, 13)] + ["YTD Budget", "YTD Actual", "Variance"]
    add_header_row(ws, 1, headers)

    current_row = 2
    plant_entity = "Kyger" if plant_code == "KC" else "Clifty"

    # Get budget data by department
    try:
        budget_query = text("""
            SELECT
                department,
                SUM(jan) as jan, SUM(feb) as feb, SUM(mar) as mar,
                SUM(apr) as apr, SUM(may) as may, SUM(jun) as jun,
                SUM(jul) as jul, SUM(aug) as aug, SUM(sep) as sep,
                SUM(oct) as oct, SUM(nov) as nov, SUM(dec) as dec,
                SUM(total) as total
            FROM budget_lines
            WHERE budget_year = :year AND budget_entity = :entity
            GROUP BY department
            ORDER BY department
        """)
        result = db.execute(budget_query, {"year": year, "entity": plant_entity})
        budget_rows = result.fetchall()
    except Exception:
        budget_rows = []

    grand_totals = [0] * 12
    grand_ytd_budget = 0
    grand_ytd_actual = 0

    for row in budget_rows:
        dept = row[0] or "UNKNOWN"
        monthly_values = [float(row[i]) if row[i] else 0 for i in range(1, 13)]
        ytd_budget = sum(monthly_values)

        # Get actuals for this department (placeholder - would query transaction_budget_groups)
        ytd_actual = 0  # Would be populated from actuals
        variance = ytd_budget - ytd_actual

        add_data_row(ws, current_row, dept, monthly_values + [ytd_budget, ytd_actual, variance], "currency")
        current_row += 1

        for i, val in enumerate(monthly_values):
            grand_totals[i] += val
        grand_ytd_budget += ytd_budget
        grand_ytd_actual += ytd_actual

    # Add grand total row
    grand_variance = grand_ytd_budget - grand_ytd_actual
    add_data_row(ws, current_row, "TOTAL", grand_totals + [grand_ytd_budget, grand_ytd_actual, grand_variance], "currency", is_total=True)

    # Set column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 17):
        ws.column_dimensions[get_column_letter(col)].width = 12


def generate_sponsor_report(
    db: Session,
    year: int,
    output_path: Path,
    plant_code: str = "KC",
) -> Path:
    """Generate complete sponsor report workbook.

    Args:
        db: Database session
        year: Year for report
        output_path: Path to save the Excel file
        plant_code: Plant code (KC, CC, or None for system)

    Returns:
        Path to generated file
    """
    logger.info(f"Generating sponsor report for {year}")

    # Create workbook
    wb = create_sponsor_workbook()

    plant_name = "Kyger Creek" if plant_code == "KC" else "Clifty Creek" if plant_code == "CC" else "System"

    # Generate summary sheet
    ws_summary = wb.active
    generate_budget_summary_sheet(ws_summary, db, year, plant_code)

    # Add cover sheet
    ws_cover = wb.create_sheet("Cover", 0)
    ws_cover["A1"] = f"OVEC Budget Report"
    ws_cover["A1"].font = Font(bold=True, size=16)
    ws_cover["A3"] = f"Plant: {plant_name}"
    ws_cover["A4"] = f"Year: {year}"
    ws_cover["A5"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_cover.column_dimensions['A'].width = 40

    # Save workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    logger.info(f"Sponsor report saved to {output_path}")
    return output_path


def generate_all_sponsor_reports(
    db: Session,
    year: int,
    output_dir: Path,
) -> List[Path]:
    """Generate sponsor reports for all plants.

    Args:
        db: Database session
        year: Year for reports
        output_dir: Directory to save reports

    Returns:
        List of generated file paths
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    reports = []

    # Kyger report
    kyger_path = output_dir / f"OVEC_Budget_{year}_Kyger.xlsx"
    reports.append(generate_sponsor_report(db, year, kyger_path, plant_code="KC"))

    # Clifty report
    clifty_path = output_dir / f"OVEC_Budget_{year}_Clifty.xlsx"
    reports.append(generate_sponsor_report(db, year, clifty_path, plant_code="CC"))

    return reports
