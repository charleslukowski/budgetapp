"""Import coal contracts from Excel budget file."""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from datetime import date
from decimal import Decimal
from src.db.postgres import get_session
from src.models.coal_contract import CoalContract, CoalDelivery
from src.models.plant import Plant


def extract_contracts_from_excel(excel_path: str) -> list:
    """Extract coal contract data from Coal Contracts Annual View sheet."""
    
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb['Coal Contracts Annual View']
    
    contracts = []
    current_contract = None
    
    # Parse the sheet - contracts are in blocks starting with contract name
    for row in range(1, 200):
        col_a = ws.cell(row=row, column=1).value
        
        if col_a is None:
            continue
        
        col_a_str = str(col_a).strip()
        
        # Skip headers and empty rows
        if col_a_str in ['', 'CC Contracted Coal Summary', 'KC Contracted Coal Summary']:
            continue
        
        # Detect contract name rows (don't contain common metric names)
        metric_names = ['Thousand Tons', 'Contracted Btu', 'Contracted SO2', 'Ash', 
                        'Coal $/ton', 'Barge $/ton', 'Total Delivered Cost']
        
        if col_a_str not in metric_names and not any(m in col_a_str for m in metric_names):
            # This is a contract name row
            if current_contract and current_contract.get('contract_name'):
                contracts.append(current_contract)
            
            current_contract = {
                'contract_name': col_a_str,
                'years': {}
            }
            continue
        
        if current_contract is None:
            continue
        
        # Get values for years 2025, 2026, etc. (columns B=2, C=3, D=4...)
        # Based on header row, column B is usually 2025
        for year_offset, col_idx in enumerate(range(2, 8)):
            year = 2025 + year_offset
            
            val = ws.cell(row=row, column=col_idx).value
            if val is None or val == 0:
                continue
            
            if year not in current_contract['years']:
                current_contract['years'][year] = {}
            
            try:
                val = float(val)
            except (ValueError, TypeError):
                continue
            
            if 'Thousand Tons' in col_a_str:
                current_contract['years'][year]['tons_thousands'] = val
            elif 'Contracted Btu' in col_a_str or col_a_str == 'Contracted Btu':
                current_contract['years'][year]['btu_per_lb'] = val
            elif 'Contracted SO2' in col_a_str or col_a_str == 'Contracted SO2':
                current_contract['years'][year]['so2_lb_per_mmbtu'] = val
            elif col_a_str == 'Ash':
                current_contract['years'][year]['ash_pct'] = val
            elif 'Coal $/ton' in col_a_str or col_a_str == 'Coal $/ton':
                current_contract['years'][year]['coal_price'] = val
            elif 'Barge $/ton' in col_a_str or col_a_str == 'Barge $/ton':
                current_contract['years'][year]['barge_price'] = val
    
    # Don't forget the last contract
    if current_contract and current_contract.get('contract_name'):
        contracts.append(current_contract)
    
    wb.close()
    
    # Filter out empty contracts
    contracts = [c for c in contracts if c.get('years')]
    
    return contracts


def determine_plant_and_region(contract_name: str) -> tuple:
    """Determine plant assignment and coal region from contract name."""
    
    name_lower = contract_name.lower()
    
    # Determine coal region based on supplier name
    region = 'NAPP'  # Default to Northern Appalachian
    if 'knight hawk' in name_lower or 'kh' in name_lower:
        region = 'ILB'  # Illinois Basin
    elif 'foresight' in name_lower:
        region = 'ILB'
    elif 'arch' in name_lower and 'prb' in name_lower:
        region = 'PRB'
    
    # Contract IDs starting with 31- are typically Clifty Creek (CC)
    # Contract IDs starting with 30- are typically Kyger Creek (KC)
    # Default to both plants (we'll need to split)
    if '31-' in contract_name:
        plant_prefix = 'CC'
    elif '30-' in contract_name:
        plant_prefix = 'KC'
    else:
        plant_prefix = 'BOTH'
    
    return plant_prefix, region


def import_contracts(excel_path: str, year: int = 2025, dry_run: bool = False):
    """Import contracts from Excel into database."""
    
    print(f"Extracting contracts from: {excel_path}")
    contracts_data = extract_contracts_from_excel(excel_path)
    
    print(f"\nFound {len(contracts_data)} contracts:")
    for c in contracts_data:
        print(f"  - {c['contract_name']}: {list(c['years'].keys())}")
    
    if dry_run:
        print("\n[DRY RUN] Would import the following contracts:")
        for c in contracts_data:
            if year in c['years']:
                data = c['years'][year]
                plant_prefix, region = determine_plant_and_region(c['contract_name'])
                print(f"\n  Contract: {c['contract_name']}")
                print(f"    Plant: {plant_prefix}")
                print(f"    Region: {region}")
                print(f"    Tons: {data.get('tons_thousands', 0) * 1000:,.0f}")
                print(f"    BTU/lb: {data.get('btu_per_lb', 0):,.0f}")
                print(f"    SO2: {data.get('so2_lb_per_mmbtu', 0):.2f} lb/MMBtu")
                print(f"    Ash: {data.get('ash_pct', 0):.1f}%")
                print(f"    Coal: ${data.get('coal_price', 0):.2f}/ton")
                print(f"    Barge: ${data.get('barge_price', 0):.2f}/ton")
                print(f"    Delivered: ${data.get('coal_price', 0) + data.get('barge_price', 0):.2f}/ton")
        return
    
    with get_session() as db:
        # Get plant IDs
        plants = {p.short_name: p.id for p in db.query(Plant).all()}
        print(f"\nPlants: {plants}")
        
        if not plants:
            print("ERROR: No plants found in database. Run migrations first.")
            return
        
        # Determine plant IDs
        kyger_id = plants.get('KC') or plants.get('Kyger') or 1
        clifty_id = plants.get('CC') or plants.get('Clifty') or 2
        
        imported = 0
        for c in contracts_data:
            if year not in c['years']:
                continue
            
            data = c['years'][year]
            plant_prefix, region = determine_plant_and_region(c['contract_name'])
            
            # Create contracts for BOTH plants, splitting tonnage 50/50
            for plant_id, plant_code in [(kyger_id, 'KC'), (clifty_id, 'CC')]:
                # Contract ID includes plant prefix
                contract_id = f"{plant_code}-{c['contract_name']}"
                
                # Check if contract already exists
                existing = db.query(CoalContract).filter(
                    CoalContract.contract_id == contract_id
                ).first()
                
                # Split tonnage between plants (50/50)
                tonnage = data.get('tons_thousands', 0) * 1000 / 2
                
                if existing:
                    print(f"  Updating: {contract_id}")
                    existing.annual_tons = Decimal(str(tonnage))
                    existing.btu_per_lb = Decimal(str(data.get('btu_per_lb', 11500)))
                    existing.so2_lb_per_mmbtu = Decimal(str(data.get('so2_lb_per_mmbtu', 5.0)))
                    existing.ash_pct = Decimal(str(data.get('ash_pct', 9) / 100))
                    existing.coal_price_per_ton = Decimal(str(data.get('coal_price', 55)))
                    existing.barge_price_per_ton = Decimal(str(data.get('barge_price', 6)))
                else:
                    print(f"  Creating: {contract_id} @ ${data.get('coal_price', 0):.2f}/ton")
                    contract = CoalContract(
                        contract_id=contract_id,
                        supplier=c['contract_name'].split()[0],
                        plant_id=plant_id,
                        start_date=date(year, 1, 1),
                        end_date=date(year, 12, 31),
                        is_active=True,
                        annual_tons=Decimal(str(tonnage)),
                        btu_per_lb=Decimal(str(data.get('btu_per_lb', 11500))),
                        so2_lb_per_mmbtu=Decimal(str(data.get('so2_lb_per_mmbtu', 5.0))),
                        ash_pct=Decimal(str(data.get('ash_pct', 9) / 100)),
                        coal_price_per_ton=Decimal(str(data.get('coal_price', 55))),
                        barge_price_per_ton=Decimal(str(data.get('barge_price', 6))),
                        coal_region=region,
                    )
                    db.add(contract)
                
                imported += 1
        
        db.commit()
        print(f"\nImported/updated {imported} contracts for {year}")
        
        # Create monthly deliveries for each contract
        print("\nCreating monthly delivery schedule...")
        contracts = db.query(CoalContract).filter(CoalContract.is_active == True).all()
        
        for contract in contracts:
            monthly_tons = float(contract.annual_tons) / 12
            
            for month in range(1, 13):
                period_yyyymm = f"{year}{month:02d}"
                
                # Check if delivery exists
                existing = db.query(CoalDelivery).filter(
                    CoalDelivery.contract_id == contract.id,
                    CoalDelivery.period_yyyymm == period_yyyymm
                ).first()
                
                if not existing:
                    delivery = CoalDelivery(
                        contract_id=contract.id,
                        period_yyyymm=period_yyyymm,
                        scheduled_tons=Decimal(str(monthly_tons)),
                        actual_tons=Decimal(str(monthly_tons)),  # Use scheduled as actual for budget
                    )
                    db.add(delivery)
        
        db.commit()
        print("Monthly deliveries created.")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Import coal contracts from Excel")
    parser.add_argument("--year", type=int, default=2025, help="Year to import")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be imported without making changes")
    parser.add_argument("--excel", type=str, 
                        default="docs/source_documents/OVEC IKEC Energy Budget- 2026 Year-End BOD Update- Final- (12-8-25).xlsm",
                        help="Path to Excel file")
    
    args = parser.parse_args()
    
    import_contracts(args.excel, args.year, args.dry_run)
