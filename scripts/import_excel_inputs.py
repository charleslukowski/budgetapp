"""Import heat rate, use factor, and coal data from OVEC Excel file to database.

This script extracts the correct input values from the Excel file and imports
them into the database tables to fix model discrepancies:

1. Heat rates with SUF correction from KC/CC Coal Burn sheets
2. Use factors from Use Factor Input sheet
3. Coal pricing (future enhancement)

Run with: python scripts/import_excel_inputs.py --year 2025
"""

import sys
import os
from pathlib import Path
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import logging
import argparse

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


def extract_heat_rates_from_coal_burn(ws, plant: str, year: int, debug: bool = True) -> Dict[int, Dict[str, float]]:
    """Extract heat rate data from KC/CC Coal Burn sheet.
    
    The Coal Burn sheet structure:
    - Row 3: Years (each month is a separate column for the same year)
    - Row 4: Months (dates like 10/1/2025, 11/1/2025, etc.)
    - Row 10: "Net Delivered Heat Rate" or "Production HR, Net" - this is baseline
    - Row 11: "SUF Correction" - additional adjustment
    - Row 12: "Adjusted Heat Rate" - baseline + SUF
    
    Args:
        ws: Worksheet object
        plant: Plant name (for logging)
        year: Target year
        
    Returns:
        Dict mapping month (1-12) to heat rate values:
        {month: {'baseline': float, 'suf_correction': float, 'adjusted': float}}
    """
    result = {}
    
    # Debug: print rows to understand structure
    if debug:
        print(f"\n  DEBUG: Analyzing {plant} Coal Burn sheet structure:")
        for row in range(1, 25):
            row_vals = []
            label = ws.cell(row=row, column=1).value
            if label is None:
                label = "--"
            for col in range(1, 8):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    if isinstance(val, datetime):
                        row_vals.append(f"DT:{val.month}/{val.year}")
                    elif isinstance(val, (int, float)):
                        row_vals.append(f"{val:,.0f}" if abs(val) > 100 else f"{val:.4f}")
                    else:
                        row_vals.append(str(val)[:15])
                else:
                    row_vals.append("--")
            # Only print rows with interesting content
            label_lower = str(label).lower() if label else ""
            if any(x in label_lower for x in ['heat', 'suf', 'produc', 'net', 'burn', 'year', 'day']):
                print(f"    Row {row:2d}: {' | '.join(row_vals)}")
    
    # Find all columns for the target year
    year_cols = {}  # col_idx -> month
    
    # Check row 1 for dates (as datetime objects) - this is the primary date row
    for col in range(2, 30):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            if isinstance(val, datetime):
                if val.year == year:
                    year_cols[col] = val.month
            elif hasattr(val, 'year'):  # date object
                if val.year == year:
                    year_cols[col] = val.month
    
    # If no dates in row 1, try using row 3 (Year) to identify target year columns
    if not year_cols:
        for col in range(2, 30):
            year_val = ws.cell(row=3, column=col).value
            if year_val is not None:
                try:
                    if int(year_val) == year:
                        # Get month from row 1
                        month_val = ws.cell(row=1, column=col).value
                        if isinstance(month_val, datetime):
                            year_cols[col] = month_val.month
                        elif hasattr(month_val, 'month'):
                            year_cols[col] = month_val.month
                except (ValueError, TypeError):
                    pass
    
    if not year_cols:
        logger.warning(f"Could not find year {year} in {plant} Coal Burn sheet")
        return result
    
    if debug:
        print(f"    Found {len(year_cols)} columns for {year}: months {sorted(year_cols.values())}")
    
    logger.info(f"  Found {len(year_cols)} months for {year}: {sorted(year_cols.values())}")
    
    # Find heat rate rows - the Excel has specific row labels
    # Row 21: "Net Plant HR (including FGD)" - this is total heat rate after aux power
    # Row 22: "Production HR" - production heat rate
    # Row 14: "SUF Correction (LSL Ops HR Penalty)" - this contains the adjusted HR values
    
    net_plant_hr_row = None      # Row 21 - total including FGD aux
    production_hr_row = None     # Row 22 - production HR
    suf_row = None               # Row 14 - labeled as SUF but contains adjusted values
    
    for row in range(5, 50):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if "net plant hr" in label_str or "net plant heat rate" in label_str:
            net_plant_hr_row = row
            logger.info(f"  Found Net Plant HR at row {row}: {label}")
        elif "production hr" in label_str and "net" not in label_str:
            production_hr_row = row
            logger.info(f"  Found Production HR at row {row}: {label}")
        elif "suf" in label_str:
            suf_row = row
            logger.info(f"  Found SUF row at row {row}: {label}")
    
    # Determine which row to use as the "effective" heat rate
    # We'll use Production HR (row 22) as the effective heat rate for cost calculations
    effective_hr_row = production_hr_row if production_hr_row else suf_row
    
    if not effective_hr_row:
        logger.warning(f"  Could not find effective heat rate row for {plant}")
        return result
    
    # Our model default baseline is 9,850 BTU/kWh
    MODEL_BASELINE = 9850.0
    
    # Extract values for each month
    for col, month in year_cols.items():
        month_data = {
            'baseline': MODEL_BASELINE,  # Our model default
            'suf_correction': 0.0,
            'adjusted': MODEL_BASELINE,
        }
        
        # Get the effective heat rate from Excel
        if effective_hr_row:
            val = ws.cell(row=effective_hr_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                effective_hr = float(val)
                month_data['adjusted'] = effective_hr
                # SUF correction is the difference from our model baseline
                month_data['suf_correction'] = effective_hr - MODEL_BASELINE
        
        # Optionally get the Net Plant HR for reference
        if net_plant_hr_row:
            val = ws.cell(row=net_plant_hr_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                # Store this as the "full" heat rate
                month_data['net_plant_hr'] = float(val)
        
        result[month] = month_data
    
    return result


def extract_use_factors(ws, year: int, debug: bool = True) -> Dict[str, Dict[int, float]]:
    """Extract use factors from Use Factor Input sheet.
    
    The Use Factor Input sheet structure:
    - Row 1-2: Headers
    - Row 3: Clifty Creek use factors (base)
    - Row 4: Kyger Creek use factors (base)
    - Columns B-M or dates: Jan-Dec values
    
    Args:
        ws: Worksheet object
        year: Target year
        
    Returns:
        Dict with plant keys ('kyger', 'clifty') mapping to
        {month (1-12): use_factor (0-1)}
    """
    result = {
        'kyger': {},
        'clifty': {},
    }
    
    # Debug: print structure
    if debug:
        print(f"\n  DEBUG: Analyzing Use Factor Input sheet structure:")
        for row in range(1, 6):
            row_vals = []
            for col in range(1, 10):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    if isinstance(val, datetime):
                        row_vals.append(f"DT:{val.month}/{val.year}")
                    elif isinstance(val, (int, float)):
                        row_vals.append(f"{val:.4f}" if val < 1.5 else f"{val:.0f}")
                    else:
                        row_vals.append(str(val)[:15])
                else:
                    row_vals.append("--")
            print(f"    Row {row}: {' | '.join(row_vals)}")
    
    # Find plant rows
    kyger_row = None
    clifty_row = None
    
    for row in range(1, 20):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if 'kyger' in label_str:
            # Use first kyger row (base, not ozone)
            if kyger_row is None:
                kyger_row = row
                logger.info(f"  Found Kyger Creek at row {row}: {label}")
        elif 'clifty' in label_str:
            if clifty_row is None:
                clifty_row = row
                logger.info(f"  Found Clifty Creek at row {row}: {label}")
    
    # Find month columns - check for dates in header rows (1, 2)
    month_cols = {}  # col_idx -> month
    
    for check_row in [1, 2]:
        for col in range(2, 20):
            val = ws.cell(row=check_row, column=col).value
            if val is None:
                continue
            
            if isinstance(val, datetime):
                if val.year == year:
                    month_cols[col] = val.month
            elif hasattr(val, 'year'):  # date object
                if val.year == year:
                    month_cols[col] = val.month
            elif isinstance(val, str):
                # Check for year in string format "Oct-25" or "10/1/2025"
                if str(year) in val or str(year)[-2:] in val:
                    month_map = {
                        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                        'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
                    }
                    val_lower = val.lower().strip()
                    for m_name, m_num in month_map.items():
                        if m_name in val_lower:
                            month_cols[col] = m_num
                            break
    
    # If we found date columns, use only those for target year
    if debug and month_cols:
        print(f"    Found year {year} columns: {month_cols}")
    
    # If no month headers found, assume columns B-M are Jan-Dec
    if not month_cols:
        logger.info("  Using default column mapping B-M for Jan-Dec")
        for i, col in enumerate(range(2, 14), start=1):
            month_cols[col] = i
    
    logger.info(f"  Found {len(month_cols)} month columns")
    
    # Extract Kyger values
    if kyger_row:
        for col, month in month_cols.items():
            val = ws.cell(row=kyger_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                # Convert percentage to decimal if needed
                use_factor = float(val)
                if use_factor > 1:  # Percentage format
                    use_factor = use_factor / 100
                result['kyger'][month] = use_factor
        logger.info(f"  Kyger: {len(result['kyger'])} months extracted")
    
    # Extract Clifty values
    if clifty_row:
        for col, month in month_cols.items():
            val = ws.cell(row=clifty_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                use_factor = float(val)
                if use_factor > 1:
                    use_factor = use_factor / 100
                result['clifty'][month] = use_factor
        logger.info(f"  Clifty: {len(result['clifty'])} months extracted")
    
    return result


def import_heat_rates_to_db(db, plant_id: int, year: int, heat_rates: Dict[int, Dict[str, float]]) -> int:
    """Import heat rate values to the database.
    
    Args:
        db: Database session
        plant_id: Plant ID (1=Kyger, 2=Clifty)
        year: Year
        heat_rates: Dict mapping month -> {baseline, suf_correction, adjusted}
        
    Returns:
        Number of records upserted
    """
    from src.models.heat_rate import upsert_heat_rate
    
    count = 0
    for month, hr_data in heat_rates.items():
        upsert_heat_rate(
            db,
            plant_id=plant_id,
            year=year,
            month=month,
            baseline_heat_rate=hr_data['baseline'],
            min_load_heat_rate=None,  # Could extract if available
            suf_correction=hr_data['suf_correction'],
            prb_blend_adjustment=0,
            notes=f"Imported from OVEC Excel {datetime.now().strftime('%Y-%m-%d')}",
            updated_by="import_excel_inputs.py",
        )
        count += 1
        logger.debug(f"  Upserted heat rate for month {month}: "
                    f"baseline={hr_data['baseline']:.0f}, suf={hr_data['suf_correction']:.0f}")
    
    return count


def import_use_factors_to_db(db, plant_id: int, year: int, use_factors: Dict[int, float]) -> int:
    """Import use factor values to the database.
    
    Args:
        db: Database session
        plant_id: Plant ID (1=Kyger, 2=Clifty)
        year: Year
        use_factors: Dict mapping month (1-12) -> use factor (0-1)
        
    Returns:
        Number of records upserted
    """
    from src.models.use_factor import upsert_use_factor
    
    count = 0
    for month, uf in use_factors.items():
        upsert_use_factor(
            db,
            plant_id=plant_id,
            year=year,
            month=month,
            use_factor_base=uf,
            use_factor_ozone_non_scr=0.0,  # Will need separate extraction for ozone
            notes=f"Imported from OVEC Excel {datetime.now().strftime('%Y-%m-%d')}",
            updated_by="import_excel_inputs.py",
        )
        count += 1
        logger.debug(f"  Upserted use factor for month {month}: {uf:.4f}")
    
    return count


def extract_outages(ws, plant: str, year: int, debug: bool = True) -> Dict[int, Dict[int, float]]:
    """Extract unit outage days from KC/CC Reliability sheet.
    
    Args:
        ws: Worksheet object
        plant: Plant name (for logging)
        year: Target year
        
    Returns:
        Dict mapping month (1-12) to {unit_num: outage_days}
    """
    result = {}
    
    # Debug: print structure
    if debug:
        print(f"\n  DEBUG: Analyzing {plant} Reliability sheet structure:")
        for row in range(1, 15):
            row_vals = []
            for col in range(1, 8):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    if isinstance(val, datetime):
                        row_vals.append(f"DT:{val.month}/{val.year}")
                    elif isinstance(val, (int, float)):
                        row_vals.append(f"{val:.1f}")
                    else:
                        row_vals.append(str(val)[:15])
                else:
                    row_vals.append("--")
            print(f"    Row {row}: {' | '.join(row_vals)}")
    
    # Find columns for target year (row 1 has dates, row 3 has year)
    year_cols = {}  # col_idx -> month
    for col in range(2, 30):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            if isinstance(val, datetime) and val.year == year:
                year_cols[col] = val.month
            elif hasattr(val, 'year') and val.year == year:
                year_cols[col] = val.month
    
    if not year_cols:
        logger.warning(f"  Could not find year {year} in {plant} Reliability sheet")
        return result
    
    if debug:
        print(f"    Found {len(year_cols)} columns for {year}: months {sorted(year_cols.values())}")
    
    # Find unit rows for PLANNED outages (rows 5-9 for KC, 5-10 for CC)
    # We need to find the section header "Planned Outage" first, then look for units below it
    planned_section_start = None
    for row in range(3, 15):
        label = ws.cell(row=row, column=1).value
        if label and isinstance(label, str):
            if "planned outage" in label.lower():
                planned_section_start = row
                break
    
    if not planned_section_start:
        logger.warning(f"  Could not find 'Planned Outage' section in {plant} Reliability sheet")
        return result
    
    # Find unit rows in the planned outage section (rows immediately after header)
    unit_rows = {}
    for row in range(planned_section_start + 1, planned_section_start + 10):
        label = ws.cell(row=row, column=1).value
        if label and isinstance(label, str):
            label_lower = label.lower()
            # Stop if we hit "Total" or another section
            if "total" in label_lower or "forced" in label_lower:
                break
            if "unit 1" in label_lower:
                unit_rows[1] = row
            elif "unit 2" in label_lower:
                unit_rows[2] = row
            elif "unit 3" in label_lower:
                unit_rows[3] = row
            elif "unit 4" in label_lower:
                unit_rows[4] = row
            elif "unit 5" in label_lower:
                unit_rows[5] = row
            elif "unit 6" in label_lower:
                unit_rows[6] = row
    
    if debug:
        print(f"    Found planned outage unit rows: {unit_rows}")
    
    # Extract outage days for each month
    for col, month in year_cols.items():
        month_outages = {}
        for unit_num, row in unit_rows.items():
            val = ws.cell(row=row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                outage_days = float(val)
                if outage_days > 0:
                    month_outages[unit_num] = outage_days
        
        if month_outages:
            result[month] = month_outages
            if debug:
                print(f"    Month {month} outages: {month_outages}")
    
    return result


def import_outages_to_db(db, plant_id: int, year: int, outages: Dict[int, Dict[int, float]]) -> int:
    """Import unit outage days to the database.
    
    Args:
        db: Database session
        plant_id: Plant ID (1=Kyger, 2=Clifty)
        year: Year
        outages: Dict mapping month -> {unit_num: outage_days}
        
    Returns:
        Number of records upserted
    """
    from src.models.unit_outage import upsert_unit_outage
    from decimal import Decimal
    
    count = 0
    for month, unit_outages in outages.items():
        for unit_num, outage_days in unit_outages.items():
            upsert_unit_outage(
                db,
                plant_id=plant_id,
                unit_number=unit_num,
                year=year,
                month=month,
                planned_outage_days=Decimal(str(outage_days)),
                notes=f"Imported from OVEC Excel {datetime.now().strftime('%Y-%m-%d')}",
                updated_by="import_excel_inputs.py",
            )
            count += 1
            logger.info(f"  Unit {unit_num} month {month}: {outage_days} days planned outage")
    
    return count


def main():
    """Main import function."""
    import openpyxl
    
    parser = argparse.ArgumentParser(description="Import Excel inputs to database")
    parser.add_argument("--year", type=int, default=2025, help="Year to import")
    parser.add_argument("--excel", type=str, 
                       default="docs/source_documents/OVEC IKEC Energy Budget- 2026 Year-End BOD Update- Final- (12-8-25).xlsm",
                       help="Path to Excel file")
    parser.add_argument("--dry-run", action="store_true", 
                       help="Extract data but don't write to database")
    parser.add_argument("--heat-rates-only", action="store_true",
                       help="Only import heat rates")
    parser.add_argument("--use-factors-only", action="store_true",
                       help="Only import use factors")
    parser.add_argument("--outages-only", action="store_true",
                       help="Only import unit outages")
    args = parser.parse_args()
    
    print("=" * 80)
    print("IMPORT EXCEL INPUTS TO DATABASE")
    print(f"Year: {args.year}")
    print(f"Excel: {args.excel}")
    print(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE IMPORT'}")
    print("=" * 80)
    
    # Open Excel file
    excel_path = Path(args.excel)
    if not excel_path.exists():
        excel_path = Path(__file__).parent.parent / args.excel
    
    if not excel_path.exists():
        print(f"\nERROR: Excel file not found: {args.excel}")
        return 1
    
    print(f"\nOpening Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return 1
    
    sheets = wb.sheetnames
    print(f"Found {len(sheets)} sheets")
    
    # Track import results
    results = {
        'heat_rates': {'kyger': 0, 'clifty': 0},
        'use_factors': {'kyger': 0, 'clifty': 0},
        'outages': {'kyger': 0, 'clifty': 0},
    }
    
    extracted_data = {
        'heat_rates': {'kyger': {}, 'clifty': {}},
        'use_factors': {'kyger': {}, 'clifty': {}},
        'outages': {'kyger': {}, 'clifty': {}},
    }
    
    # =========================================================================
    # EXTRACT HEAT RATES
    # =========================================================================
    if not args.use_factors_only:
        print("\n" + "=" * 80)
        print("EXTRACTING HEAT RATES")
        print("=" * 80)
        
        # Kyger Creek heat rates
        if 'KC Coal Burn' in sheets:
            print("\nProcessing KC Coal Burn sheet...")
            ws = wb['KC Coal Burn']
            kc_hr = extract_heat_rates_from_coal_burn(ws, "Kyger", args.year)
            extracted_data['heat_rates']['kyger'] = kc_hr
            
            if kc_hr:
                print(f"\n  Kyger Creek Heat Rates for {args.year}:")
                print(f"  {'Month':<10} {'Baseline':>12} {'SUF Corr':>12} {'Adjusted':>12}")
                print(f"  {'-' * 46}")
                for month in sorted(kc_hr.keys()):
                    data = kc_hr[month]
                    print(f"  {month:<10} {data['baseline']:>12,.0f} "
                          f"{data['suf_correction']:>12,.0f} {data['adjusted']:>12,.0f}")
        else:
            print("WARNING: KC Coal Burn sheet not found")
        
        # Clifty Creek heat rates
        if 'CC Coal Burn' in sheets:
            print("\nProcessing CC Coal Burn sheet...")
            ws = wb['CC Coal Burn']
            cc_hr = extract_heat_rates_from_coal_burn(ws, "Clifty", args.year)
            extracted_data['heat_rates']['clifty'] = cc_hr
            
            if cc_hr:
                print(f"\n  Clifty Creek Heat Rates for {args.year}:")
                print(f"  {'Month':<10} {'Baseline':>12} {'SUF Corr':>12} {'Adjusted':>12}")
                print(f"  {'-' * 46}")
                for month in sorted(cc_hr.keys()):
                    data = cc_hr[month]
                    print(f"  {month:<10} {data['baseline']:>12,.0f} "
                          f"{data['suf_correction']:>12,.0f} {data['adjusted']:>12,.0f}")
        else:
            print("WARNING: CC Coal Burn sheet not found")
    
    # =========================================================================
    # EXTRACT USE FACTORS
    # =========================================================================
    if not args.heat_rates_only:
        print("\n" + "=" * 80)
        print("EXTRACTING USE FACTORS")
        print("=" * 80)
        
        if 'Use Factor Input' in sheets:
            print("\nProcessing Use Factor Input sheet...")
            ws = wb['Use Factor Input']
            use_factors = extract_use_factors(ws, args.year)
            extracted_data['use_factors'] = use_factors
            
            # Print extracted use factors
            for plant in ['kyger', 'clifty']:
                if use_factors[plant]:
                    print(f"\n  {plant.title()} Creek Use Factors for {args.year}:")
                    print(f"  {'Month':<10} {'Use Factor':>12}")
                    print(f"  {'-' * 22}")
                    for month in sorted(use_factors[plant].keys()):
                        uf = use_factors[plant][month]
                        print(f"  {month:<10} {uf:>12.4f}")
        else:
            print("WARNING: Use Factor Input sheet not found")
    
    # =========================================================================
    # EXTRACT UNIT OUTAGES
    # =========================================================================
    if not args.heat_rates_only and not args.use_factors_only:
        print("\n" + "=" * 80)
        print("EXTRACTING UNIT OUTAGES")
        print("=" * 80)
        
        # Kyger Creek outages
        if 'KC Reliability' in sheets:
            print("\nProcessing KC Reliability sheet...")
            ws = wb['KC Reliability']
            kc_outages = extract_outages(ws, "Kyger", args.year)
            extracted_data['outages']['kyger'] = kc_outages
            
            if kc_outages:
                print(f"\n  Kyger Creek Planned Outages for {args.year}:")
                for month in sorted(kc_outages.keys()):
                    units = kc_outages[month]
                    total_days = sum(units.values())
                    print(f"  Month {month}: {total_days} total unit-days ({units})")
        else:
            print("WARNING: KC Reliability sheet not found")
        
        # Clifty Creek outages
        if 'CC Reliability' in sheets:
            print("\nProcessing CC Reliability sheet...")
            ws = wb['CC Reliability']
            cc_outages = extract_outages(ws, "Clifty", args.year)
            extracted_data['outages']['clifty'] = cc_outages
            
            if cc_outages:
                print(f"\n  Clifty Creek Planned Outages for {args.year}:")
                for month in sorted(cc_outages.keys()):
                    units = cc_outages[month]
                    total_days = sum(units.values())
                    print(f"  Month {month}: {total_days} total unit-days ({units})")
        else:
            print("WARNING: CC Reliability sheet not found")
    
    wb.close()
    
    # =========================================================================
    # IMPORT TO DATABASE
    # =========================================================================
    if args.dry_run:
        print("\n" + "=" * 80)
        print("DRY RUN - No database changes made")
        print("=" * 80)
        print("\nRe-run without --dry-run to import data to database")
        return 0
    
    print("\n" + "=" * 80)
    print("IMPORTING TO DATABASE")
    print("=" * 80)
    
    from src.db.postgres import get_session
    
    with get_session() as db:
        # Import heat rates
        if not args.use_factors_only:
            print("\nImporting heat rates...")
            
            if extracted_data['heat_rates']['kyger']:
                count = import_heat_rates_to_db(db, 1, args.year, 
                                                extracted_data['heat_rates']['kyger'])
                results['heat_rates']['kyger'] = count
                print(f"  Kyger Creek: {count} months imported")
            
            if extracted_data['heat_rates']['clifty']:
                count = import_heat_rates_to_db(db, 2, args.year,
                                                extracted_data['heat_rates']['clifty'])
                results['heat_rates']['clifty'] = count
                print(f"  Clifty Creek: {count} months imported")
        
        # Import use factors
        if not args.heat_rates_only:
            print("\nImporting use factors...")
            
            if extracted_data['use_factors']['kyger']:
                count = import_use_factors_to_db(db, 1, args.year,
                                                 extracted_data['use_factors']['kyger'])
                results['use_factors']['kyger'] = count
                print(f"  Kyger Creek: {count} months imported")
            
            if extracted_data['use_factors']['clifty']:
                count = import_use_factors_to_db(db, 2, args.year,
                                                 extracted_data['use_factors']['clifty'])
                results['use_factors']['clifty'] = count
                print(f"  Clifty Creek: {count} months imported")
        
        # Import unit outages
        if not args.heat_rates_only and not args.use_factors_only:
            print("\nImporting unit outages...")
            
            if extracted_data['outages']['kyger']:
                count = import_outages_to_db(db, 1, args.year,
                                            extracted_data['outages']['kyger'])
                results['outages']['kyger'] = count
                print(f"  Kyger Creek: {count} unit-months imported")
            
            if extracted_data['outages']['clifty']:
                count = import_outages_to_db(db, 2, args.year,
                                            extracted_data['outages']['clifty'])
                results['outages']['clifty'] = count
                print(f"  Clifty Creek: {count} unit-months imported")
        
        db.commit()
    
    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("IMPORT COMPLETE")
    print("=" * 80)
    
    print(f"\nHeat Rates Imported:")
    print(f"  Kyger Creek: {results['heat_rates']['kyger']} months")
    print(f"  Clifty Creek: {results['heat_rates']['clifty']} months")
    
    print(f"\nUse Factors Imported:")
    print(f"  Kyger Creek: {results['use_factors']['kyger']} months")
    print(f"  Clifty Creek: {results['use_factors']['clifty']} months")
    
    print(f"\nUnit Outages Imported:")
    print(f"  Kyger Creek: {results['outages']['kyger']} unit-months")
    print(f"  Clifty Creek: {results['outages']['clifty']} unit-months")
    
    print("\nNext steps:")
    print("  1. Re-run comparison: python scripts/compare_model_to_excel.py --year 2025")
    print("  2. Verify heat rate and generation differences are reduced")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
