"""
Deep dive into source sheets to extract actual calculation formulas.

Based on the initial analysis, we need to look at:
- 'CC Consumables' / 'KC Consumables' sheets for:
  - Row 25: Urea cost
  - Row 65: Limestone cost  
  - Row 84: Hydrated Lime cost
  - Row 179-180: Bioreactor Reagent + Support
  - Row 182: WWTP Reagent
  - Row 184: Misc Reagents
- 'CC Coal Burn' / 'KC Coal Burn' sheets for:
  - Row 128: Fuel Oil cost
  - Row 131: Labor & Handling cost
- 'CC Emissions' / 'KC Emissions' sheets for:
  - Row 135: Allowance cost (NOx)
- 'CC Forecast Inputs' / 'KC Forecast Inputs' sheets for:
  - Row 217: Temporary Coal Storage
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

SOURCE_DIR = Path("docs/source_documents")
MAIN_FILE = "OVEC IKEC Energy Budget- 2026 Year-End BOD Update- Final- (12-8-25).xlsm"


def extract_row_context(ws, row_num: int, context_rows: int = 5) -> list:
    """Extract a row and its context."""
    results = []
    start_row = max(1, row_num - context_rows)
    end_row = min(ws.max_row, row_num + context_rows)
    
    for r in range(start_row, end_row + 1):
        row_data = {
            "row": r,
            "is_target": r == row_num,
            "cells": []
        }
        for c in range(1, min(ws.max_column + 1, 8)):  # First 8 columns
            cell = ws.cell(row=r, column=c)
            value = cell.value
            if value is not None:
                row_data["cells"].append({
                    "col": get_column_letter(c),
                    "value": str(value)[:100] if value else "",
                    "is_formula": isinstance(value, str) and value.startswith("=")
                })
        results.append(row_data)
    
    return results


def analyze_sheet_formulas(ws, sheet_name: str, target_rows: dict) -> list:
    """Analyze specific rows in a sheet."""
    results = []
    
    for row_num, description in target_rows.items():
        # Get row label from column A
        label_cell = ws.cell(row=row_num, column=1)
        row_label = str(label_cell.value) if label_cell.value else f"Row {row_num}"
        
        result = {
            "sheet": sheet_name,
            "row": row_num,
            "description": description,
            "row_label": row_label,
            "formulas": [],
            "context": []
        }
        
        # Get formulas from this row (columns B through G typically have the logic)
        for c in range(2, 8):
            cell = ws.cell(row=row_num, column=c)
            if cell.value:
                col_header_cell = ws.cell(row=1, column=c)
                col_header = str(col_header_cell.value) if col_header_cell.value else f"Col {c}"
                
                formula_info = {
                    "col": get_column_letter(c),
                    "col_header": col_header,
                    "value": str(cell.value),
                    "is_formula": isinstance(cell.value, str) and cell.value.startswith("=")
                }
                result["formulas"].append(formula_info)
        
        # Get context (rows above to understand inputs)
        result["context"] = extract_row_context(ws, row_num, context_rows=10)
        
        results.append(result)
    
    return results


def main():
    output_lines = []
    output_lines.append("=" * 100)
    output_lines.append("DETAILED FORMULA EXTRACTION FOR FUEL COST CATEGORIES")
    output_lines.append("=" * 100)
    
    filepath = SOURCE_DIR / MAIN_FILE
    if not filepath.exists():
        print(f"File not found: {filepath}")
        return
    
    wb = load_workbook(filepath, data_only=False)
    
    # Define sheets and target rows
    sheet_targets = {
        "CC Consumables": {
            25: "Urea Cost (Total)",
            24: "Urea Usage Rate",
            23: "Urea Price",
            22: "Urea Tons",
            65: "Limestone Cost (Total)",
            64: "Limestone Usage",
            63: "Limestone Price",
            84: "Hydrated Lime Cost (Total)",
            83: "Hydrated Lime Usage",
            82: "Hydrated Lime Price",
            179: "Bioreactor Reagent Cost",
            180: "Bioreactor Support Cost",
            182: "WWTP Reagent Cost",
            184: "Misc Reagents Cost",
        },
        "KC Consumables": {
            25: "Urea Cost (Total)",
            65: "Limestone Cost (Total)",
            84: "Hydrated Lime Cost (Total)",
            179: "Bioreactor Reagent Cost",
            180: "Bioreactor Support Cost",
            182: "WWTP Reagent Cost",
            184: "Misc Reagents Cost",
        },
        "CC Coal Burn": {
            128: "Fuel Oil Cost",
            127: "Fuel Oil Usage",
            126: "Fuel Oil Price",
            131: "Labor & Handling Cost",
            130: "Coal Handling",
        },
        "KC Coal Burn": {
            128: "Fuel Oil Cost",
            131: "Labor & Handling Cost",
        },
        "CC Emissions": {
            135: "NOx Allowance Cost",
            134: "NOx Allowance Rate",
            133: "NOx Emissions",
        },
        "KC Emissions": {
            135: "NOx Allowance Cost",
            134: "NOx Allowance Rate",
            133: "NOx Emissions",
        },
        "CC Forecast Inputs": {
            216: "Temp Storage Header",
            217: "Temp Storage Cost",
        },
        "KC Forecast Inputs": {
            216: "Temp Storage Header",
            217: "Temp Storage Cost",
        },
    }
    
    for sheet_name, target_rows in sheet_targets.items():
        if sheet_name not in wb.sheetnames:
            output_lines.append(f"\n[SKIPPED] Sheet not found: {sheet_name}")
            continue
        
        ws = wb[sheet_name]
        output_lines.append(f"\n{'='*80}")
        output_lines.append(f"SHEET: {sheet_name}")
        output_lines.append("=" * 80)
        
        results = analyze_sheet_formulas(ws, sheet_name, target_rows)
        
        for result in results:
            output_lines.append(f"\n--- Row {result['row']}: {result['description']} ---")
            output_lines.append(f"Row Label: {result['row_label']}")
            
            output_lines.append("\nFORMULAS/VALUES:")
            for f in result["formulas"]:
                marker = "FORMULA" if f["is_formula"] else "VALUE"
                output_lines.append(f"  [{marker}] {f['col']}: {f['value']}")
            
            # Show condensed context
            output_lines.append("\nCONTEXT (surrounding rows):")
            for ctx_row in result["context"]:
                marker = ">>>" if ctx_row["is_target"] else "   "
                cells_str = " | ".join([f"{c['col']}:{c['value'][:50]}" for c in ctx_row["cells"][:4]])
                output_lines.append(f"  {marker} Row {ctx_row['row']:3d}: {cells_str}")
    
    wb.close()
    
    # Write output
    output_path = Path("docs/excel_source_formulas.txt")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))
    
    print(f"\nAnalysis complete! Output written to: {output_path}")
    
    # Also print summary to console
    print("\n" + "=" * 60)
    print("KEY FORMULAS FOUND:")
    print("=" * 60)
    for line in output_lines:
        if "FORMULA" in line or "Row Label:" in line or "---" in line:
            print(line)


if __name__ == "__main__":
    main()

