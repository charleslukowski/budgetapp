"""
Test script for Excel import functionality.

Tests the import of use factors, heat rates, and coal contracts
from Excel files into the database.

Usage:
    python scripts/test_excel_import.py
"""

import sys
import os
import tempfile
from decimal import Decimal
from datetime import date

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))


def print_result(test_name: str, passed: bool, message: str = ""):
    """Print test result."""
    status = "[PASS]" if passed else "[FAIL]"
    print(f"  {status} {test_name}")
    if message:
        print(f"         {message}")


def create_test_workbook():
    """Create a test Excel workbook with sample data."""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return None
    
    wb = openpyxl.Workbook()
    
    # Create Use Factor Input sheet
    ws_uf = wb.active
    ws_uf.title = "Use Factor Input"
    
    # Headers
    ws_uf['A1'] = "Plant"
    for i, month in enumerate(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], start=2):
        ws_uf.cell(row=1, column=i, value=month)
    
    # Kyger Creek data
    ws_uf['A2'] = "Kyger Creek"
    for i in range(12):
        ws_uf.cell(row=2, column=i+2, value=85 + i * 0.5)  # 85% to 90.5%
    
    # Kyger Creek ozone (not needed for KC since all units have SCR)
    ws_uf['A3'] = "Kyger Creek Ozone"
    for i in range(12):
        ws_uf.cell(row=3, column=i+2, value=85 + i * 0.5)
    
    # Clifty Creek data
    ws_uf['A4'] = "Clifty Creek"
    for i in range(12):
        ws_uf.cell(row=4, column=i+2, value=80 + i * 0.4)  # 80% to 84.4%
    
    # Clifty Unit 6 ozone data
    ws_uf['A5'] = "Clifty Creek Unit 6 Ozone"
    for i in range(12):
        month = i + 1
        if month in {5, 6, 7, 8, 9}:  # Ozone season
            ws_uf.cell(row=5, column=i+2, value=0)  # Curtailed
        else:
            ws_uf.cell(row=5, column=i+2, value=80 + i * 0.4)
    
    # Create Heat Rate sheet
    ws_hr = wb.create_sheet("Heat Rate")
    
    # Headers
    ws_hr['A1'] = "Plant"
    for i, month in enumerate(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], start=2):
        ws_hr.cell(row=1, column=i, value=month)
    
    # Kyger Creek baseline
    ws_hr['A2'] = "Kyger Creek Baseline"
    for i in range(12):
        ws_hr.cell(row=2, column=i+2, value=9800 + i * 10)
    
    # Kyger Creek SUF
    ws_hr['A3'] = "Kyger Creek SUF Correction"
    ws_hr['B3'] = 50
    
    # Clifty Creek baseline
    ws_hr['A4'] = "Clifty Creek Baseline"
    for i in range(12):
        ws_hr.cell(row=4, column=i+2, value=9850 + i * 12)
    
    # Clifty Creek SUF
    ws_hr['A5'] = "Clifty Creek SUF Correction"
    ws_hr['B5'] = 45
    
    # Create Coal Contracts sheet
    ws_cc = wb.create_sheet("Coal Contracts Annual View")
    
    # Headers
    headers = ['Contract ID', 'Supplier', 'Plant', 'Start Date', 'End Date',
               'Annual Tons', 'BTU/lb', 'Coal Price', 'Barge Price', 'Region']
    for i, header in enumerate(headers, start=1):
        ws_cc.cell(row=1, column=i, value=header)
    
    # Sample contracts
    contracts = [
        ('TEST-KC-001', 'Alpha Coal', 'Kyger Creek', date(2025, 1, 1), date(2025, 12, 31),
         500000, 12600, 55.00, 6.50, 'NAPP'),
        ('TEST-KC-002', 'Beta Mining', 'Kyger Creek', date(2025, 1, 1), date(2025, 12, 31),
         300000, 12400, 52.00, 6.00, 'NAPP'),
        ('TEST-CC-001', 'Gamma Energy', 'Clifty Creek', date(2025, 1, 1), date(2025, 12, 31),
         450000, 12500, 54.00, 7.00, 'NAPP'),
    ]
    
    for row_idx, contract in enumerate(contracts, start=2):
        for col_idx, value in enumerate(contract, start=1):
            ws_cc.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name
    
    wb.save(tmp_path)
    wb.close()
    
    return tmp_path


def cleanup_test_data(db):
    """Clean up test data from the database."""
    from src.models.use_factor import UseFactorInput
    from src.models.heat_rate import HeatRateInput
    from src.models.coal_contract import CoalContract
    
    # Delete test contracts
    db.query(CoalContract).filter(
        CoalContract.contract_id.like('TEST-%')
    ).delete(synchronize_session=False)
    
    # Delete use factors for 2025 (test year)
    db.query(UseFactorInput).filter(
        UseFactorInput.year == 2025
    ).delete(synchronize_session=False)
    
    # Delete heat rates for 2025
    db.query(HeatRateInput).filter(
        HeatRateInput.year == 2025
    ).delete(synchronize_session=False)
    
    db.commit()


def test_import_use_factors():
    """Test importing use factors from Excel."""
    print("\n" + "=" * 60)
    print("Testing Use Factor Import")
    print("=" * 60)
    
    from src.db.postgres import get_session
    from src.etl.fuel_inputs_import import import_use_factors_from_excel
    from src.models.use_factor import get_use_factors_for_year
    
    passed = 0
    failed = 0
    
    # Create test workbook
    test_file = create_test_workbook()
    if not test_file:
        return 0, 1
    
    try:
        with get_session() as db:
            cleanup_test_data(db)
            
            # Test 1: Import Kyger Creek use factors
            result = import_use_factors_from_excel(
                db,
                file_path=test_file,
                plant_id=1,
                year=2025,
                sheet_name="Use Factor Input",
            )
            
            if result["months_imported"] == 12:
                print_result("Import KC use factors", True, "12 months imported")
                passed += 1
            else:
                print_result("Import KC use factors", False, 
                           f"Only {result['months_imported']} months, errors: {result.get('errors')}")
                failed += 1
            
            # Test 2: Verify imported values
            use_factors = get_use_factors_for_year(db, plant_id=1, year=2025)
            
            if len(use_factors) >= 12:
                print_result("Use factors stored in DB", True)
                passed += 1
            else:
                print_result("Use factors stored in DB", False, f"Only {len(use_factors)} months")
                failed += 1
            
            # Test 3: Check value accuracy
            jan_uf = use_factors.get(1, {})
            expected_jan = 0.85  # 85%
            actual_jan = float(jan_uf.get("base", 0))
            
            if abs(actual_jan - expected_jan) < 0.01:
                print_result("Use factor value correct", True, f"Jan = {actual_jan:.2%}")
                passed += 1
            else:
                print_result("Use factor value correct", False, 
                           f"Expected {expected_jan:.2%}, got {actual_jan:.2%}")
                failed += 1
            
            # Test 4: Import Clifty Creek with ozone factors
            result = import_use_factors_from_excel(
                db,
                file_path=test_file,
                plant_id=2,
                year=2025,
                sheet_name="Use Factor Input",
            )
            
            if result["months_imported"] > 0:
                print_result("Import CC use factors", True, f"{result['months_imported']} months")
                passed += 1
            else:
                print_result("Import CC use factors", False, str(result.get("errors")))
                failed += 1
            
            cleanup_test_data(db)
            
    finally:
        os.unlink(test_file)
    
    return passed, failed


def test_import_heat_rates():
    """Test importing heat rates from Excel."""
    print("\n" + "=" * 60)
    print("Testing Heat Rate Import")
    print("=" * 60)
    
    from src.db.postgres import get_session
    from src.etl.fuel_inputs_import import import_heat_rates_from_excel
    from src.models.heat_rate import get_heat_rates_for_year
    
    passed = 0
    failed = 0
    
    # Create test workbook
    test_file = create_test_workbook()
    if not test_file:
        return 0, 1
    
    try:
        with get_session() as db:
            cleanup_test_data(db)
            
            # Test 1: Import Kyger Creek heat rates
            result = import_heat_rates_from_excel(
                db,
                file_path=test_file,
                plant_id=1,
                year=2025,
                sheet_name="Heat Rate",
            )
            
            if result["months_imported"] == 12:
                print_result("Import KC heat rates", True, "12 months imported")
                passed += 1
            else:
                print_result("Import KC heat rates", False, 
                           f"Only {result['months_imported']} months, errors: {result.get('errors')}")
                failed += 1
            
            # Test 2: Verify imported values
            heat_rates = get_heat_rates_for_year(db, plant_id=1, year=2025)
            
            if len(heat_rates) >= 12:
                print_result("Heat rates stored in DB", True)
                passed += 1
            else:
                print_result("Heat rates stored in DB", False, f"Only {len(heat_rates)} months")
                failed += 1
            
            # Test 3: Check value accuracy
            jan_hr = heat_rates.get(1, {})
            expected_jan = 9800
            actual_jan = float(jan_hr.get("plant", {}).get("baseline_heat_rate", 0))
            
            if abs(actual_jan - expected_jan) < 20:
                print_result("Heat rate value correct", True, f"Jan = {actual_jan:.0f} BTU/kWh")
                passed += 1
            else:
                print_result("Heat rate value correct", False, 
                           f"Expected {expected_jan}, got {actual_jan}")
                failed += 1
            
            cleanup_test_data(db)
            
    finally:
        os.unlink(test_file)
    
    return passed, failed


def test_import_coal_contracts():
    """Test importing coal contracts from Excel."""
    print("\n" + "=" * 60)
    print("Testing Coal Contract Import")
    print("=" * 60)
    
    from src.db.postgres import get_session
    from src.etl.fuel_inputs_import import import_coal_contracts_from_excel
    from src.models.coal_contract import CoalContract
    
    passed = 0
    failed = 0
    
    # Create test workbook
    test_file = create_test_workbook()
    if not test_file:
        return 0, 1
    
    try:
        with get_session() as db:
            cleanup_test_data(db)
            
            # Test 1: Import contracts
            result = import_coal_contracts_from_excel(
                db,
                file_path=test_file,
                sheet_name="Coal Contracts Annual View",
            )
            
            total_processed = result["contracts_imported"] + result["contracts_updated"]
            
            if total_processed >= 3:
                print_result("Import coal contracts", True, f"{total_processed} contracts")
                passed += 1
            else:
                print_result("Import coal contracts", False, 
                           f"Only {total_processed}, errors: {result.get('errors')}")
                failed += 1
            
            # Test 2: Verify contracts in DB
            contracts = db.query(CoalContract).filter(
                CoalContract.contract_id.like('TEST-%')
            ).all()
            
            if len(contracts) >= 3:
                print_result("Contracts stored in DB", True, f"{len(contracts)} found")
                passed += 1
            else:
                print_result("Contracts stored in DB", False, f"Only {len(contracts)} found")
                failed += 1
            
            # Test 3: Verify contract details
            alpha_contract = db.query(CoalContract).filter(
                CoalContract.contract_id == 'TEST-KC-001'
            ).first()
            
            if alpha_contract:
                if alpha_contract.supplier == "Alpha Coal":
                    print_result("Contract supplier correct", True)
                    passed += 1
                else:
                    print_result("Contract supplier correct", False, 
                               f"Expected Alpha Coal, got {alpha_contract.supplier}")
                    failed += 1
                
                if float(alpha_contract.annual_tons) == 500000:
                    print_result("Contract tons correct", True, "500,000 tons")
                    passed += 1
                else:
                    print_result("Contract tons correct", False, 
                               f"Expected 500000, got {float(alpha_contract.annual_tons)}")
                    failed += 1
            else:
                print_result("Contract lookup", False, "Contract TEST-KC-001 not found")
                failed += 2
            
            cleanup_test_data(db)
            
    finally:
        os.unlink(test_file)
    
    return passed, failed


def test_import_all():
    """Test importing all fuel inputs at once."""
    print("\n" + "=" * 60)
    print("Testing Full Import")
    print("=" * 60)
    
    from src.db.postgres import get_session
    from src.etl.fuel_inputs_import import import_all_fuel_inputs
    
    passed = 0
    failed = 0
    
    # Create test workbook
    test_file = create_test_workbook()
    if not test_file:
        return 0, 1
    
    try:
        with get_session() as db:
            cleanup_test_data(db)
            
            # Test: Import everything
            result = import_all_fuel_inputs(
                db,
                file_path=test_file,
                year=2025,
            )
            
            summary = result["summary"]
            
            if summary["use_factors_imported"] > 0:
                print_result("Use factors imported", True, 
                           f"{summary['use_factors_imported']} total")
                passed += 1
            else:
                print_result("Use factors imported", False, "None imported")
                failed += 1
            
            if summary["heat_rates_imported"] > 0:
                print_result("Heat rates imported", True, 
                           f"{summary['heat_rates_imported']} total")
                passed += 1
            else:
                print_result("Heat rates imported", False, "None imported")
                failed += 1
            
            if summary["contracts_imported"] > 0 or summary["contracts_updated"] > 0:
                print_result("Contracts imported", True, 
                           f"{summary['contracts_imported']} new, {summary['contracts_updated']} updated")
                passed += 1
            else:
                print_result("Contracts imported", False, "None imported")
                failed += 1
            
            cleanup_test_data(db)
            
    finally:
        os.unlink(test_file)
    
    return passed, failed


def main():
    """Run all Excel import tests."""
    print("\n" + "=" * 60)
    print("EXCEL IMPORT TESTS")
    print("=" * 60)
    print("Testing import of use factors, heat rates, and contracts from Excel")
    
    # Check dependencies
    try:
        import openpyxl
    except ImportError:
        print("\n[ERROR] openpyxl not installed")
        print("Run: pip install openpyxl")
        return False
    
    # Check database connection
    try:
        from src.db.postgres import get_session
        with get_session() as db:
            pass
        print("\nDatabase connection successful.")
    except Exception as e:
        print(f"\n[ERROR] Cannot connect to database: {e}")
        return False
    
    total_passed = 0
    total_failed = 0
    
    # Run tests
    p, f = test_import_use_factors()
    total_passed += p
    total_failed += f
    
    p, f = test_import_heat_rates()
    total_passed += p
    total_failed += f
    
    p, f = test_import_coal_contracts()
    total_passed += p
    total_failed += f
    
    p, f = test_import_all()
    total_passed += p
    total_failed += f
    
    # Summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)
    print(f"  Passed: {total_passed}")
    print(f"  Failed: {total_failed}")
    print(f"  Total:  {total_passed + total_failed}")
    
    if total_failed == 0:
        print("\n  All tests passed!")
    else:
        print(f"\n  {total_failed} test(s) failed.")
    
    return total_failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

