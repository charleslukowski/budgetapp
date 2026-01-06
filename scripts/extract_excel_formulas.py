"""
Excel Formula Extraction Script

Extracts formulas from OVEC fuel cost Excel files to understand
calculation logic for missing cost categories.
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

# Target cost categories to look for
TARGET_CATEGORIES = [
    "allowance",
    "co2",
    "carbon",
    "hydrated lime",
    "bioreactor",
    "wwtp",
    "wastewater",
    "misc",
    "miscellaneous",
    "fuel oil",
    "oil",
    "labor",
    "handling",
    "storage",
    "temp coal",
    "temporary coal",
    "reagent",
    "urea",
    "limestone",
]

# Files to analyze
SOURCE_DIR = Path("docs/source_documents")
FILES = [
    "2025 - Fuel.xlsx",
    "OVEC IKEC Energy Budget- 2026 Year-End BOD Update- Final- (12-8-25).xlsm",
    "2025 Kyger Forecast.xlsx",
]


def matches_category(text: str) -> list[str]:
    """Check if text matches any target category."""
    if not text:
        return []
    text_lower = text.lower()
    matches = []
    for cat in TARGET_CATEGORIES:
        if cat in text_lower:
            matches.append(cat)
    return matches


def extract_formulas_from_sheet(ws, sheet_name: str) -> list[dict]:
    """Extract formulas and their context from a worksheet."""
    results = []
    
    # First pass: find row and column headers
    row_headers = {}  # row_num -> header text
    col_headers = {}  # col_num -> header text
    
    # Scan first 5 columns for row headers
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                row_headers[row] = cell.value.strip()
                break
    
    # Scan first 10 rows for column headers
    for col in range(1, min(ws.max_column + 1, 50)):
        for row in range(1, 11):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                col_headers[col] = cell.value.strip()
                break
    
    # Second pass: find cells with formulas that relate to our categories
    for row in range(1, min(ws.max_row + 1, 200)):
        row_header = row_headers.get(row, "")
        row_matches = matches_category(row_header)
        
        for col in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row=row, column=col)
            col_header = col_headers.get(col, "")
            col_matches = matches_category(col_header)
            
            # Check if this row or column is relevant
            if row_matches or col_matches:
                cell_ref = f"{get_column_letter(col)}{row}"
                
                # Get formula or value
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula = cell.value
                    else:
                        formula = None
                    
                    results.append({
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "row_header": row_header,
                        "col_header": col_header,
                        "formula": formula,
                        "value": str(cell.value) if not formula else None,
                        "categories": list(set(row_matches + col_matches)),
                    })
    
    return results


def analyze_sheet_structure(ws, sheet_name: str) -> dict:
    """Analyze the overall structure of a sheet."""
    # Get all unique row headers in first column
    row_headers = []
    for row in range(1, min(ws.max_row + 1, 100)):
        cell = ws.cell(row=row, column=1)
        if cell.value and isinstance(cell.value, str):
            row_headers.append((row, cell.value.strip()))
    
    # Get column headers from first row
    col_headers = []
    for col in range(1, min(ws.max_column + 1, 20)):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            col_headers.append((col, str(cell.value).strip()))
    
    return {
        "sheet_name": sheet_name,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "row_headers": row_headers,
        "col_headers": col_headers,
    }


def extract_all_formulas_in_range(ws, sheet_name: str, start_row: int, end_row: int, start_col: int, end_col: int) -> list[dict]:
    """Extract all formulas in a specific range."""
    results = []
    
    for row in range(start_row, end_row + 1):
        row_label = ws.cell(row=row, column=1).value
        if row_label:
            row_label = str(row_label).strip()
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell_ref = f"{get_column_letter(col)}{row}"
                
                # Try to get formula
                formula = None
                value = cell.value
                
                # In openpyxl, formulas are stored as strings starting with =
                if isinstance(value, str) and value.startswith("="):
                    formula = value
                    value = None
                
                col_header = ws.cell(row=1, column=col).value
                if col_header:
                    col_header = str(col_header).strip()
                
                results.append({
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "row": row,
                    "col": col,
                    "row_label": row_label,
                    "col_header": col_header,
                    "formula": formula,
                    "value": value,
                })
    
    return results


def main():
    output_lines = []
    output_lines.append("=" * 80)
    output_lines.append("EXCEL FORMULA ANALYSIS FOR FUEL COST CATEGORIES")
    output_lines.append("=" * 80)
    output_lines.append("")
    
    for filename in FILES:
        filepath = SOURCE_DIR / filename
        if not filepath.exists():
            output_lines.append(f"\n[SKIPPED] File not found: {filepath}")
            continue
        
        output_lines.append(f"\n{'=' * 80}")
        output_lines.append(f"FILE: {filename}")
        output_lines.append("=" * 80)
        
        try:
            # Load with data_only=False to get formulas
            wb = load_workbook(filepath, data_only=False)
            
            output_lines.append(f"\nSheets found: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Skip empty or very small sheets
                if ws.max_row < 3 or ws.max_column < 2:
                    continue
                
                # Analyze structure
                structure = analyze_sheet_structure(ws, sheet_name)
                
                # Look for cost-related sheets
                sheet_relevant = any(
                    cat in sheet_name.lower() 
                    for cat in ["fuel", "cost", "energy", "budget", "kyger", "clifty", "ovec"]
                )
                
                # Check row headers for relevance
                row_relevance = []
                for row_num, header in structure["row_headers"]:
                    matches = matches_category(header)
                    if matches:
                        row_relevance.append((row_num, header, matches))
                
                if sheet_relevant or row_relevance:
                    output_lines.append(f"\n{'-' * 60}")
                    output_lines.append(f"SHEET: {sheet_name}")
                    output_lines.append(f"Size: {ws.max_row} rows x {ws.max_column} cols")
                    output_lines.append("-" * 60)
                    
                    # Print all row headers
                    output_lines.append("\nRow Headers (Column A):")
                    for row_num, header in structure["row_headers"][:50]:
                        marker = " *** " if matches_category(header) else "     "
                        output_lines.append(f"  {marker}Row {row_num:3d}: {header}")
                    
                    # Print column headers
                    if structure["col_headers"]:
                        output_lines.append("\nColumn Headers (Row 1):")
                        for col_num, header in structure["col_headers"]:
                            output_lines.append(f"       Col {get_column_letter(col_num)}: {header}")
                    
                    # Extract and show relevant formulas
                    if row_relevance:
                        output_lines.append("\n\nRELEVANT FORMULAS:")
                        output_lines.append("-" * 40)
                        
                        for row_num, header, matches in row_relevance:
                            output_lines.append(f"\n>>> Row {row_num}: {header}")
                            output_lines.append(f"    Categories: {matches}")
                            
                            # Get all cells in this row
                            for col in range(2, min(ws.max_column + 1, 20)):
                                cell = ws.cell(row=row_num, column=col)
                                if cell.value is not None:
                                    cell_ref = f"{get_column_letter(col)}{row_num}"
                                    col_header = ws.cell(row=1, column=col).value or f"Col {col}"
                                    
                                    if isinstance(cell.value, str) and cell.value.startswith("="):
                                        output_lines.append(f"    {cell_ref} ({col_header}): FORMULA = {cell.value}")
                                    else:
                                        val = cell.value
                                        if isinstance(val, float):
                                            val = f"{val:,.2f}"
                                        output_lines.append(f"    {cell_ref} ({col_header}): VALUE = {val}")
            
            wb.close()
            
        except Exception as e:
            output_lines.append(f"\n[ERROR] Failed to process {filename}: {e}")
            import traceback
            output_lines.append(traceback.format_exc())
    
    # Write output
    output_path = Path("docs/excel_formula_analysis.txt")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))
    
    print(f"\nAnalysis complete! Output written to: {output_path}")
    print("\n" + "=" * 60)
    print("SUMMARY OF FINDINGS:")
    print("=" * 60)
    
    # Print a summary to console
    for line in output_lines[-100:]:  # Last 100 lines
        print(line)


if __name__ == "__main__":
    main()

