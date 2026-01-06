"""Fuel forecast workflow page routes.

Provides the multi-step guided workflow for creating fuel forecasts:
1. Start Point - Choose roll forward, copy, or fresh start
2. Coal Position - Update inventory, contracts, prices, blend, transport
3. Generation - Set capacity, use factors, heat rates, deductions
4. Other Costs - Reagents, byproducts, escalation
5. Review & Save - Calculate and save forecast

Field names are aligned with the driver framework in src/engine/default_drivers.py
"""

from fastapi import APIRouter, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from pathlib import Path
from datetime import datetime, date
from typing import Optional
from decimal import Decimal

from src.engine.forecast_session import (
    create_session,
    get_session,
    update_session,
    ForecastSession,
    STEP_CONFIG,
    TOTAL_STEPS,
)
from src.db.postgres import get_session as get_db_session

router = APIRouter(prefix="/fuel-forecast", tags=["Fuel Forecast"])

# Set up templates
templates_path = Path(__file__).parent.parent.parent.parent / "templates"
templates = Jinja2Templates(directory=str(templates_path))


# =============================================================================
# Default values matching driver framework defaults
# =============================================================================

# Coal Price Drivers
DEFAULTS = {
    # Inventory
    "inventory_beginning_kc": 180000,
    "inventory_beginning_cc": 150000,
    "inventory_target_days": 50,
    "contracted_deliveries_tons": 75000,
    
    # Coal Prices ($/ton FOB mine)
    "coal_price_eastern": 55.00,
    "coal_price_ilb": 45.00,
    "coal_price_prb": 15.00,
    
    # Coal Blend (%)
    "coal_blend_eastern_pct": 100,
    "coal_blend_ilb_pct": 0,
    "coal_blend_prb_pct": 0,
    
    # Coal Quality (BTU/lb)
    "coal_btu_eastern": 12600,
    "coal_btu_ilb": 11400,
    
    # Transportation ($/ton)
    "barge_rate_ohio": 6.00,
    "barge_rate_upper_ohio": 7.50,
    "rail_rate_prb": 30.00,
    
    # Capacity (MW)
    "capacity_mw_kc": 1025,
    "capacity_mw_cc": 1025,
    
    # Use Factor & Capacity Factor (%)
    "use_factor": 85,
    "capacity_factor_target": 70,
    
    # Heat Rates (BTU/kWh)
    "heat_rate_baseline_kc": 9850,
    "heat_rate_baseline_cc": 9900,
    "heat_rate_suf_correction": 0,
    "heat_rate_prb_penalty": 100,
    
    # Outages (days)
    "outage_days_planned": 0,
    "outage_days_forced": 0,
    
    # Deductions
    "fgd_aux_pct": 2.5,
    "gsu_loss_pct": 0.5545,
    "reserve_mw": 10,
    
    # Escalation (% annual)
    "escalation_coal_annual": 2.0,
    "escalation_transport_annual": 2.5,
    "escalation_reagent_annual": 2.0,
    
    # Reagent Prices ($/ton)
    "urea_price_per_ton": 350,
    "limestone_price_per_ton": 25,
    "hydrated_lime_price_per_ton": 150,
    "mercury_reagent_cost_monthly": 5000,
    
    # Byproduct Prices ($/ton)
    "ash_sale_price_per_ton": 8.00,
    "ash_disposal_cost_per_ton": 15.00,
    "gypsum_sale_price_per_ton": 5.00,
    "fly_ash_sale_pct": 50,
}


def calculate_forecast_from_session(session: ForecastSession, year: int) -> tuple:
    """Calculate fuel costs from session pending changes.
    
    Args:
        session: The forecast session with pending changes
        year: Forecast year
        
    Returns:
        Tuple of (summary_results dict, monthly_results list)
    """
    from src.engine.default_drivers import create_default_fuel_model
    from src.engine.fuel_model import (
        FuelModelInputs,
        calculate_annual_fuel_costs,
        summarize_annual_fuel_costs,
        create_kyger_params,
        create_clifty_params,
    )
    from src.engine.generation import PlantParams
    from src.engine.coal_burn import HeatRateParams, CoalQuality
    
    # Create the fuel model and populate with session values
    model = create_default_fuel_model()
    
    # Map session pending changes to driver values
    driver_mapping = {
        # Coal prices
        "coal_price_eastern": "coal_price_eastern",
        "coal_price_ilb": "coal_price_ilb",
        "coal_price_prb": "coal_price_prb",
        "coal_blend_eastern_pct": "coal_blend_eastern_pct",
        "coal_blend_ilb_pct": "coal_blend_ilb_pct",
        "coal_blend_prb_pct": "coal_blend_prb_pct",
        "coal_btu_eastern": "coal_btu_eastern",
        "coal_btu_ilb": "coal_btu_ilb",
        # Transportation
        "barge_rate_ohio": "barge_rate_ohio",
        "barge_rate_upper_ohio": "barge_rate_upper_ohio",
        "rail_rate_prb": "rail_rate_prb",
        # Generation
        "use_factor": "use_factor",
        "capacity_factor_target": "capacity_factor_target",
        # Heat rates
        "heat_rate_suf_correction": "heat_rate_suf_correction",
        "heat_rate_prb_penalty": "heat_rate_prb_penalty",
        # Deductions
        "fgd_aux_pct": "fgd_aux_pct",
        "gsu_loss_pct": "gsu_loss_pct",
        "reserve_mw": "reserve_mw",
        # Escalation
        "escalation_coal_annual": "escalation_coal_annual",
        "escalation_transport_annual": "escalation_transport_annual",
        "escalation_reagent_annual": "escalation_reagent_annual",
        # Inventory
        "inventory_target_days": "inventory_target_days",
        "contracted_deliveries_tons": "contracted_deliveries_tons",
    }
    
    # Apply pending changes to the model
    for ui_field, driver_name in driver_mapping.items():
        value = session.get_pending_change(ui_field)
        if value is not None and driver_name in model.drivers:
            model.set_driver_value(driver_name, year, None, Decimal(str(value)))
    
    # Handle plant-specific values
    # Capacity
    capacity_kc = get_value(session, "capacity_mw_kc")
    capacity_cc = get_value(session, "capacity_mw_cc")
    model.set_driver_value("capacity_mw", year, None, Decimal(str(capacity_kc)), plant_id=1)
    model.set_driver_value("capacity_mw", year, None, Decimal(str(capacity_cc)), plant_id=2)
    
    # Heat rate baselines
    hr_kc = get_value(session, "heat_rate_baseline_kc")
    hr_cc = get_value(session, "heat_rate_baseline_cc")
    model.set_driver_value("heat_rate_baseline", year, None, Decimal(str(hr_kc)), plant_id=1)
    model.set_driver_value("heat_rate_baseline", year, None, Decimal(str(hr_cc)), plant_id=2)
    
    # Beginning inventory
    inv_kc = get_value(session, "inventory_beginning_kc")
    inv_cc = get_value(session, "inventory_beginning_cc")
    model.set_driver_value("inventory_beginning_tons", year, None, Decimal(str(inv_kc)), plant_id=1)
    model.set_driver_value("inventory_beginning_tons", year, None, Decimal(str(inv_cc)), plant_id=2)
    
    # Build FuelModelInputs for each plant using session values
    use_factor_pct = Decimal(str(get_value(session, "use_factor")))
    use_factor = use_factor_pct / Decimal("100")
    
    coal_price = Decimal(str(get_value(session, "coal_price_eastern")))
    barge_rate = Decimal(str(get_value(session, "barge_rate_ohio")))
    
    # Create plant params with session capacity values
    # Kyger Creek has 5 units, Clifty Creek has 6 units
    from src.engine.generation import UnitParams
    
    kyger_unit_capacity = Decimal(str(capacity_kc)) / Decimal("5")  # 5 units
    kyger_units = [UnitParams(unit_number=i, capacity_mw=kyger_unit_capacity) for i in range(1, 6)]
    kyger_params = PlantParams(
        plant_name="Kyger Creek",
        units=kyger_units,
    )
    
    clifty_unit_capacity = Decimal(str(capacity_cc)) / Decimal("6")  # 6 units
    clifty_units = [UnitParams(unit_number=i, capacity_mw=clifty_unit_capacity) for i in range(1, 7)]
    clifty_params = PlantParams(
        plant_name="Clifty Creek",
        units=clifty_units,
    )
    
    # Heat rate params
    heat_rate_kc = HeatRateParams(
        baseline_heat_rate=Decimal(str(hr_kc)),
        suf_correction=Decimal(str(get_value(session, "heat_rate_suf_correction"))),
        prb_blend_pct=Decimal(str(get_value(session, "coal_blend_prb_pct"))) / Decimal("100"),
    )
    
    heat_rate_cc = HeatRateParams(
        baseline_heat_rate=Decimal(str(hr_cc)),
        suf_correction=Decimal(str(get_value(session, "heat_rate_suf_correction"))),
        prb_blend_pct=Decimal(str(get_value(session, "coal_blend_prb_pct"))) / Decimal("100"),
    )
    
    # Coal quality from BTU values
    coal_quality = CoalQuality(
        name="Blended",
        btu_per_lb=Decimal(str(get_value(session, "coal_btu_eastern"))),
        moisture_pct=Decimal("0.08"),
        ash_pct=Decimal("0.10"),
        so2_lb_per_mmbtu=Decimal("5.0"),
    )
    
    # Create inputs for each plant
    kyger_inputs = FuelModelInputs(
        plant_params=kyger_params,
        heat_rate_params=heat_rate_kc,
        coal_quality=coal_quality,
        use_factor=use_factor,
        coal_price_per_ton=coal_price,
        barge_price_per_ton=barge_rate,
    )
    
    clifty_inputs = FuelModelInputs(
        plant_params=clifty_params,
        heat_rate_params=heat_rate_cc,
        coal_quality=coal_quality,
        use_factor=use_factor,
        coal_price_per_ton=coal_price,
        barge_price_per_ton=barge_rate,
    )
    
    # Calculate annual fuel costs for each plant
    with get_db_session() as db:
        kyger_monthly = calculate_annual_fuel_costs(db, 1, year, kyger_inputs)
        clifty_monthly = calculate_annual_fuel_costs(db, 2, year, clifty_inputs)
    
    # Summarize results
    kyger_annual = summarize_annual_fuel_costs(kyger_monthly)
    clifty_annual = summarize_annual_fuel_costs(clifty_monthly)
    
    # Combine system totals
    total_mwh = kyger_annual.get("total_mwh", 0) + clifty_annual.get("total_mwh", 0)
    total_fuel_cost = kyger_annual.get("total_fuel_cost", 0) + clifty_annual.get("total_fuel_cost", 0)
    avg_capacity_factor = (kyger_annual.get("avg_capacity_factor", 0) + clifty_annual.get("avg_capacity_factor", 0)) / 2
    
    results = {
        "total_cost": total_fuel_cost,
        "cost_per_mwh": total_fuel_cost / total_mwh if total_mwh > 0 else 0,
        "total_mwh": total_mwh,
        "capacity_factor": avg_capacity_factor,
        "kyger": kyger_annual,
        "clifty": clifty_annual,
    }
    
    # Build monthly results for display
    monthly_results = []
    for i in range(12):
        month_num = i + 1
        kyger_month = kyger_monthly[i] if i < len(kyger_monthly) else None
        clifty_month = clifty_monthly[i] if i < len(clifty_monthly) else None
        
        if kyger_month and clifty_month:
            monthly_results.append({
                "month": month_num,
                "mwh": float(kyger_month.net_delivered_mwh + clifty_month.net_delivered_mwh),
                "coal_cost": float(kyger_month.coal_cost + clifty_month.coal_cost),
                "consumables_cost": float(kyger_month.consumables_cost + clifty_month.consumables_cost),
                "byproduct_net": float(kyger_month.byproduct_net_cost + clifty_month.byproduct_net_cost),
                "byproduct_sales_revenue": float(kyger_month.byproduct_sales_revenue + clifty_month.byproduct_sales_revenue),
                "byproduct_disposal_costs": float(kyger_month.byproduct_disposal_costs + clifty_month.byproduct_disposal_costs),
                "byproduct_misc_expense": float(kyger_month.byproduct_misc_expense + clifty_month.byproduct_misc_expense),
                "total_cost": float(kyger_month.total_fuel_cost + clifty_month.total_fuel_cost),
                "cost_per_mwh": float((kyger_month.total_fuel_cost + clifty_month.total_fuel_cost) / 
                                     (kyger_month.net_delivered_mwh + clifty_month.net_delivered_mwh))
                                if (kyger_month.net_delivered_mwh + clifty_month.net_delivered_mwh) > 0 else 0,
            })
    
    return results, monthly_results


def save_forecast_scenario(session: ForecastSession, name: str, notes: str, year: int) -> dict:
    """Save the forecast scenario to the database.
    
    Creates a Scenario and ScenarioInputSnapshot so the scenario can be 
    compared with other scenarios using the comparison feature.
    
    Args:
        session: The forecast session with pending changes
        name: Scenario name
        notes: Optional notes
        year: Forecast year
        
    Returns:
        Dict with scenario_id and status
    """
    from src.models.scenario import Scenario, ScenarioType, ScenarioStatus
    from src.models.scenario_inputs import ScenarioInputSnapshot
    
    with get_db_session() as db:
        # Create new scenario
        scenario = Scenario(
            name=name,
            description=notes,
            scenario_type=ScenarioType.INTERNAL_FORECAST,
            status=ScenarioStatus.DRAFT,
            created_by="forecast_workflow",
            parent_scenario_id=session.base_scenario_id,
        )
        db.add(scenario)
        db.flush()  # Get the ID
        
        # Create input snapshot for comparison feature
        # Extract key metrics from session pending changes
        snapshot = ScenarioInputSnapshot(
            scenario_id=scenario.id,
            year=year,
            use_factors=_extract_use_factors_from_session(session, year),
            heat_rates=_extract_heat_rates_from_session(session),
            coal_pricing=_extract_coal_pricing_from_session(session),
            outages={},  # Outages are in the database, not in session
            other_params={
                "limestone_price": session.pending_changes.get("limestone_price_per_ton", 30),
                "urea_price": session.pending_changes.get("urea_price_per_ton", 410),
                "ash_sale_price": session.pending_changes.get("ash_sale_price_per_ton", 15),
                "gypsum_sale_price": session.pending_changes.get("gypsum_sale_price_per_ton", 3),
            },
            avg_use_factor_kc=Decimal(str(session.pending_changes.get("use_factor_base_kc", 0.85))),
            avg_use_factor_cc=Decimal(str(session.pending_changes.get("use_factor_base_cc", 0.85))),
            avg_heat_rate_kc=Decimal(str(session.pending_changes.get("heat_rate_baseline_kc", 9850))),
            avg_heat_rate_cc=Decimal(str(session.pending_changes.get("heat_rate_baseline_cc", 9850))),
            total_outage_days_kc=Decimal("0"),  # Would need to sum from db
            total_outage_days_cc=Decimal("0"),
            avg_coal_price=Decimal(str(session.pending_changes.get("coal_price_per_ton", 61))),
            change_summary=notes,
            created_by="forecast_workflow",
        )
        db.add(snapshot)
        
        db.commit()
        
        return {
            "scenario_id": scenario.id,
            "name": scenario.name,
        }


def _extract_use_factors_from_session(session: ForecastSession, year: int) -> dict:
    """Extract use factors from session for snapshot."""
    result = {"1": {}, "2": {}}
    for month in range(1, 13):
        kc_uf = session.pending_changes.get(f"kc_use_factor_{month}", 85) / 100
        cc_uf = session.pending_changes.get(f"cc_use_factor_{month}", 85) / 100
        result["1"][str(month)] = {"base": kc_uf, "ozone_non_scr": kc_uf}
        result["2"][str(month)] = {"base": cc_uf, "ozone_non_scr": cc_uf}
    return result


def _extract_heat_rates_from_session(session: ForecastSession) -> dict:
    """Extract heat rates from session for snapshot."""
    kc_baseline = session.pending_changes.get("heat_rate_baseline_kc", 9850)
    cc_baseline = session.pending_changes.get("heat_rate_baseline_cc", 9850)
    result = {"1": {}, "2": {}}
    for month in range(1, 13):
        result["1"][str(month)] = {"baseline": kc_baseline, "suf": 0, "prb": 0}
        result["2"][str(month)] = {"baseline": cc_baseline, "suf": 0, "prb": 0}
    return result


def _extract_coal_pricing_from_session(session: ForecastSession) -> dict:
    """Extract coal pricing from session for snapshot."""
    coal_price = session.pending_changes.get("coal_price_per_ton", 61)
    barge_price = session.pending_changes.get("barge_price_per_ton", 6.5)
    btu = session.pending_changes.get("coal_btu_per_lb", 11500)
    result = {"1": {}, "2": {}}
    for month in range(1, 13):
        result["1"][str(month)] = {"coal_price": coal_price, "barge_price": barge_price, "btu": btu}
        result["2"][str(month)] = {"coal_price": coal_price, "barge_price": barge_price, "btu": btu}
    return result


def load_scenario_into_session(session: ForecastSession, scenario_id: int, year: int) -> dict:
    """Load driver values from a scenario into the session's pending changes.
    
    Args:
        session: The forecast session to populate
        scenario_id: Source scenario ID
        year: Year to load values for
        
    Returns:
        Dict with loaded driver count and scenario name
    """
    from src.models.scenario import Scenario
    from src.engine.scenario_drivers import export_scenario_drivers
    
    with get_db_session() as db:
        # Get scenario details
        scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
        if not scenario:
            return {"loaded": 0, "name": None}
        
        session.base_scenario_name = scenario.name
        
        # Export driver values from the scenario
        driver_set = export_scenario_drivers(db, scenario_id, year)
        
        loaded_count = 0
        prior_values = {}
        
        # Mapping from driver names to UI field names
        driver_to_ui = {
            # Coal prices
            "coal_price_eastern": "coal_price_eastern",
            "coal_price_ilb": "coal_price_ilb",
            "coal_price_prb": "coal_price_prb",
            "coal_blend_eastern_pct": "coal_blend_eastern_pct",
            "coal_blend_ilb_pct": "coal_blend_ilb_pct",
            "coal_blend_prb_pct": "coal_blend_prb_pct",
            "coal_btu_eastern": "coal_btu_eastern",
            "coal_btu_ilb": "coal_btu_ilb",
            # Transportation
            "barge_rate_ohio": "barge_rate_ohio",
            "barge_rate_upper_ohio": "barge_rate_upper_ohio",
            "rail_rate_prb": "rail_rate_prb",
            # Generation
            "use_factor": "use_factor",
            "capacity_factor_target": "capacity_factor_target",
            # Heat rates
            "heat_rate_suf_correction": "heat_rate_suf_correction",
            "heat_rate_prb_penalty": "heat_rate_prb_penalty",
            # Deductions
            "fgd_aux_pct": "fgd_aux_pct",
            "gsu_loss_pct": "gsu_loss_pct",
            "reserve_mw": "reserve_mw",
            # Outages
            "outage_days_planned": "outage_days_planned",
            "outage_days_forced": "outage_days_forced",
            # Escalation
            "escalation_coal_annual": "escalation_coal_annual",
            "escalation_transport_annual": "escalation_transport_annual",
            "escalation_reagent_annual": "escalation_reagent_annual",
            # Inventory
            "inventory_target_days": "inventory_target_days",
            "contracted_deliveries_tons": "contracted_deliveries_tons",
            # Reagents
            "urea_price_per_ton": "urea_price_per_ton",
            "limestone_price_per_ton": "limestone_price_per_ton",
            "hydrated_lime_price_per_ton": "hydrated_lime_price_per_ton",
            # Byproducts  
            "ash_sale_price_per_ton": "ash_sale_price_per_ton",
            "ash_disposal_cost_per_ton": "ash_disposal_cost_per_ton",
            "gypsum_sale_price_per_ton": "gypsum_sale_price_per_ton",
        }
        
        # Load system-wide values
        for driver_name, ui_field in driver_to_ui.items():
            if driver_name in driver_set.driver_values:
                val_data = driver_set.driver_values[driver_name]
                value = val_data.get("annual")
                if value is not None:
                    session.set_pending_change(ui_field, value)
                    prior_values[ui_field] = value
                    loaded_count += 1
        
        # Load plant-specific values
        # Capacity
        if "capacity_mw" in driver_set.driver_values:
            cap_data = driver_set.driver_values["capacity_mw"]
            if "1" in cap_data.get("plant_specific", {}) or 1 in cap_data.get("plant_specific", {}):
                plant_1 = cap_data["plant_specific"].get("1") or cap_data["plant_specific"].get(1, {})
                val = plant_1.get("annual")
                if val is not None:
                    session.set_pending_change("capacity_mw_kc", val)
                    prior_values["capacity_mw_kc"] = val
                    loaded_count += 1
            if "2" in cap_data.get("plant_specific", {}) or 2 in cap_data.get("plant_specific", {}):
                plant_2 = cap_data["plant_specific"].get("2") or cap_data["plant_specific"].get(2, {})
                val = plant_2.get("annual")
                if val is not None:
                    session.set_pending_change("capacity_mw_cc", val)
                    prior_values["capacity_mw_cc"] = val
                    loaded_count += 1
        
        # Heat rate baseline
        if "heat_rate_baseline" in driver_set.driver_values:
            hr_data = driver_set.driver_values["heat_rate_baseline"]
            if "1" in hr_data.get("plant_specific", {}) or 1 in hr_data.get("plant_specific", {}):
                plant_1 = hr_data["plant_specific"].get("1") or hr_data["plant_specific"].get(1, {})
                val = plant_1.get("annual")
                if val is not None:
                    session.set_pending_change("heat_rate_baseline_kc", val)
                    prior_values["heat_rate_baseline_kc"] = val
                    loaded_count += 1
            if "2" in hr_data.get("plant_specific", {}) or 2 in hr_data.get("plant_specific", {}):
                plant_2 = hr_data["plant_specific"].get("2") or hr_data["plant_specific"].get(2, {})
                val = plant_2.get("annual")
                if val is not None:
                    session.set_pending_change("heat_rate_baseline_cc", val)
                    prior_values["heat_rate_baseline_cc"] = val
                    loaded_count += 1
        
        # Beginning inventory
        if "inventory_beginning_tons" in driver_set.driver_values:
            inv_data = driver_set.driver_values["inventory_beginning_tons"]
            if "1" in inv_data.get("plant_specific", {}) or 1 in inv_data.get("plant_specific", {}):
                plant_1 = inv_data["plant_specific"].get("1") or inv_data["plant_specific"].get(1, {})
                val = plant_1.get("annual")
                if val is not None:
                    session.set_pending_change("inventory_beginning_kc", val)
                    prior_values["inventory_beginning_kc"] = val
                    loaded_count += 1
            if "2" in inv_data.get("plant_specific", {}) or 2 in inv_data.get("plant_specific", {}):
                plant_2 = inv_data["plant_specific"].get("2") or inv_data["plant_specific"].get(2, {})
                val = plant_2.get("annual")
                if val is not None:
                    session.set_pending_change("inventory_beginning_cc", val)
                    prior_values["inventory_beginning_cc"] = val
                    loaded_count += 1
        
        # Store prior values for comparison
        session.pending_changes["_prior_values"] = prior_values
        
        return {
            "loaded": loaded_count,
            "name": scenario.name,
        }


def get_base_context(request: Request, session: ForecastSession, step: int) -> dict:
    """Build base template context for workflow pages."""
    return {
        "request": request,
        "session": session,
        "session_id": session.session_id,
        "current_step": step,
        "total_steps": TOTAL_STEPS,
        "steps": [s.to_dict() for s in session.get_steps()],
        "year": datetime.now().year,
        "plant_code": "KC",  # For nav compatibility
        "active_page": "fuel_forecast",
    }


def get_value(session: ForecastSession, key: str):
    """Get value from session pending changes or defaults."""
    return session.get_pending_change(key, DEFAULTS.get(key, 0))


def get_prior_value(session: ForecastSession, key: str):
    """Get prior value from session if roll-forward was used.
    
    Returns None if no prior value exists.
    """
    prior_values = session.pending_changes.get("_prior_values", {})
    return prior_values.get(key)


def get_all_prior_values(session: ForecastSession) -> dict:
    """Get all prior values from session.
    
    Returns empty dict if no prior values exist.
    """
    return session.pending_changes.get("_prior_values", {})


# =============================================================================
# Workflow Entry Points
# =============================================================================

@router.get("/new", response_class=HTMLResponse)
async def start_new_forecast(request: Request):
    """Start a new fuel forecast workflow.
    
    Creates a session and redirects to Step 1.
    """
    # Create new session with defaults
    session = create_session(
        start_mode="roll_forward",
        as_of_date=date.today(),
    )
    
    # Redirect to step 1
    return RedirectResponse(
        url=f"/fuel-forecast/{session.session_id}/step/1",
        status_code=302
    )


@router.get("/{session_id}", response_class=RedirectResponse)
async def resume_forecast(session_id: str):
    """Resume a forecast workflow at the current step."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/{session.current_step}",
        status_code=302
    )


@router.get("/{session_id}/summary")
async def get_session_summary(session_id: str):
    """Get the current session state as JSON."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return session.to_dict()


# =============================================================================
# Step 1: Start Point
# =============================================================================

@router.get("/{session_id}/step/1", response_class=HTMLResponse)
async def step1_start(request: Request, session_id: str):
    """Render Step 1: Start Point selection."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Update current step
    session.current_step = 1
    update_session(session)
    
    # Get available scenarios for copy/roll-forward
    scenarios = []
    try:
        with get_db_session() as db:
            from src.models.scenario import Scenario
            db_scenarios = db.query(Scenario).order_by(Scenario.created_at.desc()).limit(10).all()
            scenarios = [
                {"id": s.id, "name": s.name, "created_at": s.created_at.strftime("%Y-%m-%d") if s.created_at else ""}
                for s in db_scenarios
            ]
    except Exception:
        # If no scenarios table or empty, provide empty list
        scenarios = []
    
    context = get_base_context(request, session, 1)
    context.update({
        "scenarios": scenarios,
        "selected_mode": session.start_mode,
        "selected_scenario_id": session.base_scenario_id,
    })
    
    return templates.TemplateResponse("fuel_forecast/step1_start.html", context)


@router.post("/{session_id}/step/1", response_class=RedirectResponse)
async def step1_save(
    request: Request,
    session_id: str,
    start_mode: str = Form("roll_forward"),
    scenario_id: Optional[int] = Form(None),
    as_of_month: Optional[str] = Form(None),
):
    """Save Step 1 and advance to Step 2."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Update session with selections
    session.start_mode = start_mode
    session.base_scenario_id = scenario_id if start_mode in ("roll_forward", "copy") else None
    
    # Parse as-of date
    if as_of_month:
        try:
            # Expect format like "2025-11"
            year, month = as_of_month.split("-")
            session.as_of_date = date(int(year), int(month), 1)
        except ValueError:
            session.as_of_date = date.today()
    else:
        session.as_of_date = date.today()
    
    # Load source scenario driver values if roll-forward or copy mode
    if start_mode in ("roll_forward", "copy") and scenario_id:
        load_year = session.as_of_date.year if session.as_of_date else datetime.now().year
        load_result = load_scenario_into_session(session, scenario_id, load_year)
        if load_result["loaded"] > 0:
            session.add_change_note(f"Loaded {load_result['loaded']} driver values from '{load_result['name']}'")
    
    # Mark step done and advance
    session.mark_step_done(1)
    session.advance_to_step(2)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/2",
        status_code=302
    )


# =============================================================================
# Step 2: Coal Position
# =============================================================================

@router.get("/{session_id}/step/2", response_class=HTMLResponse)
async def step2_coal(request: Request, session_id: str):
    """Render Step 2: Coal Position."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(2):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/1", status_code=302)
    
    session.current_step = 2
    update_session(session)
    
    # Calculate blended price for display
    eastern_pct = get_value(session, "coal_blend_eastern_pct")
    ilb_pct = get_value(session, "coal_blend_ilb_pct")
    prb_pct = get_value(session, "coal_blend_prb_pct")
    total_pct = eastern_pct + ilb_pct + prb_pct
    
    if total_pct > 0:
        blended_price = (
            get_value(session, "coal_price_eastern") * eastern_pct +
            get_value(session, "coal_price_ilb") * ilb_pct +
            get_value(session, "coal_price_prb") * prb_pct
        ) / total_pct
    else:
        blended_price = get_value(session, "coal_price_eastern")
    
    delivered_cost = blended_price + get_value(session, "barge_rate_ohio")
    
    # Get prior values for comparison display
    prior = get_all_prior_values(session)
    has_prior = len(prior) > 0
    
    context = get_base_context(request, session, 2)
    context.update({
        # Inventory
        "inventory_beginning_kc": get_value(session, "inventory_beginning_kc"),
        "inventory_beginning_cc": get_value(session, "inventory_beginning_cc"),
        "inventory_target_days": get_value(session, "inventory_target_days"),
        "contracted_deliveries_tons": get_value(session, "contracted_deliveries_tons"),
        
        # Coal Prices
        "coal_price_eastern": get_value(session, "coal_price_eastern"),
        "coal_price_ilb": get_value(session, "coal_price_ilb"),
        "coal_price_prb": get_value(session, "coal_price_prb"),
        "blended_price": blended_price,
        
        # Coal Blend
        "coal_blend_eastern_pct": get_value(session, "coal_blend_eastern_pct"),
        "coal_blend_ilb_pct": get_value(session, "coal_blend_ilb_pct"),
        "coal_blend_prb_pct": get_value(session, "coal_blend_prb_pct"),
        
        # Coal Quality
        "coal_btu_eastern": get_value(session, "coal_btu_eastern"),
        "coal_btu_ilb": get_value(session, "coal_btu_ilb"),
        
        # Transportation
        "barge_rate_ohio": get_value(session, "barge_rate_ohio"),
        "barge_rate_upper_ohio": get_value(session, "barge_rate_upper_ohio"),
        "rail_rate_prb": get_value(session, "rail_rate_prb"),
        "delivered_cost": delivered_cost,
        
        # Prior values for comparison
        "has_prior": has_prior,
        "prior": prior,
        "base_scenario_name": session.base_scenario_name,
    })
    
    return templates.TemplateResponse("fuel_forecast/step2_coal.html", context)


@router.post("/{session_id}/step/2", response_class=RedirectResponse)
async def step2_save(
    request: Request,
    session_id: str,
    # Inventory
    inventory_beginning_kc: float = Form(DEFAULTS["inventory_beginning_kc"]),
    inventory_beginning_cc: float = Form(DEFAULTS["inventory_beginning_cc"]),
    inventory_target_days: float = Form(DEFAULTS["inventory_target_days"]),
    contracted_deliveries_tons: float = Form(DEFAULTS["contracted_deliveries_tons"]),
    # Coal Prices
    coal_price_eastern: float = Form(DEFAULTS["coal_price_eastern"]),
    coal_price_ilb: float = Form(DEFAULTS["coal_price_ilb"]),
    coal_price_prb: float = Form(DEFAULTS["coal_price_prb"]),
    # Coal Blend
    coal_blend_eastern_pct: float = Form(DEFAULTS["coal_blend_eastern_pct"]),
    coal_blend_ilb_pct: float = Form(DEFAULTS["coal_blend_ilb_pct"]),
    coal_blend_prb_pct: float = Form(DEFAULTS["coal_blend_prb_pct"]),
    # Coal Quality
    coal_btu_eastern: float = Form(DEFAULTS["coal_btu_eastern"]),
    coal_btu_ilb: float = Form(DEFAULTS["coal_btu_ilb"]),
    # Transportation
    barge_rate_ohio: float = Form(DEFAULTS["barge_rate_ohio"]),
    barge_rate_upper_ohio: float = Form(DEFAULTS["barge_rate_upper_ohio"]),
    rail_rate_prb: float = Form(DEFAULTS["rail_rate_prb"]),
):
    """Save Step 2 and advance to Step 3."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Store all pending changes with driver-aligned names
    session.set_pending_change("inventory_beginning_kc", inventory_beginning_kc)
    session.set_pending_change("inventory_beginning_cc", inventory_beginning_cc)
    session.set_pending_change("inventory_target_days", inventory_target_days)
    session.set_pending_change("contracted_deliveries_tons", contracted_deliveries_tons)
    
    session.set_pending_change("coal_price_eastern", coal_price_eastern)
    session.set_pending_change("coal_price_ilb", coal_price_ilb)
    session.set_pending_change("coal_price_prb", coal_price_prb)
    
    session.set_pending_change("coal_blend_eastern_pct", coal_blend_eastern_pct)
    session.set_pending_change("coal_blend_ilb_pct", coal_blend_ilb_pct)
    session.set_pending_change("coal_blend_prb_pct", coal_blend_prb_pct)
    
    session.set_pending_change("coal_btu_eastern", coal_btu_eastern)
    session.set_pending_change("coal_btu_ilb", coal_btu_ilb)
    
    session.set_pending_change("barge_rate_ohio", barge_rate_ohio)
    session.set_pending_change("barge_rate_upper_ohio", barge_rate_upper_ohio)
    session.set_pending_change("rail_rate_prb", rail_rate_prb)
    
    session.mark_step_done(2)
    session.advance_to_step(3)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/3",
        status_code=302
    )


# =============================================================================
# Step 3: Coal Contracts
# =============================================================================

@router.get("/{session_id}/step/3", response_class=HTMLResponse)
async def step3_contracts(request: Request, session_id: str):
    """Render Step 3: Coal Contracts."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(3):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/2", status_code=302)
    
    session.current_step = 3
    update_session(session)
    
    year = session.pending_changes.get("forecast_year",
                                        session.as_of_date.year if session.as_of_date else datetime.now().year)
    
    # Load active contracts from database
    contracts = []
    kc_blended = {"delivered_cost_per_ton": 61, "cost_per_mmbtu": 2.44, "source_breakdown": [{"source": "Default", "type": "default", "pct": 100}]}
    cc_blended = {"delivered_cost_per_ton": 61, "cost_per_mmbtu": 2.44, "source_breakdown": [{"source": "Default", "type": "default", "pct": 100}]}
    kc_uncommitted = {"price_per_ton": 55, "barge_per_ton": 6, "btu_per_lb": 12500}
    cc_uncommitted = {"price_per_ton": 55, "barge_per_ton": 6, "btu_per_lb": 12500}
    
    try:
        with get_db_session() as db:
            from src.models.coal_contract import CoalContract, UncommittedCoalPrice
            from src.engine.coal_supply import get_weighted_coal_pricing
            from datetime import date
            
            period_date = date(year, 1, 1)
            
            # Get active contracts
            active_contracts = db.query(CoalContract).filter(
                CoalContract.is_active == True,
                CoalContract.start_date <= period_date,
                CoalContract.end_date >= period_date,
            ).all()
            
            for c in active_contracts:
                contracts.append({
                    "id": c.id,
                    "contract_id": c.contract_id,
                    "supplier": c.supplier,
                    "plant_name": "Kyger Creek" if c.plant_id == 1 else "Clifty Creek",
                    "annual_tons": float(c.annual_tons),
                    "coal_price_per_ton": float(c.coal_price_per_ton),
                    "barge_price_per_ton": float(c.barge_price_per_ton or 0),
                    "delivered_cost_per_ton": float(c.coal_price_per_ton + (c.barge_price_per_ton or 0)),
                    "btu_per_lb": float(c.btu_per_lb),
                })
            
            # Get weighted pricing for each plant (using first month)
            kc_pricing = get_weighted_coal_pricing(db, 1, year, 1)
            cc_pricing = get_weighted_coal_pricing(db, 2, year, 1)
            
            kc_blended = {
                "delivered_cost_per_ton": float(kc_pricing["delivered_cost_per_ton"]),
                "cost_per_mmbtu": float(kc_pricing["cost_per_mmbtu"]),
                "source_breakdown": kc_pricing["source_breakdown"],
            }
            cc_blended = {
                "delivered_cost_per_ton": float(cc_pricing["delivered_cost_per_ton"]),
                "cost_per_mmbtu": float(cc_pricing["cost_per_mmbtu"]),
                "source_breakdown": cc_pricing["source_breakdown"],
            }
            
            # Get uncommitted pricing
            kc_unc = db.query(UncommittedCoalPrice).filter(
                UncommittedCoalPrice.plant_id == 1,
                UncommittedCoalPrice.year == year,
                UncommittedCoalPrice.month == 1,
            ).first()
            
            cc_unc = db.query(UncommittedCoalPrice).filter(
                UncommittedCoalPrice.plant_id == 2,
                UncommittedCoalPrice.year == year,
                UncommittedCoalPrice.month == 1,
            ).first()
            
            if kc_unc:
                kc_uncommitted = {
                    "price_per_ton": float(kc_unc.price_per_ton),
                    "barge_per_ton": float(kc_unc.barge_per_ton or 0),
                    "btu_per_lb": float(kc_unc.btu_per_lb),
                }
            if cc_unc:
                cc_uncommitted = {
                    "price_per_ton": float(cc_unc.price_per_ton),
                    "barge_per_ton": float(cc_unc.barge_per_ton or 0),
                    "btu_per_lb": float(cc_unc.btu_per_lb),
                }
    except Exception as e:
        print(f"Warning: Could not load contracts: {e}")
    
    context = get_base_context(request, session, 3)
    context.update({
        "year": year,
        "contracts": contracts,
        "kc_blended": kc_blended,
        "cc_blended": cc_blended,
        "kc_uncommitted": kc_uncommitted,
        "cc_uncommitted": cc_uncommitted,
    })
    
    return templates.TemplateResponse("fuel_forecast/step3_contracts.html", context)


@router.post("/{session_id}/step/3", response_class=RedirectResponse)
async def step3_save_contracts(request: Request, session_id: str):
    """Save Step 3 (Coal Contracts) and advance to Step 4.
    
    Uncommitted coal pricing is saved directly to uncommitted_coal_prices table.
    """
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    form_data = await request.form()
    
    year = session.pending_changes.get("forecast_year",
                                        session.as_of_date.year if session.as_of_date else datetime.now().year)
    
    # Save uncommitted coal pricing to database
    try:
        with get_db_session() as db:
            from src.models.coal_contract import UncommittedCoalPrice
            from decimal import Decimal
            
            # Save KC uncommitted for all months
            kc_price = Decimal(str(form_data.get("kc_uncommitted_price", 55)))
            kc_barge = Decimal(str(form_data.get("kc_uncommitted_barge", 6)))
            kc_btu = Decimal(str(form_data.get("kc_uncommitted_btu", 12500)))
            
            for month in range(1, 13):
                existing = db.query(UncommittedCoalPrice).filter(
                    UncommittedCoalPrice.plant_id == 1,
                    UncommittedCoalPrice.year == year,
                    UncommittedCoalPrice.month == month,
                ).first()
                
                if existing:
                    existing.price_per_ton = kc_price
                    existing.barge_per_ton = kc_barge
                    existing.btu_per_lb = kc_btu
                else:
                    db.add(UncommittedCoalPrice(
                        plant_id=1,
                        year=year,
                        month=month,
                        price_per_ton=kc_price,
                        barge_per_ton=kc_barge,
                        btu_per_lb=kc_btu,
                        source_name="NAPP Spot",
                    ))
            
            # Save CC uncommitted for all months
            cc_price = Decimal(str(form_data.get("cc_uncommitted_price", 55)))
            cc_barge = Decimal(str(form_data.get("cc_uncommitted_barge", 6)))
            cc_btu = Decimal(str(form_data.get("cc_uncommitted_btu", 12500)))
            
            for month in range(1, 13):
                existing = db.query(UncommittedCoalPrice).filter(
                    UncommittedCoalPrice.plant_id == 2,
                    UncommittedCoalPrice.year == year,
                    UncommittedCoalPrice.month == month,
                ).first()
                
                if existing:
                    existing.price_per_ton = cc_price
                    existing.barge_per_ton = cc_barge
                    existing.btu_per_lb = cc_btu
                else:
                    db.add(UncommittedCoalPrice(
                        plant_id=2,
                        year=year,
                        month=month,
                        price_per_ton=cc_price,
                        barge_per_ton=cc_barge,
                        btu_per_lb=cc_btu,
                        source_name="NAPP Spot",
                    ))
            
            db.commit()
    except Exception as e:
        print(f"Warning: Could not save uncommitted pricing: {e}")
    
    session.mark_step_done(3)
    session.advance_to_step(4)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/4",
        status_code=302
    )


# =============================================================================
# Step 4: Use Factors
# =============================================================================

MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

OZONE_MONTHS = {5, 6, 7, 8, 9}  # May through September


@router.get("/{session_id}/step/4", response_class=HTMLResponse)
async def step4_use_factors(request: Request, session_id: str):
    """Render Step 4: Use Factors."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(4):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/3", status_code=302)
    
    session.current_step = 4
    update_session(session)
    
    # Get prior values for comparison display
    prior = get_all_prior_values(session)
    has_prior = len(prior) > 0
    
    # Build use factor values from session or defaults
    kc_use_factors = {}
    cc_use_factors = {}
    cc_unit6_ozone_use_factors = {}
    
    for month in range(1, 13):
        kc_use_factors[month] = get_value(session, f"kc_use_factor_{month}") / 100 if session.get_pending_change(f"kc_use_factor_{month}") else 0.85
        cc_use_factors[month] = get_value(session, f"cc_use_factor_{month}") / 100 if session.get_pending_change(f"cc_use_factor_{month}") else 0.85
        # Unit 6 ozone: default 0% during ozone season, else match other units
        if month in OZONE_MONTHS:
            cc_unit6_ozone_use_factors[month] = get_value(session, f"cc_unit6_ozone_{month}") / 100 if session.get_pending_change(f"cc_unit6_ozone_{month}") else 0.0
        else:
            cc_unit6_ozone_use_factors[month] = cc_use_factors[month]
    
    context = get_base_context(request, session, 4)
    context.update({
        "kc_use_factors": kc_use_factors,
        "cc_use_factors": cc_use_factors,
        "cc_unit6_ozone_use_factors": cc_unit6_ozone_use_factors,
        "month_names": MONTH_NAMES,
        "ozone_months": OZONE_MONTHS,
        
        # Prior values for comparison
        "has_prior": has_prior,
        "prior": prior,
        "base_scenario_name": session.base_scenario_name,
    })
    
    return templates.TemplateResponse("fuel_forecast/step2_use_factors.html", context)


@router.post("/{session_id}/step/4", response_class=RedirectResponse)
async def step4_save_use_factors(request: Request, session_id: str):
    """Save Step 4 (Use Factors) and advance to Step 5.
    
    Use factors are saved directly to the use_factor_inputs table.
    No longer stored in session.pending_changes to avoid redundancy.
    """
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Parse form data
    form_data = await request.form()
    
    # Save directly to database (use_factor_inputs table) - this is the source of truth
    year = session.as_of_date.year if session.as_of_date else datetime.now().year
    try:
        with get_db_session() as db:
            from src.models.use_factor import upsert_use_factor
            
            # Save KC use factors
            for month in range(1, 13):
                # form_data.get() returns None if key not found, so use 'or' for default
                base_uf = float(form_data.get(f"kc_use_factor_{month}") or 85) / 100
                upsert_use_factor(
                    db,
                    plant_id=1,  # Kyger Creek
                    year=year,
                    month=month,
                    use_factor_base=base_uf,
                    use_factor_ozone_non_scr=base_uf,  # KC has SCR on all units
                )
            
            # Save CC use factors
            for month in range(1, 13):
                base_uf = float(form_data.get(f"cc_use_factor_{month}") or 85) / 100
                ozone_uf = float(form_data.get(f"cc_unit6_ozone_use_factor_{month}") or 0) / 100
                upsert_use_factor(
                    db,
                    plant_id=2,  # Clifty Creek
                    year=year,
                    month=month,
                    use_factor_base=base_uf,
                    use_factor_ozone_non_scr=ozone_uf if month in OZONE_MONTHS else base_uf,
                )
    except Exception as e:
        # Log but don't fail the workflow
        print(f"Warning: Could not save use factors to database: {e}")
    
    # Store year in pending changes for later steps to reference
    session.set_pending_change("forecast_year", year)
    
    session.mark_step_done(4)
    session.advance_to_step(5)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/5",
        status_code=302
    )


# =============================================================================
# Step 5: Heat Rates
# =============================================================================

@router.get("/{session_id}/step/5", response_class=HTMLResponse)
async def step5_heat_rates(request: Request, session_id: str):
    """Render Step 5: Heat Rates."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(5):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/4", status_code=302)
    
    session.current_step = 5
    update_session(session)
    
    # Get prior values for comparison display
    prior = get_all_prior_values(session)
    has_prior = len(prior) > 0
    
    # Build heat rate values from session or defaults
    kc_heat_rates = {}
    cc_heat_rates = {}
    
    for month in range(1, 13):
        kc_heat_rates[month] = {
            "baseline_heat_rate": get_value(session, f"kc_hr_{month}") if session.get_pending_change(f"kc_hr_{month}") else get_value(session, "heat_rate_baseline_kc")
        }
        cc_heat_rates[month] = {
            "baseline_heat_rate": get_value(session, f"cc_hr_{month}") if session.get_pending_change(f"cc_hr_{month}") else get_value(session, "heat_rate_baseline_cc")
        }
    
    # Annual adjustment values
    kc_heat_rates_annual = {
        "suf_correction": get_value(session, "heat_rate_suf_correction"),
        "prb_blend_adjustment": get_value(session, "heat_rate_prb_penalty"),
    }
    cc_heat_rates_annual = {
        "suf_correction": get_value(session, "heat_rate_suf_correction"),
        "prb_blend_adjustment": get_value(session, "heat_rate_prb_penalty"),
    }
    
    # Note: Template step3_heat_rates.html is used for workflow Step 5 (Heat Rates)
    # Build kyger_heat_rate and clifty_heat_rate objects for template compatibility
    kyger_heat_rate = {
        "baseline": get_value(session, "heat_rate_baseline_kc") or 9850,
        "min_load": get_value(session, "heat_rate_min_load_kc") or 10500,
        "suf_correction": kc_heat_rates_annual.get("suf_correction", 0),
        "prb_adjustment": kc_heat_rates_annual.get("prb_blend_adjustment", 0),
    }
    clifty_heat_rate = {
        "baseline": get_value(session, "heat_rate_baseline_cc") or 9850,
        "min_load": get_value(session, "heat_rate_min_load_cc") or 10500,
        "suf_correction": cc_heat_rates_annual.get("suf_correction", 0),
        "prb_adjustment": cc_heat_rates_annual.get("prb_blend_adjustment", 0),
    }
    
    context = get_base_context(request, session, 5)
    context.update({
        "kc_heat_rates": kc_heat_rates,
        "cc_heat_rates": cc_heat_rates,
        "kc_heat_rates_annual": kc_heat_rates_annual,
        "cc_heat_rates_annual": cc_heat_rates_annual,
        "kyger_heat_rate": kyger_heat_rate,
        "clifty_heat_rate": clifty_heat_rate,
        "month_names": MONTH_NAMES,
        
        # Prior values for comparison
        "has_prior": has_prior,
        "prior": prior,
        "base_scenario_name": session.base_scenario_name,
    })
    
    return templates.TemplateResponse("fuel_forecast/step3_heat_rates.html", context)


@router.post("/{session_id}/step/5", response_class=RedirectResponse)
async def step5_save_heat_rates(request: Request, session_id: str):
    """Save Step 5 (Heat Rates) and advance to Step 6.
    
    Heat rates are saved directly to the heat_rate_inputs table.
    No longer stored in session.pending_changes to avoid redundancy.
    """
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Parse form data
    form_data = await request.form()
    
    # Save directly to database (heat_rate_inputs table) - this is the source of truth
    year = session.pending_changes.get("forecast_year", 
                                        session.as_of_date.year if session.as_of_date else datetime.now().year)
    try:
        with get_db_session() as db:
            from src.models.heat_rate import upsert_heat_rate
            
            # Save KC heat rates
            suf_kc = float(form_data.get("kc_hr_suf_correction", 0))
            prb_kc = float(form_data.get("kc_hr_prb_adjustment", 0))
            for month in range(1, 13):
                baseline = float(form_data.get(f"kc_hr_baseline_{month}", 9850))
                upsert_heat_rate(
                    db,
                    plant_id=1,  # Kyger Creek
                    year=year,
                    month=month,
                    baseline_heat_rate=baseline,
                    suf_correction=suf_kc,
                    prb_blend_adjustment=prb_kc,
                )
            
            # Save CC heat rates
            suf_cc = float(form_data.get("cc_hr_suf_correction", 0))
            prb_cc = float(form_data.get("cc_hr_prb_adjustment", 0))
            for month in range(1, 13):
                baseline = float(form_data.get(f"cc_hr_baseline_{month}", 9900))
                upsert_heat_rate(
                    db,
                    plant_id=2,  # Clifty Creek
                    year=year,
                    month=month,
                    baseline_heat_rate=baseline,
                    suf_correction=suf_cc,
                    prb_blend_adjustment=prb_cc,
                )
    except Exception as e:
        # Log but don't fail the workflow
        print(f"Warning: Could not save heat rates to database: {e}")
    
    session.mark_step_done(5)
    session.advance_to_step(6)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/6",
        status_code=302
    )


# =============================================================================
# Step 6: Generation
# =============================================================================

@router.get("/{session_id}/step/6", response_class=HTMLResponse)
async def step6_generation(request: Request, session_id: str):
    """Render Step 6: Generation Profile."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(6):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/5", status_code=302)
    
    session.current_step = 6
    update_session(session)
    
    # Get prior values for comparison display
    prior = get_all_prior_values(session)
    has_prior = len(prior) > 0
    
    context = get_base_context(request, session, 6)
    context.update({
        # Capacity
        "capacity_mw_kc": get_value(session, "capacity_mw_kc"),
        "capacity_mw_cc": get_value(session, "capacity_mw_cc"),
        
        # Use Factor (summary from step 4)
        "use_factor": get_value(session, "use_factor"),
        "capacity_factor_target": get_value(session, "capacity_factor_target"),
        
        # Heat Rates (summary from step 5)
        "heat_rate_baseline_kc": get_value(session, "heat_rate_baseline_kc"),
        "heat_rate_baseline_cc": get_value(session, "heat_rate_baseline_cc"),
        "heat_rate_suf_correction": get_value(session, "heat_rate_suf_correction"),
        "heat_rate_prb_penalty": get_value(session, "heat_rate_prb_penalty"),
        
        # Outages
        "outage_days_planned": get_value(session, "outage_days_planned"),
        "outage_days_forced": get_value(session, "outage_days_forced"),
        
        # Deductions
        "fgd_aux_pct": get_value(session, "fgd_aux_pct"),
        "gsu_loss_pct": get_value(session, "gsu_loss_pct"),
        "reserve_mw": get_value(session, "reserve_mw"),
        
        # Prior values for comparison
        "has_prior": has_prior,
        "prior": prior,
        "base_scenario_name": session.base_scenario_name,
    })
    
    return templates.TemplateResponse("fuel_forecast/step3_generation.html", context)


@router.post("/{session_id}/step/6", response_class=RedirectResponse)
async def step6_save(
    request: Request,
    session_id: str,
    # Capacity - kept in pending_changes (plant-level config)
    capacity_mw_kc: float = Form(DEFAULTS["capacity_mw_kc"]),
    capacity_mw_cc: float = Form(DEFAULTS["capacity_mw_cc"]),
    # Use Factor - summary only, detailed values are in use_factor_inputs table
    use_factor: float = Form(DEFAULTS["use_factor"]),
    capacity_factor_target: float = Form(DEFAULTS["capacity_factor_target"]),
    # Heat Rates - summary only, detailed values are in heat_rate_inputs table
    heat_rate_baseline_kc: float = Form(DEFAULTS["heat_rate_baseline_kc"]),
    heat_rate_baseline_cc: float = Form(DEFAULTS["heat_rate_baseline_cc"]),
    heat_rate_suf_correction: float = Form(DEFAULTS["heat_rate_suf_correction"]),
    heat_rate_prb_penalty: float = Form(DEFAULTS["heat_rate_prb_penalty"]),
    # Outages - summary only, detailed values are in unit_outage_inputs table
    outage_days_planned: float = Form(DEFAULTS["outage_days_planned"]),
    outage_days_forced: float = Form(DEFAULTS["outage_days_forced"]),
    # Deductions - kept in pending_changes (simple scalar values)
    fgd_aux_pct: float = Form(DEFAULTS["fgd_aux_pct"]),
    gsu_loss_pct: float = Form(DEFAULTS["gsu_loss_pct"]),
    reserve_mw: float = Form(DEFAULTS["reserve_mw"]),
):
    """Save Step 6 and advance to Step 7.
    
    Note: Use factors, heat rates, and outages are now stored in their respective
    input tables (use_factor_inputs, heat_rate_inputs, unit_outage_inputs).
    The summary values here are for display purposes only in the review step.
    The actual calculations use data from the input tables.
    """
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Capacity and deductions are kept in pending_changes (simple plant-level values)
    session.set_pending_change("capacity_mw_kc", capacity_mw_kc)
    session.set_pending_change("capacity_mw_cc", capacity_mw_cc)
    session.set_pending_change("capacity_factor_target", capacity_factor_target)
    
    session.set_pending_change("fgd_aux_pct", fgd_aux_pct)
    session.set_pending_change("gsu_loss_pct", gsu_loss_pct)
    session.set_pending_change("reserve_mw", reserve_mw)
    
    # Note: use_factor, heat_rate_*, outage_days_* are NOT stored in pending_changes
    # They come from the input tables (use_factor_inputs, heat_rate_inputs, unit_outage_inputs)
    # We'll read them from the database in the review step
    
    session.mark_step_done(6)
    session.advance_to_step(7)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/7",
        status_code=302
    )


# =============================================================================
# Step 7: Other Costs
# =============================================================================

@router.get("/{session_id}/step/7", response_class=HTMLResponse)
async def step7_other(request: Request, session_id: str):
    """Render Step 7: Other Costs."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(7):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/6", status_code=302)
    
    session.current_step = 7
    update_session(session)
    
    # Get prior values for comparison display
    prior = get_all_prior_values(session)
    has_prior = len(prior) > 0
    
    context = get_base_context(request, session, 7)
    context.update({
        # Escalation
        "escalation_coal_annual": get_value(session, "escalation_coal_annual"),
        "escalation_transport_annual": get_value(session, "escalation_transport_annual"),
        "escalation_reagent_annual": get_value(session, "escalation_reagent_annual"),
        
        # Keep prior flags
        "keep_reagents": session.keep_prior.get("reagents", True),
        "keep_byproducts": session.keep_prior.get("byproducts", True),
        
        # Reagent prices
        "urea_price_per_ton": get_value(session, "urea_price_per_ton"),
        "limestone_price_per_ton": get_value(session, "limestone_price_per_ton"),
        "hydrated_lime_price_per_ton": get_value(session, "hydrated_lime_price_per_ton"),
        "mercury_reagent_cost_monthly": get_value(session, "mercury_reagent_cost_monthly"),
        
        # Byproduct prices
        "ash_sale_price_per_ton": get_value(session, "ash_sale_price_per_ton"),
        "ash_disposal_cost_per_ton": get_value(session, "ash_disposal_cost_per_ton"),
        "gypsum_sale_price_per_ton": get_value(session, "gypsum_sale_price_per_ton"),
        "fly_ash_sale_pct": get_value(session, "fly_ash_sale_pct"),
        
        # Prior values for comparison
        "has_prior": has_prior,
        "prior": prior,
        "base_scenario_name": session.base_scenario_name,
    })
    
    return templates.TemplateResponse("fuel_forecast/step4_other.html", context)


@router.post("/{session_id}/step/7", response_class=RedirectResponse)
async def step7_save(
    request: Request,
    session_id: str,
    # Escalation
    escalation_coal_annual: float = Form(DEFAULTS["escalation_coal_annual"]),
    escalation_transport_annual: float = Form(DEFAULTS["escalation_transport_annual"]),
    escalation_reagent_annual: float = Form(DEFAULTS["escalation_reagent_annual"]),
    # Keep prior flags
    keep_reagents: Optional[str] = Form(None),
    keep_byproducts: Optional[str] = Form(None),
    # Reagent prices
    urea_price_per_ton: float = Form(DEFAULTS["urea_price_per_ton"]),
    limestone_price_per_ton: float = Form(DEFAULTS["limestone_price_per_ton"]),
    hydrated_lime_price_per_ton: float = Form(DEFAULTS["hydrated_lime_price_per_ton"]),
    mercury_reagent_cost_monthly: float = Form(DEFAULTS["mercury_reagent_cost_monthly"]),
    # Byproduct prices
    ash_sale_price_per_ton: float = Form(DEFAULTS["ash_sale_price_per_ton"]),
    ash_disposal_cost_per_ton: float = Form(DEFAULTS["ash_disposal_cost_per_ton"]),
    gypsum_sale_price_per_ton: float = Form(DEFAULTS["gypsum_sale_price_per_ton"]),
    fly_ash_sale_pct: float = Form(DEFAULTS["fly_ash_sale_pct"]),
):
    """Save Step 7 and advance to Step 8."""
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Escalation
    session.set_pending_change("escalation_coal_annual", escalation_coal_annual)
    session.set_pending_change("escalation_transport_annual", escalation_transport_annual)
    session.set_pending_change("escalation_reagent_annual", escalation_reagent_annual)
    
    # Keep prior flags (checkbox sends value only when checked)
    session.set_keep_prior("reagents", keep_reagents is not None)
    session.set_keep_prior("byproducts", keep_byproducts is not None)
    
    # Only save reagent/byproduct values if not keeping prior
    if not session.keep_prior.get("reagents", True):
        session.set_pending_change("urea_price_per_ton", urea_price_per_ton)
        session.set_pending_change("limestone_price_per_ton", limestone_price_per_ton)
        session.set_pending_change("hydrated_lime_price_per_ton", hydrated_lime_price_per_ton)
        session.set_pending_change("mercury_reagent_cost_monthly", mercury_reagent_cost_monthly)
    
    if not session.keep_prior.get("byproducts", True):
        session.set_pending_change("ash_sale_price_per_ton", ash_sale_price_per_ton)
        session.set_pending_change("ash_disposal_cost_per_ton", ash_disposal_cost_per_ton)
        session.set_pending_change("gypsum_sale_price_per_ton", gypsum_sale_price_per_ton)
        session.set_pending_change("fly_ash_sale_pct", fly_ash_sale_pct)
    
    session.mark_step_done(7)
    session.advance_to_step(8)
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/8",
        status_code=302
    )


# =============================================================================
# Step 8: Review & Save
# =============================================================================

@router.get("/{session_id}/step/8", response_class=HTMLResponse)
async def step8_review(request: Request, session_id: str):
    """Render Step 8: Review & Save.
    
    Reads use factors, heat rates, and outages from database input tables.
    Other values are read from pending_changes.
    """
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if not session.can_proceed_to_step(8):
        return RedirectResponse(url=f"/fuel-forecast/{session_id}/step/7", status_code=302)
    
    session.current_step = 8
    update_session(session)
    
    # Get forecast year from pending changes or session date
    year = session.pending_changes.get("forecast_year",
                                        session.as_of_date.year if session.as_of_date else datetime.now().year)
    
    # Load use factors, heat rates, and outages from database input tables
    use_factor_summary = {"kc_avg": 85.0, "cc_avg": 85.0, "unit6_ozone_avg": 0.0}
    heat_rate_summary = {"kc_baseline": 9850, "cc_baseline": 9900, "suf_correction": 0, "prb_penalty": 0}
    outage_summary = {"kc_planned": 0.0, "kc_forced": 0.0, "cc_planned": 0.0, "cc_forced": 0.0}
    
    try:
        with get_db_session() as db:
            from src.models.use_factor import get_use_factors_for_year
            from src.models.heat_rate import get_heat_rates_for_year
            from src.models.unit_outage import get_unit_outages_for_year
            
            # Load use factor averages
            kc_use_factors = get_use_factors_for_year(db, 1, year)
            cc_use_factors = get_use_factors_for_year(db, 2, year)
            
            if kc_use_factors:
                use_factor_summary["kc_avg"] = sum(v["base"] * 100 for v in kc_use_factors.values()) / len(kc_use_factors)
            if cc_use_factors:
                use_factor_summary["cc_avg"] = sum(v["base"] * 100 for v in cc_use_factors.values()) / len(cc_use_factors)
                # Calculate Unit 6 ozone average for months 5-9
                ozone_values = [v["ozone_non_scr"] * 100 for m, v in cc_use_factors.items() if m in {5, 6, 7, 8, 9}]
                if ozone_values:
                    use_factor_summary["unit6_ozone_avg"] = sum(ozone_values) / len(ozone_values)
            
            # Load heat rate averages
            # get_heat_rates_for_year returns: {month: {"plant": {...}, "units": {...}}}
            kc_heat_rates = get_heat_rates_for_year(db, 1, year)
            cc_heat_rates = get_heat_rates_for_year(db, 2, year)
            
            if kc_heat_rates:
                hr_values = [kc_heat_rates[m]["plant"]["baseline_heat_rate"] for m in kc_heat_rates]
                heat_rate_summary["kc_baseline"] = float(sum(hr_values) / len(hr_values)) if hr_values else 9850
                heat_rate_summary["suf_correction"] = float(kc_heat_rates[1]["plant"]["suf_correction"] or 0)
            if cc_heat_rates:
                hr_values = [cc_heat_rates[m]["plant"]["baseline_heat_rate"] for m in cc_heat_rates]
                heat_rate_summary["cc_baseline"] = float(sum(hr_values) / len(hr_values)) if hr_values else 9850
                heat_rate_summary["prb_penalty"] = float(cc_heat_rates[1]["plant"]["prb_blend_adjustment"] or 0)
            
            # Load outage summaries
            kc_outages = get_unit_outages_for_year(db, 1, year)
            cc_outages = get_unit_outages_for_year(db, 2, year)
            
            if kc_outages:
                outage_summary["kc_planned"] = sum(float(o.planned_outage_days) for o in kc_outages)
                outage_summary["kc_forced"] = sum(float(o.forced_outage_days) for o in kc_outages)
            if cc_outages:
                outage_summary["cc_planned"] = sum(float(o.planned_outage_days) for o in cc_outages)
                outage_summary["cc_forced"] = sum(float(o.forced_outage_days) for o in cc_outages)
    except Exception as e:
        print(f"Warning: Could not load input table summaries: {e}")
    
    # Build comprehensive summary of all inputs
    summary = {
        "start_mode": session.start_mode,
        "base_scenario": session.base_scenario_name or "None",
        "as_of_date": session.as_of_date.strftime("%B %Y") if session.as_of_date else "Not set",
        
        "inventory": {
            "kc_beginning": get_value(session, "inventory_beginning_kc"),
            "cc_beginning": get_value(session, "inventory_beginning_cc"),
            "target_days": get_value(session, "inventory_target_days"),
            "contracted_tons": get_value(session, "contracted_deliveries_tons"),
        },
        
        "coal_prices": {
            "eastern": get_value(session, "coal_price_eastern"),
            "ilb": get_value(session, "coal_price_ilb"),
            "prb": get_value(session, "coal_price_prb"),
        },
        
        "coal_blend": {
            "eastern_pct": get_value(session, "coal_blend_eastern_pct"),
            "ilb_pct": get_value(session, "coal_blend_ilb_pct"),
            "prb_pct": get_value(session, "coal_blend_prb_pct"),
        },
        
        "transportation": {
            "barge_ohio": get_value(session, "barge_rate_ohio"),
            "barge_upper_ohio": get_value(session, "barge_rate_upper_ohio"),
            "rail_prb": get_value(session, "rail_rate_prb"),
        },
        
        # Generation: capacity from pending_changes, use factors from DB
        "generation": {
            "capacity_kc": get_value(session, "capacity_mw_kc"),
            "capacity_cc": get_value(session, "capacity_mw_cc"),
            "use_factor_kc": use_factor_summary["kc_avg"],
            "use_factor_cc": use_factor_summary["cc_avg"],
            "unit6_ozone_factor": use_factor_summary["unit6_ozone_avg"],
            "capacity_factor_target": get_value(session, "capacity_factor_target"),
        },
        
        # Heat rates from DB input table
        "heat_rates": {
            "kc_baseline": heat_rate_summary["kc_baseline"],
            "cc_baseline": heat_rate_summary["cc_baseline"],
            "suf_correction": heat_rate_summary["suf_correction"],
            "prb_penalty": heat_rate_summary["prb_penalty"],
        },
        
        # Outages from DB input table
        "outages": {
            "kc_planned_days": outage_summary["kc_planned"],
            "kc_forced_days": outage_summary["kc_forced"],
            "cc_planned_days": outage_summary["cc_planned"],
            "cc_forced_days": outage_summary["cc_forced"],
        },
        
        "deductions": {
            "fgd_aux_pct": get_value(session, "fgd_aux_pct"),
            "gsu_loss_pct": get_value(session, "gsu_loss_pct"),
            "reserve_mw": get_value(session, "reserve_mw"),
        },
        
        "escalation": {
            "coal": get_value(session, "escalation_coal_annual"),
            "transport": get_value(session, "escalation_transport_annual"),
            "reagent": get_value(session, "escalation_reagent_annual"),
        },
        
        "keep_prior": {
            "reagents": session.keep_prior.get("reagents", True),
            "byproducts": session.keep_prior.get("byproducts", True),
        },
    }
    
    context = get_base_context(request, session, 8)
    context.update({
        "summary": summary,
        "change_notes": session.change_notes,
    })
    
    return templates.TemplateResponse("fuel_forecast/step5_review.html", context)


@router.post("/{session_id}/step/8", response_class=HTMLResponse)
async def step8_calculate(
    request: Request,
    session_id: str,
    scenario_name: str = Form(""),
    notes: str = Form(""),
    action: str = Form("calculate"),
):
    """Calculate forecast and optionally save."""
    import json
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Determine forecast year from as_of_date
    year = session.as_of_date.year if session.as_of_date else datetime.now().year
    
    # Calculate the forecast using the fuel model engine
    results, monthly_results = calculate_forecast_from_session(session, year)
    
    session.mark_step_done(8)
    update_session(session)
    
    context = get_base_context(request, session, 8)
    context.update({
        "results": results,
        "monthly_results": monthly_results,
        "scenario_name": scenario_name or f"{datetime.now().strftime('%B %Y')} Forecast",
        "notes": notes,
        "calculated": True,
        "year": year,
    })
    
    # Build prior comparison and waterfall if base scenario exists
    prior_comparison = None
    waterfall_data = []
    
    if session.base_scenario_id:
        # Get prior values from session
        prior = get_all_prior_values(session)
        
        # Estimate prior cost based on prior driver values (simplified)
        prior_coal_price = prior.get("coal_price_eastern", 55)
        current_coal_price = get_value(session, "coal_price_eastern")
        prior_use_factor = prior.get("use_factor", 85) / 100
        current_use_factor = get_value(session, "use_factor") / 100
        
        # Calculate estimated prior cost using same logic
        base_generation = 8760 * 2050 * 0.70
        base_coal_per_mwh = 0.45
        transport_factor = 1.1
        
        prior_gen = base_generation * prior_use_factor
        prior_coal_cost = prior_gen * base_coal_per_mwh * prior_coal_price * transport_factor
        prior_cost = prior_coal_cost
        
        current_cost = results["total_cost"]
        variance = current_cost - prior_cost
        
        prior_comparison = {
            "prior_cost": prior_cost,
            "current_cost": current_cost,
            "variance": variance,
            "variance_pct": (variance / prior_cost * 100) if prior_cost > 0 else 0,
        }
        
        # Build waterfall data
        waterfall_data = [
            {"label": "Prior Forecast", "value": prior_cost},
        ]
        
        # Coal price impact
        coal_impact = (current_coal_price - prior_coal_price) * prior_gen * base_coal_per_mwh * transport_factor
        if abs(coal_impact) > 100000:
            waterfall_data.append({"label": "Coal Price", "value": coal_impact})
        
        # Generation impact
        gen_impact = (current_use_factor - prior_use_factor) * base_generation * base_coal_per_mwh * current_coal_price * transport_factor
        if abs(gen_impact) > 100000:
            waterfall_data.append({"label": "Generation", "value": gen_impact})
        
        # Other impacts (residual)
        explained = sum(d["value"] for d in waterfall_data[1:]) if len(waterfall_data) > 1 else 0
        other = variance - explained
        if abs(other) > 100000:
            waterfall_data.append({"label": "Other", "value": other})
        
        waterfall_data.append({"label": "Current Forecast", "value": current_cost})
        
        context["prior_comparison"] = prior_comparison
        context["waterfall_data"] = json.dumps(waterfall_data)
    
    # If saving, handle the save action
    if action == "save" and scenario_name:
        save_result = save_forecast_scenario(session, scenario_name, notes, year)
        context["saved"] = True
        context["saved_scenario_id"] = save_result.get("scenario_id")
    
    return templates.TemplateResponse("fuel_forecast/step5_review.html", context)


# =============================================================================
# Multi-Year Projection
# =============================================================================

@router.get("/{session_id}/multi-year", response_class=HTMLResponse)
async def multi_year_projection(request: Request, session_id: str, end_year: int = 2040):
    """Render multi-year fuel cost projection view."""
    import json
    from src.engine.projections import project_multi_year, AnnualProjection
    from src.engine.default_drivers import create_default_fuel_model
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get base year from session
    base_year = session.as_of_date.year if session.as_of_date else datetime.now().year
    
    # Create a FuelModel from session values
    model = create_default_fuel_model()
    
    # Apply session pending changes to the model
    for driver_name, value in session.pending_changes.items():
        if driver_name.startswith("_"):
            continue
        if driver_name in model.drivers:
            try:
                model.set_driver_value(driver_name, base_year, None, Decimal(str(value)))
            except (ValueError, TypeError):
                pass
    
    # Handle plant-specific values
    plant_specific = [
        ("capacity_mw_kc", "capacity_mw", 1),
        ("capacity_mw_cc", "capacity_mw", 2),
        ("heat_rate_baseline_kc", "heat_rate_baseline", 1),
        ("heat_rate_baseline_cc", "heat_rate_baseline", 2),
        ("inventory_beginning_kc", "inventory_beginning_tons", 1),
        ("inventory_beginning_cc", "inventory_beginning_tons", 2),
    ]
    
    for ui_name, driver_name, plant_id in plant_specific:
        if ui_name in session.pending_changes:
            try:
                model.set_driver_value(driver_name, base_year, None, 
                                       Decimal(str(session.pending_changes[ui_name])), plant_id=plant_id)
            except (ValueError, TypeError):
                pass
    
    # Run projection
    with get_db_session() as db:
        projection = project_multi_year(db, model, base_year, end_year)
    
    # Build projection data with percent change
    projections = []
    prev_cost = None
    
    for sys_proj in projection.system_projections:
        pct_change = None
        if prev_cost is not None and prev_cost > 0:
            pct_change = ((sys_proj.total_fuel_cost - prev_cost) / prev_cost) * 100
        
        projections.append({
            "year": sys_proj.year,
            "total_fuel_cost": sys_proj.total_fuel_cost,
            "total_mwh": sys_proj.total_mwh,
            "avg_fuel_cost_per_mwh": sys_proj.avg_fuel_cost_per_mwh,
            "coal_price_eastern": sys_proj.coal_price_eastern,
            "barge_rate": sys_proj.barge_rate,
            "pct_change": pct_change,
        })
        prev_cost = sys_proj.total_fuel_cost
    
    # Calculate summary metrics
    total_cost = sum(p["total_fuel_cost"] for p in projections)
    total_mwh = sum(p["total_mwh"] for p in projections)
    avg_cost_per_mwh = total_cost / total_mwh if total_mwh > 0 else 0
    
    # Calculate CAGR
    if len(projections) >= 2 and projections[0]["total_fuel_cost"] > 0:
        years = end_year - base_year
        cagr = ((projections[-1]["total_fuel_cost"] / projections[0]["total_fuel_cost"]) ** (1/years) - 1) * 100
    else:
        cagr = 0
    
    context = get_base_context(request, session, 7)
    context.update({
        "base_year": base_year,
        "end_year": end_year,
        "projections": projections,
        "projections_json": json.dumps(projections),
        "total_cost": total_cost,
        "total_mwh": total_mwh,
        "avg_cost_per_mwh": avg_cost_per_mwh,
        "cagr": cagr,
        "escalation": {
            "coal": get_value(session, "escalation_coal_annual"),
            "transport": get_value(session, "escalation_transport_annual"),
            "reagent": get_value(session, "escalation_reagent_annual"),
        },
        "scenario_name": session.base_scenario_name or "Current Forecast",
    })
    
    return templates.TemplateResponse("fuel_forecast/multi_year.html", context)


@router.get("/{session_id}/multi-year/export")
async def export_multi_year(session_id: str, format: str = "excel"):
    """Export multi-year projection to Excel."""
    from fastapi.responses import FileResponse, StreamingResponse
    from src.engine.projections import project_multi_year
    from src.engine.default_drivers import create_default_fuel_model
    import io
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    base_year = session.as_of_date.year if session.as_of_date else datetime.now().year
    end_year = 2040
    
    # Create model and run projection
    model = create_default_fuel_model()
    for driver_name, value in session.pending_changes.items():
        if driver_name.startswith("_"):
            continue
        if driver_name in model.drivers:
            try:
                model.set_driver_value(driver_name, base_year, None, Decimal(str(value)))
            except (ValueError, TypeError):
                pass
    
    with get_db_session() as db:
        projection = project_multi_year(db, model, base_year, end_year)
    
    # Create Excel file
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Multi-Year Projection"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"OVEC Fuel Cost Projection {base_year}-{end_year}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ["Year", "Total Cost ($M)", "Generation (GWh)", "$/MWh", "% Change", "Coal Price", "Barge Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        prev_cost = None
        for row, proj in enumerate(projection.system_projections, 4):
            pct_change = ((proj.total_fuel_cost - prev_cost) / prev_cost * 100) if prev_cost else None
            
            ws.cell(row=row, column=1, value=proj.year).border = thin_border
            ws.cell(row=row, column=2, value=round(proj.total_fuel_cost / 1000000, 1)).border = thin_border
            ws.cell(row=row, column=3, value=round(proj.total_mwh / 1000, 0)).border = thin_border
            ws.cell(row=row, column=4, value=round(proj.avg_fuel_cost_per_mwh, 2)).border = thin_border
            ws.cell(row=row, column=5, value=f"{pct_change:+.1f}%" if pct_change else "—").border = thin_border
            ws.cell(row=row, column=6, value=round(proj.coal_price_eastern, 2)).border = thin_border
            ws.cell(row=row, column=7, value=round(proj.barge_rate, 2)).border = thin_border
            
            prev_cost = proj.total_fuel_cost
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"OVEC_Projection_{base_year}_{end_year}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed for Excel export")


# =============================================================================
# Scenario Comparison
# =============================================================================

# Driver display labels for comparison
DRIVER_LABELS = {
    "coal_price_eastern": ("Coal Prices", "Eastern (NAPP)"),
    "coal_price_ilb": ("Coal Prices", "Illinois Basin"),
    "coal_price_prb": ("Coal Prices", "PRB"),
    "coal_blend_eastern_pct": ("Coal Blend", "Eastern %"),
    "coal_blend_ilb_pct": ("Coal Blend", "ILB %"),
    "coal_blend_prb_pct": ("Coal Blend", "PRB %"),
    "barge_rate_ohio": ("Transportation", "Ohio River Barge"),
    "barge_rate_upper_ohio": ("Transportation", "Upper Ohio Barge"),
    "rail_rate_prb": ("Transportation", "PRB Rail"),
    "use_factor": ("Generation", "Use Factor"),
    "capacity_factor_target": ("Generation", "Target Capacity Factor"),
    "heat_rate_baseline": ("Heat Rates", "Baseline Heat Rate"),
    "heat_rate_suf_correction": ("Heat Rates", "SUF Correction"),
    "escalation_coal_annual": ("Escalation", "Coal Escalation"),
    "escalation_transport_annual": ("Escalation", "Transport Escalation"),
    "escalation_reagent_annual": ("Escalation", "Reagent Escalation"),
}


def format_driver_value(driver_name: str, value) -> str:
    """Format a driver value for display."""
    if value is None:
        return "—"
    
    try:
        val = float(value)
    except (ValueError, TypeError):
        return str(value)
    
    # Format based on driver type
    if "price" in driver_name or "rate" in driver_name or "cost" in driver_name:
        return f"${val:,.2f}"
    elif "pct" in driver_name or "factor" in driver_name:
        return f"{val:.1f}%"
    elif "mw" in driver_name.lower():
        return f"{val:,.0f} MW"
    elif "tons" in driver_name or "inventory" in driver_name:
        return f"{val:,.0f} tons"
    elif "days" in driver_name:
        return f"{val:.0f} days"
    else:
        return f"{val:,.2f}"


@router.get("/{session_id}/compare", response_class=HTMLResponse)
async def compare_scenarios(
    request: Request, 
    session_id: str,
    scenario_a: Optional[int] = None,
    scenario_b: Optional[int] = None,
):
    """Render scenario comparison view."""
    import json
    from src.models.scenario import Scenario
    from src.engine.scenario_drivers import compare_scenario_drivers, export_scenario_drivers
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    year = session.as_of_date.year if session.as_of_date else datetime.now().year
    
    # Get available scenarios
    with get_db_session() as db:
        scenarios = db.query(Scenario).filter(
            Scenario.is_active == True
        ).order_by(Scenario.created_at.desc()).limit(50).all()
        
        available_scenarios = [
            {"id": s.id, "name": s.name, "created_at": s.created_at}
            for s in scenarios
        ]
        
        comparison = None
        scenario_a_obj = None
        scenario_b_obj = None
        waterfall_data = []
        
        if scenario_a and scenario_b:
            # Get scenario objects
            scenario_a_obj = db.query(Scenario).get(scenario_a)
            scenario_b_obj = db.query(Scenario).get(scenario_b)
            
            if scenario_a_obj and scenario_b_obj:
                # Get driver comparison
                driver_comparison = compare_scenario_drivers(db, scenario_a, scenario_b, year)
                
                # Get full driver sets for calculation
                set_a = export_scenario_drivers(db, scenario_a, year)
                set_b = export_scenario_drivers(db, scenario_b, year)
                
                # Calculate aggregate metrics (simplified - use driver values)
                # For full calculation, we'd run calculate_forecast_from_session for each
                # Here we use placeholder calculations based on key drivers
                
                # Extract key values for comparison metrics
                def get_driver_val(driver_set, name, default=0):
                    val = driver_set.driver_values.get(name, {})
                    if isinstance(val, dict):
                        return float(val.get("annual", default))
                    return float(val) if val else default
                
                # Approximate costs based on drivers (simplified model)
                coal_price_a = get_driver_val(set_a, "coal_price_eastern", 55)
                coal_price_b = get_driver_val(set_b, "coal_price_eastern", 55)
                use_factor_a = get_driver_val(set_a, "use_factor", 85) / 100
                use_factor_b = get_driver_val(set_b, "use_factor", 85) / 100
                
                # Very rough estimates for demo purposes
                # In production, this would run the full fuel model
                base_generation = 8760 * 2050 * 0.70  # ~12.5M MWh
                base_coal_per_mwh = 0.45  # tons/MWh
                
                gen_a = base_generation * use_factor_a
                gen_b = base_generation * use_factor_b
                coal_cost_a = gen_a * base_coal_per_mwh * coal_price_a
                coal_cost_b = gen_b * base_coal_per_mwh * coal_price_b
                
                # Add transport and other costs
                transport_factor = 1.1
                total_a = coal_cost_a * transport_factor
                total_b = coal_cost_b * transport_factor
                
                comparison = {
                    "a": {
                        "total_cost": total_a,
                        "total_mwh": gen_a,
                        "cost_per_mwh": total_a / gen_a if gen_a > 0 else 0,
                    },
                    "b": {
                        "total_cost": total_b,
                        "total_mwh": gen_b,
                        "cost_per_mwh": total_b / gen_b if gen_b > 0 else 0,
                    },
                    "variance": {
                        "total_cost": total_b - total_a,
                        "total_cost_pct": ((total_b - total_a) / total_a * 100) if total_a > 0 else 0,
                        "total_mwh": gen_b - gen_a,
                        "cost_per_mwh": (total_b / gen_b if gen_b > 0 else 0) - (total_a / gen_a if gen_a > 0 else 0),
                    },
                    "different_count": driver_comparison["different_count"],
                    "differences": [],
                }
                
                # Format differences for display
                for diff in driver_comparison["differences"]:
                    driver_name = diff["driver"]
                    category, label = DRIVER_LABELS.get(driver_name, ("Other", driver_name))
                    
                    val_a = diff["scenario_a"]
                    val_b = diff["scenario_b"]
                    
                    # Handle dict values (monthly/annual)
                    if isinstance(val_a, dict):
                        val_a = val_a.get("annual", 0)
                    if isinstance(val_b, dict):
                        val_b = val_b.get("annual", 0)
                    
                    try:
                        variance = float(val_b) - float(val_a)
                    except (ValueError, TypeError):
                        variance = 0
                    
                    comparison["differences"].append({
                        "driver": driver_name,
                        "driver_label": label,
                        "category": category,
                        "value_a": val_a,
                        "value_b": val_b,
                        "value_a_formatted": format_driver_value(driver_name, val_a),
                        "value_b_formatted": format_driver_value(driver_name, val_b),
                        "variance": variance,
                        "variance_formatted": format_driver_value(driver_name, variance) if variance != 0 else "—",
                    })
                
                # Build waterfall data for chart
                # This shows the contribution of each driver change to total cost change
                waterfall_data = [
                    {"label": "Baseline", "value": total_a},
                ]
                
                # Calculate impact of each major driver change
                coal_impact = (coal_price_b - coal_price_a) * gen_a * base_coal_per_mwh * transport_factor
                gen_impact = (gen_b - gen_a) * base_coal_per_mwh * coal_price_b * transport_factor
                
                if abs(coal_impact) > 100000:  # Only show significant impacts
                    waterfall_data.append({"label": "Coal Price", "value": coal_impact})
                if abs(gen_impact) > 100000:
                    waterfall_data.append({"label": "Generation", "value": gen_impact})
                
                # Other impact (residual)
                explained = sum(d["value"] for d in waterfall_data[1:]) if len(waterfall_data) > 1 else 0
                other_impact = (total_b - total_a) - explained
                if abs(other_impact) > 100000:
                    waterfall_data.append({"label": "Other", "value": other_impact})
                
                waterfall_data.append({"label": "New Forecast", "value": total_b})
    
    context = get_base_context(request, session, 7)
    context.update({
        "available_scenarios": available_scenarios,
        "scenario_a": scenario_a_obj,
        "scenario_b": scenario_b_obj,
        "comparison": comparison,
        "waterfall_data": json.dumps(waterfall_data) if waterfall_data else "[]",
    })
    
    return templates.TemplateResponse("fuel_forecast/compare.html", context)


# =============================================================================
# Unit-Level Outage Input (Sub-page from Step 5)
# =============================================================================

@router.get("/{session_id}/step/outages", response_class=HTMLResponse)
async def step_outages(request: Request, session_id: str):
    """Render Unit Outage Schedule input page.
    
    This is a detailed sub-page for entering outages by unit by month,
    accessible from the Generation step (step 5).
    """
    from src.models.unit_outage import get_unit_outages_for_year
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get the year from pending changes or default
    year = session.pending_changes.get("forecast_year", datetime.now().year)
    
    # Load outage data from database
    outage_matrix_kc = {}
    outage_matrix_cc = {}
    
    with get_db_session() as db:
        # Kyger Creek (plant_id=1)
        kc_outages = get_unit_outages_for_year(db, 1, year)
        for outage in kc_outages:
            if outage.unit_number not in outage_matrix_kc:
                outage_matrix_kc[outage.unit_number] = {}
            outage_matrix_kc[outage.unit_number][outage.month] = {
                "planned_days": float(outage.planned_outage_days or 0),
                "forced_days": float(outage.forced_outage_days or 0),
            }
        
        # Clifty Creek (plant_id=2)
        cc_outages = get_unit_outages_for_year(db, 2, year)
        for outage in cc_outages:
            if outage.unit_number not in outage_matrix_cc:
                outage_matrix_cc[outage.unit_number] = {}
            outage_matrix_cc[outage.unit_number][outage.month] = {
                "planned_days": float(outage.planned_outage_days or 0),
                "forced_days": float(outage.forced_outage_days or 0),
            }
    
    context = get_base_context(request, session, 5)  # Still on step 5
    context.update({
        "outage_matrix_kc": outage_matrix_kc,
        "outage_matrix_cc": outage_matrix_cc,
        "year": year,
        "forced_outage_rate_kc": session.pending_changes.get("forced_outage_rate_kc", 5),
        "forced_outage_rate_cc": session.pending_changes.get("forced_outage_rate_cc", 5),
    })
    
    return templates.TemplateResponse("fuel_forecast/step_outages.html", context)


@router.post("/{session_id}/step/outages", response_class=RedirectResponse)
async def step_outages_save(request: Request, session_id: str):
    """Save unit outage data and return to Generation step."""
    from src.models.unit_outage import upsert_unit_outage
    
    session = get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    form = await request.form()
    year = session.pending_changes.get("forecast_year", datetime.now().year)
    
    with get_db_session() as db:
        # Save Kyger Creek outages (5 units)
        for unit in range(1, 6):
            for month in range(1, 13):
                field_name = f"kc_{unit}_{month}"
                value = float(form.get(field_name, 0))
                if value > 0:
                    upsert_unit_outage(
                        db,
                        plant_id=1,
                        unit_number=unit,
                        year=year,
                        month=month,
                        planned_outage_days=Decimal(str(value)),
                        forced_outage_days=Decimal("0"),  # Forced is estimated separately
                        updated_by="forecast_workflow",
                    )
        
        # Save Clifty Creek outages (6 units)
        for unit in range(1, 7):
            for month in range(1, 13):
                field_name = f"cc_{unit}_{month}"
                value = float(form.get(field_name, 0))
                if value > 0:
                    upsert_unit_outage(
                        db,
                        plant_id=2,
                        unit_number=unit,
                        year=year,
                        month=month,
                        planned_outage_days=Decimal(str(value)),
                        forced_outage_days=Decimal("0"),
                        updated_by="forecast_workflow",
                    )
    
    # Store EFOR estimates in session
    forced_rate_kc = float(form.get("forced_outage_rate_kc", 5))
    forced_rate_cc = float(form.get("forced_outage_rate_cc", 5))
    session.set_pending_change("forced_outage_rate_kc", forced_rate_kc)
    session.set_pending_change("forced_outage_rate_cc", forced_rate_cc)
    
    # Calculate total outage days for summary
    total_kc = sum(
        float(form.get(f"kc_{unit}_{month}", 0))
        for unit in range(1, 6)
        for month in range(1, 13)
    )
    total_cc = sum(
        float(form.get(f"cc_{unit}_{month}", 0))
        for unit in range(1, 7)
        for month in range(1, 13)
    )
    
    session.set_pending_change("outage_days_planned_kc", total_kc)
    session.set_pending_change("outage_days_planned_cc", total_cc)
    session.set_pending_change("outage_days_planned", total_kc + total_cc)
    
    session.add_change_note(f"Updated unit outages: KC={total_kc:.0f} days, CC={total_cc:.0f} days")
    update_session(session)
    
    return RedirectResponse(
        url=f"/fuel-forecast/{session_id}/step/6",
        status_code=302
    )
