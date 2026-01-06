"""Excel import for fuel model drivers.

Imports driver values from OVEC fuel model Excel workbooks:
- Use factors (monthly by plant)
- Heat rates (baseline, SUF correction)
- Coal prices (NAPP, ILB)
- Barge rates
- Inventory targets
- Generation parameters

Supports both the simple "2025 - Fuel.xlsx" and the comprehensive 
"OVEC IKEC Energy Budget" xlsm files.
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import re

from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


@dataclass
class ImportResult:
    """Result of an Excel import operation."""
    success: bool
    file_path: str
    scenario_id: int
    year: int
    drivers_imported: int = 0
    values_imported: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    details: Dict[str, Any] = field(default_factory=dict)


@dataclass
class CellMapping:
    """Maps an Excel cell to a driver value."""
    driver_name: str
    sheet: str
    cell: Optional[str] = None  # Single cell like "B5"
    row: Optional[int] = None   # For dynamic lookup
    col: Optional[int] = None
    transform: Optional[callable] = None  # Optional value transform


@dataclass
class RangeMapping:
    """Maps an Excel range to monthly driver values."""
    driver_name: str
    sheet: str
    start_row: int
    start_col: int  # Column for first month
    plant_id: Optional[int] = None
    plant_name: Optional[str] = None
    transform: Optional[callable] = None


# =============================================================================
# Excel Cell Mappings for OVEC IKEC Energy Budget workbook
# =============================================================================

# KC Forecast Inputs sheet mappings
KC_INPUTS_MAPPINGS = [
    CellMapping("escalation_coal_annual", "KC Forecast Inputs", cell="B5", 
                transform=lambda x: Decimal(str(x * 100)) if x else Decimal("0")),  # Convert to %
    CellMapping("reserve_mw", "KC Forecast Inputs", cell="B7"),
    CellMapping("gsu_loss_pct", "KC Forecast Inputs", cell="B8",
                transform=lambda x: Decimal(str(x * 100)) if x else Decimal("0")),  # Convert to %
    CellMapping("capacity_mw", "KC Forecast Inputs", cell="B11",
                transform=lambda x: Decimal(str(x * 5)) if x else Decimal("1025")),  # 5 units
    CellMapping("fgd_aux_pct", "KC Forecast Inputs", cell="B13",
                transform=lambda x: Decimal(str(x / 205 * 100)) if x else Decimal("2.5")),  # Convert MW to %
    CellMapping("inventory_target_days", "KC Forecast Inputs", cell="B20"),
]

# CC Forecast Inputs sheet mappings (same structure, different plant)
CC_INPUTS_MAPPINGS = [
    CellMapping("capacity_mw", "CC Forecast Inputs", cell="B11",
                transform=lambda x: Decimal(str(x * 6)) if x else Decimal("1302")),  # 6 units
]

# Coal price mappings from Northern Appalachia Coal Fcst
COAL_PRICE_MAPPINGS = [
    # Row 15 has 2025 prices, Row 16 has 2026, etc.
    # Column B is 12500 BTU, Column E is Scenario Base
    CellMapping("coal_price_eastern", "Northern Appalachia Coal Fcst", row=15, col=2),  # 2025
]


def _cell_to_coords(cell_ref: str) -> Tuple[int, int]:
    """Convert A1-style cell reference to (row, col) tuple."""
    import openpyxl.utils
    col_letter = ''.join(c for c in cell_ref if c.isalpha())
    row_num = int(''.join(c for c in cell_ref if c.isdigit()))
    col_num = openpyxl.utils.column_index_from_string(col_letter)
    return (row_num, col_num)


def _get_cell_value(ws, cell_ref: str = None, row: int = None, col: int = None) -> Any:
    """Get a cell value by reference or coordinates."""
    if cell_ref:
        row, col = _cell_to_coords(cell_ref)
    return ws.cell(row=row, column=col).value


def _find_year_row(ws, year: int, col: int = 1, start_row: int = 10, end_row: int = 50) -> Optional[int]:
    """Find the row containing a specific year value."""
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val == year:
            return row
    return None


def _parse_use_factors(ws, plant_name: str, year: int) -> Dict[int, Decimal]:
    """Parse monthly use factors from the Use Factor Input sheet.
    
    Args:
        ws: Worksheet object
        plant_name: 'Kyger Creek' or 'Clifty Creek'
        year: Target year
        
    Returns:
        Dict mapping month (1-12) to use factor (as percentage 0-100)
    """
    result = {}
    
    # Find the row for the plant
    plant_row = None
    for row in range(1, 30):
        val = ws.cell(row=row, column=1).value
        if val and plant_name.lower() in str(val).lower():
            plant_row = row
            break
    
    if not plant_row:
        logger.warning(f"Could not find {plant_name} in Use Factor Input sheet")
        return result
    
    # Find columns for target year months
    # The header row is typically one row before the plant row
    header_row = plant_row - 1
    
    for col in range(2, 20):  # Check columns B through T
        date_val = ws.cell(row=header_row, column=col).value
        if date_val and isinstance(date_val, datetime):
            if date_val.year == year:
                month = date_val.month
                use_factor = ws.cell(row=plant_row, column=col).value
                if use_factor is not None:
                    # Convert to percentage (0-100)
                    result[month] = Decimal(str(use_factor * 100))
    
    return result


def _parse_heat_rates(ws, year: int) -> Dict[str, Decimal]:
    """Parse heat rate values from KC/CC Coal Burn sheet.
    
    Returns dict with keys: 'baseline', 'suf_correction'
    """
    result = {}
    
    # Find the column for the target year
    year_col = None
    for col in range(2, 20):
        # Check row 3 for year value
        val = ws.cell(row=3, column=col).value
        if val == year:
            year_col = col
            break
    
    if not year_col:
        logger.warning(f"Could not find year {year} in Coal Burn sheet")
        return result
    
    # Row 13 has baseline heat rate
    baseline = ws.cell(row=13, column=year_col).value
    if baseline:
        result['baseline'] = Decimal(str(baseline))
    
    # Row 14 has SUF correction
    suf = ws.cell(row=14, column=year_col).value
    if suf:
        # SUF is typically the full corrected rate, so we calculate the correction
        if baseline and suf:
            result['suf_correction'] = Decimal(str(suf - baseline))
        else:
            result['suf_correction'] = Decimal("0")
    
    return result


def import_ovec_energy_budget(
    db: Session,
    scenario_id: int,
    excel_path: str,
    year: int,
    plant_id: Optional[int] = None,
) -> ImportResult:
    """Import driver values from OVEC IKEC Energy Budget workbook.
    
    Args:
        db: Database session
        scenario_id: Target scenario ID
        excel_path: Path to Excel file
        year: Target year to import
        plant_id: Optional specific plant (1=Kyger, 2=Clifty), None for both
        
    Returns:
        ImportResult with import details
    """
    import openpyxl
    
    result = ImportResult(
        success=False,
        file_path=excel_path,
        scenario_id=scenario_id,
        year=year,
    )
    
    # Verify file exists
    path = Path(excel_path)
    if not path.exists():
        result.errors.append(f"File not found: {excel_path}")
        return result
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        result.errors.append(f"Failed to open Excel file: {e}")
        return result
    
    # Import from drivers module
    from src.engine.drivers import FuelModel
    from src.engine.default_drivers import create_default_fuel_model
    from src.engine.scenario_drivers import save_driver_values_to_scenario
    
    model = create_default_fuel_model()
    imported_drivers = set()
    
    try:
        # Import Kyger Creek inputs
        if plant_id is None or plant_id == 1:
            if 'KC Forecast Inputs' in wb.sheetnames:
                ws = wb['KC Forecast Inputs']
                
                for mapping in KC_INPUTS_MAPPINGS:
                    try:
                        if mapping.cell:
                            value = _get_cell_value(ws, cell_ref=mapping.cell)
                        else:
                            value = _get_cell_value(ws, row=mapping.row, col=mapping.col)
                        
                        if value is not None:
                            if mapping.transform:
                                value = mapping.transform(value)
                            else:
                                value = Decimal(str(value))
                            
                            model.set_driver_value(mapping.driver_name, year, None, value, plant_id=1)
                            imported_drivers.add(mapping.driver_name)
                            result.values_imported += 1
                    except Exception as e:
                        result.warnings.append(f"Error importing {mapping.driver_name}: {e}")
            
            # Import Kyger heat rates
            if 'KC Coal Burn' in wb.sheetnames:
                ws = wb['KC Coal Burn']
                heat_rates = _parse_heat_rates(ws, year)
                if 'baseline' in heat_rates:
                    model.set_driver_value('heat_rate_baseline', year, None, heat_rates['baseline'], plant_id=1)
                    imported_drivers.add('heat_rate_baseline')
                    result.values_imported += 1
                if 'suf_correction' in heat_rates:
                    model.set_driver_value('heat_rate_suf_correction', year, None, heat_rates['suf_correction'], plant_id=1)
                    imported_drivers.add('heat_rate_suf_correction')
                    result.values_imported += 1
        
        # Import Clifty Creek inputs
        if plant_id is None or plant_id == 2:
            if 'CC Forecast Inputs' in wb.sheetnames:
                ws = wb['CC Forecast Inputs']
                
                for mapping in CC_INPUTS_MAPPINGS:
                    try:
                        if mapping.cell:
                            value = _get_cell_value(ws, cell_ref=mapping.cell)
                        else:
                            value = _get_cell_value(ws, row=mapping.row, col=mapping.col)
                        
                        if value is not None:
                            if mapping.transform:
                                value = mapping.transform(value)
                            else:
                                value = Decimal(str(value))
                            
                            model.set_driver_value(mapping.driver_name, year, None, value, plant_id=2)
                            imported_drivers.add(mapping.driver_name)
                            result.values_imported += 1
                    except Exception as e:
                        result.warnings.append(f"Error importing {mapping.driver_name}: {e}")
        
        # Import use factors
        if 'Use Factor Input' in wb.sheetnames:
            ws = wb['Use Factor Input']
            
            if plant_id is None or plant_id == 1:
                kc_factors = _parse_use_factors(ws, 'Kyger Creek', year)
                for month, value in kc_factors.items():
                    model.set_driver_value('use_factor', year, month, value, plant_id=1)
                    result.values_imported += 1
                if kc_factors:
                    imported_drivers.add('use_factor')
            
            if plant_id is None or plant_id == 2:
                cc_factors = _parse_use_factors(ws, 'Clifty Creek', year)
                for month, value in cc_factors.items():
                    model.set_driver_value('use_factor', year, month, value, plant_id=2)
                    result.values_imported += 1
                if cc_factors:
                    imported_drivers.add('use_factor')
        
        # Import coal prices
        if 'Northern Appalachia Coal Fcst' in wb.sheetnames:
            ws = wb['Northern Appalachia Coal Fcst']
            year_row = _find_year_row(ws, year, col=1, start_row=10)
            if year_row:
                price = ws.cell(row=year_row, column=2).value  # Column B has 12500 BTU price
                if price:
                    model.set_driver_value('coal_price_eastern', year, None, Decimal(str(price)))
                    imported_drivers.add('coal_price_eastern')
                    result.values_imported += 1
        
        if 'Illinois Basin Coal Fcst' in wb.sheetnames:
            ws = wb['Illinois Basin Coal Fcst']
            year_row = _find_year_row(ws, year, col=1, start_row=10)
            if year_row:
                price = ws.cell(row=year_row, column=2).value
                if price:
                    model.set_driver_value('coal_price_ilb', year, None, Decimal(str(price)))
                    imported_drivers.add('coal_price_ilb')
                    result.values_imported += 1
        
        # Save imported values to scenario
        save_result = save_driver_values_to_scenario(db, model, scenario_id, year)
        
        result.drivers_imported = len(imported_drivers)
        result.details['drivers'] = list(imported_drivers)
        result.success = True
        
    except Exception as e:
        result.errors.append(f"Import failed: {e}")
        logger.exception("Excel import failed")
    
    finally:
        wb.close()
    
    return result


def import_simple_fuel_file(
    db: Session,
    scenario_id: int,
    excel_path: str,
    year: int,
) -> ImportResult:
    """Import from the simple 2025 - Fuel.xlsx format.
    
    This file has monthly sheets with limited metrics:
    - Coal Inventory - # Days of Coal in Inventory
    - Generation (MW)
    - Other Fuel - Total Cost ($)
    - Other Fuel Related Costs - Gypsum Sales ($)
    
    Args:
        db: Database session
        scenario_id: Target scenario ID
        excel_path: Path to Excel file
        year: Target year
        
    Returns:
        ImportResult
    """
    import openpyxl
    
    result = ImportResult(
        success=False,
        file_path=excel_path,
        scenario_id=scenario_id,
        year=year,
    )
    
    path = Path(excel_path)
    if not path.exists():
        result.errors.append(f"File not found: {excel_path}")
        return result
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        result.errors.append(f"Failed to open Excel file: {e}")
        return result
    
    from src.engine.drivers import FuelModel
    from src.engine.default_drivers import create_default_fuel_model
    from src.engine.scenario_drivers import save_driver_values_to_scenario
    
    model = create_default_fuel_model()
    
    # Map month names to numbers
    month_map = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }
    
    # Parse each monthly sheet
    for sheet_name, month_num in month_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Parse rows to find data
        for row in range(2, 20):
            location = ws.cell(row=row, column=3).value  # Column C
            measure = ws.cell(row=row, column=4).value   # Column D
            projected = ws.cell(row=row, column=5).value  # Column E (Month Projected)
            
            if not location or not measure or projected is None:
                continue
            
            plant_id = None
            if 'Kyger' in str(location):
                plant_id = 1
            elif 'Clifty' in str(location):
                plant_id = 2
            # Skip Corporate (combined) for now
            
            if plant_id is None:
                continue
            
            # Map measures to drivers
            if 'Days of Coal' in str(measure):
                model.set_driver_value('inventory_target_days', year, month_num, 
                                      Decimal(str(projected)), plant_id=plant_id)
                result.values_imported += 1
            elif 'Generation' in str(measure):
                # This is actual generation, could use to back-calculate use factor
                pass
    
    try:
        save_driver_values_to_scenario(db, model, scenario_id, year)
        result.success = True
        result.drivers_imported = 1  # Just inventory target days
    except Exception as e:
        result.errors.append(f"Failed to save: {e}")
    
    finally:
        wb.close()
    
    return result


def import_fuel_drivers_from_excel(
    db: Session,
    scenario_id: int,
    excel_path: str,
    year: int,
    plant_id: Optional[int] = None,
) -> ImportResult:
    """Import fuel driver values from Excel file.
    
    Auto-detects the file type and uses appropriate parser.
    
    Args:
        db: Database session
        scenario_id: Target scenario ID
        excel_path: Path to Excel file
        year: Year to import data for
        plant_id: Optional specific plant ID
        
    Returns:
        ImportResult with import details
    """
    path = Path(excel_path)
    
    # Detect file type by extension and name patterns
    if path.suffix.lower() == '.xlsm':
        # Macro-enabled workbook - likely the comprehensive one
        return import_ovec_energy_budget(db, scenario_id, excel_path, year, plant_id)
    elif 'Fuel' in path.name:
        # Simple fuel file
        return import_simple_fuel_file(db, scenario_id, excel_path, year)
    else:
        # Default to comprehensive import
        return import_ovec_energy_budget(db, scenario_id, excel_path, year, plant_id)

