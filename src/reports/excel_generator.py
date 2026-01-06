"""Excel report generator for sponsor reports."""

from io import BytesIO
from datetime import datetime
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func

from src.models import Forecast, Scenario, Plant, CostCategory, Period
from src.models.cost_category import CostSection
from src.models.period import Granularity


def create_styles():
    """Create reusable Excel styles."""
    styles = {}
    
    # Header style
    styles['header'] = NamedStyle(name='header')
    styles['header'].font = Font(bold=True, size=12, color='FFFFFF')
    styles['header'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    styles['header'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section header
    styles['section'] = NamedStyle(name='section')
    styles['section'].font = Font(bold=True, size=11)
    styles['section'].fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    
    # Subtotal row
    styles['subtotal'] = NamedStyle(name='subtotal')
    styles['subtotal'].font = Font(bold=True)
    styles['subtotal'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    # Currency format
    styles['currency'] = NamedStyle(name='currency')
    styles['currency'].number_format = '#,##0'
    
    # $/MWhr format
    styles['per_mwh'] = NamedStyle(name='per_mwh')
    styles['per_mwh'].number_format = '$#,##0.00'
    
    # Generation format
    styles['generation'] = NamedStyle(name='generation')
    styles['generation'].number_format = '#,##0'
    
    return styles


def generate_sponsor_report(
    db: Session,
    scenario_id: int,
    years: int = 2,
    include_monthly: bool = True,
) -> BytesIO:
    """
    Generate Excel report for sponsors.
    
    Args:
        db: Database session
        scenario_id: ID of the scenario to report on
        years: Number of years to include (default 2)
        include_monthly: Include monthly detail for first 2 years
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    styles = create_styles()
    
    # Get scenario info
    scenario = db.query(Scenario).filter(Scenario.id == scenario_id).first()
    
    # Get all plants
    plants = db.query(Plant).filter(Plant.is_active == True).all()
    
    # Get cost categories ordered by section and sort_order
    categories = (
        db.query(CostCategory)
        .filter(CostCategory.is_active == True)
        .order_by(CostCategory.section, CostCategory.sort_order)
        .all()
    )
    
    # Determine year range
    current_year = datetime.now().year
    year_range = range(current_year, current_year + years)
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, db, scenario, plants, categories, year_range, styles)
    
    # Create Monthly Detail sheet if requested
    if include_monthly and years >= 1:
        ws_monthly = wb.create_sheet("Monthly Detail")
        _create_monthly_sheet(ws_monthly, db, scenario, plants, categories, current_year, min(years, 2), styles)
    
    # Create sheets by plant
    for plant in plants:
        ws_plant = wb.create_sheet(plant.short_name)
        _create_plant_sheet(ws_plant, db, scenario, plant, categories, year_range, styles)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def _create_summary_sheet(ws, db, scenario, plants, categories, year_range, styles):
    """Create the summary sheet with annual totals."""
    # Title
    ws['A1'] = f"OVEC Financial Forecast - {scenario.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = f"Scenario Type: {scenario.scenario_type.value.replace('_', ' ').title()}"
    
    # Headers starting at row 5
    row = 5
    headers = ['Cost Category'] + [str(y) for y in year_range] + ['Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Generation row
    ws.cell(row=row, column=1, value="GENERATION (MWh)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    total_gen = Decimal(0)
    for col, year in enumerate(year_range, 2):
        gen = _get_generation_for_year(db, scenario.id, year)
        ws.cell(row=row, column=col, value=float(gen) if gen else 0)
        ws.cell(row=row, column=col).number_format = '#,##0'
        total_gen += gen or Decimal(0)
    ws.cell(row=row, column=len(list(year_range)) + 2, value=float(total_gen))
    ws.cell(row=row, column=len(list(year_range)) + 2).number_format = '#,##0'
    row += 1
    
    # Cost sections
    sections = [
        (CostSection.FUEL, "FUEL COSTS"),
        (CostSection.OPERATING, "OPERATING COSTS"),
        (CostSection.NON_OPERATING, "NON-OPERATING COSTS"),
        (CostSection.CAPITAL, "CAPITAL COSTS"),
    ]
    
    grand_total = {y: Decimal(0) for y in year_range}
    
    for section, section_name in sections:
        # Section header
        row += 1
        ws.cell(row=row, column=1, value=section_name)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        row += 1
        
        section_cats = [c for c in categories if c.section == section and not c.is_subtotal]
        section_total = {y: Decimal(0) for y in year_range}
        
        for cat in section_cats:
            ws.cell(row=row, column=1, value=f"  {cat.name}")
            for col, year in enumerate(year_range, 2):
                cost = _get_cost_for_category_year(db, scenario.id, cat.id, year)
                ws.cell(row=row, column=col, value=float(cost) if cost else 0)
                ws.cell(row=row, column=col).number_format = '#,##0'
                section_total[year] += cost or Decimal(0)
            row += 1
        
        # Section subtotal
        ws.cell(row=row, column=1, value=f"{section_name} SUBTOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        section_grand = Decimal(0)
        for col, year in enumerate(year_range, 2):
            ws.cell(row=row, column=col, value=float(section_total[year]))
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True)
            grand_total[year] += section_total[year]
            section_grand += section_total[year]
        ws.cell(row=row, column=len(list(year_range)) + 2, value=float(section_grand))
        ws.cell(row=row, column=len(list(year_range)) + 2).number_format = '#,##0'
        ws.cell(row=row, column=len(list(year_range)) + 2).font = Font(bold=True)
        
        # $/MWhr for section
        row += 1
        ws.cell(row=row, column=1, value=f"  $/MWhr")
        for col, year in enumerate(year_range, 2):
            gen = _get_generation_for_year(db, scenario.id, year)
            if gen and gen > 0:
                cpm = section_total[year] / gen
                ws.cell(row=row, column=col, value=float(cpm))
                ws.cell(row=row, column=col).number_format = '$#,##0.00'
        row += 1
    
    # Grand total
    row += 1
    ws.cell(row=row, column=1, value="TOTAL ALL-IN COST")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    all_total = Decimal(0)
    for col, year in enumerate(year_range, 2):
        ws.cell(row=row, column=col, value=float(grand_total[year]))
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFF')
        all_total += grand_total[year]
    ws.cell(row=row, column=len(list(year_range)) + 2, value=float(all_total))
    
    # Total $/MWhr
    row += 1
    ws.cell(row=row, column=1, value="ALL-IN $/MWhr")
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col, year in enumerate(year_range, 2):
        gen = _get_generation_for_year(db, scenario.id, year)
        if gen and gen > 0:
            cpm = grand_total[year] / gen
            ws.cell(row=row, column=col, value=float(cpm))
            ws.cell(row=row, column=col).number_format = '$#,##0.00'
            ws.cell(row=row, column=col).font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    for col in range(2, len(list(year_range)) + 3):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_monthly_sheet(ws, db, scenario, plants, categories, start_year, num_years, styles):
    """Create monthly detail sheet."""
    ws['A1'] = f"Monthly Detail - {scenario.name}"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Build month headers
    row = 3
    headers = ['Cost Category']
    months = []
    for year in range(start_year, start_year + num_years):
        for month in range(1, 13):
            from calendar import month_abbr
            headers.append(f"{month_abbr[month]} {year}")
            months.append((year, month))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Data rows
    for section in CostSection:
        section_cats = [c for c in categories if c.section == section and not c.is_subtotal]
        
        # Section header
        ws.cell(row=row, column=1, value=section.value.upper().replace('_', ' '))
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        row += 1
        
        for cat in section_cats:
            ws.cell(row=row, column=1, value=f"  {cat.name}")
            for col, (year, month) in enumerate(months, 2):
                cost = _get_cost_for_category_month(db, scenario.id, cat.id, year, month)
                ws.cell(row=row, column=col, value=float(cost) if cost else 0)
                ws.cell(row=row, column=col).number_format = '#,##0'
            row += 1
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


def _create_plant_sheet(ws, db, scenario, plant, categories, year_range, styles):
    """Create a sheet for a specific plant."""
    ws['A1'] = f"{plant.name} - {scenario.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Capacity: {plant.capacity_mw} MW ({plant.unit_count} x {plant.unit_capacity_mw} MW units)"
    
    # Similar structure to summary but filtered by plant
    row = 4
    headers = ['Cost Category'] + [str(y) for y in year_range]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    
    row += 1
    
    # Generation
    ws.cell(row=row, column=1, value="GENERATION (MWh)")
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col, year in enumerate(year_range, 2):
        gen = _get_generation_for_year(db, scenario.id, year, plant.id)
        ws.cell(row=row, column=col, value=float(gen) if gen else 0)
        ws.cell(row=row, column=col).number_format = '#,##0'
    
    row += 2
    
    # Costs by section
    for section in CostSection:
        section_cats = [c for c in categories if c.section == section and not c.is_subtotal]
        
        ws.cell(row=row, column=1, value=section.value.upper().replace('_', ' '))
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
        row += 1
        
        for cat in section_cats:
            ws.cell(row=row, column=1, value=f"  {cat.name}")
            for col, year in enumerate(year_range, 2):
                cost = _get_cost_for_category_year(db, scenario.id, cat.id, year, plant.id)
                ws.cell(row=row, column=col, value=float(cost) if cost else 0)
                ws.cell(row=row, column=col).number_format = '#,##0'
            row += 1
        
        row += 1
    
    ws.column_dimensions['A'].width = 35
    for col in range(2, len(list(year_range)) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _get_generation_for_year(db, scenario_id: int, year: int, plant_id: int = None) -> Optional[Decimal]:
    """Get total generation for a year."""
    query = (
        db.query(func.sum(Forecast.generation_mwh))
        .join(Period)
        .filter(Forecast.scenario_id == scenario_id)
        .filter(Period.year == year)
    )
    if plant_id:
        query = query.filter(Forecast.plant_id == plant_id)
    
    result = query.scalar()
    return Decimal(str(result)) if result else Decimal(0)


def _get_cost_for_category_year(db, scenario_id: int, category_id: int, year: int, plant_id: int = None) -> Optional[Decimal]:
    """Get total cost for a category in a year."""
    query = (
        db.query(func.sum(Forecast.cost_dollars))
        .join(Period)
        .filter(Forecast.scenario_id == scenario_id)
        .filter(Forecast.category_id == category_id)
        .filter(Period.year == year)
    )
    if plant_id:
        query = query.filter(Forecast.plant_id == plant_id)
    
    result = query.scalar()
    return Decimal(str(result)) if result else Decimal(0)


def _get_cost_for_category_month(db, scenario_id: int, category_id: int, year: int, month: int) -> Optional[Decimal]:
    """Get cost for a category in a specific month."""
    query = (
        db.query(func.sum(Forecast.cost_dollars))
        .join(Period)
        .filter(Forecast.scenario_id == scenario_id)
        .filter(Forecast.category_id == category_id)
        .filter(Period.year == year)
        .filter(Period.month == month)
    )
    
    result = query.scalar()
    return Decimal(str(result)) if result else Decimal(0)

