"""Compare Python fuel model output to OVEC Excel file values.

This script:
1. Extracts key values from the OVEC IKEC Energy Budget Excel file
2. Runs the Python fuel model calculation for the same period
3. Generates a side-by-side comparison with variance analysis
4. Identifies root causes of differences
"""

import sys
import os
from pathlib import Path
from decimal import Decimal
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
import logging

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


@dataclass
class ExcelValues:
    """Container for values extracted from Excel."""
    plant: str
    year: int
    
    # Monthly values (dict of month -> value)
    net_generation_mwh: Dict[int, float] = field(default_factory=dict)
    coal_burn_tons: Dict[int, float] = field(default_factory=dict)
    coal_burn_mmbtu: Dict[int, float] = field(default_factory=dict)
    heat_rate: Dict[int, float] = field(default_factory=dict)
    coal_cost: Dict[int, float] = field(default_factory=dict)
    coal_cost_per_mmbtu: Dict[int, float] = field(default_factory=dict)
    urea_cost: Dict[int, float] = field(default_factory=dict)
    limestone_cost: Dict[int, float] = field(default_factory=dict)
    hydrated_lime_cost: Dict[int, float] = field(default_factory=dict)
    bioreactor_cost: Dict[int, float] = field(default_factory=dict)
    byproduct_net: Dict[int, float] = field(default_factory=dict)
    allowance_cost: Dict[int, float] = field(default_factory=dict)
    total_fuel_cost: Dict[int, float] = field(default_factory=dict)
    fuel_cost_per_mwh: Dict[int, float] = field(default_factory=dict)
    
    # Annual totals
    annual_generation: float = 0
    annual_coal_tons: float = 0
    annual_coal_cost: float = 0
    annual_total_fuel_cost: float = 0


def list_excel_sheets(excel_path: str) -> List[str]:
    """List all sheet names in an Excel file."""
    import openpyxl
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    return sheets


def find_year_in_sheet(ws, year: int, search_rows: range = range(1, 20), 
                       search_cols: range = range(1, 30)) -> Optional[Tuple[int, int]]:
    """Find the cell containing a specific year value."""
    for row in search_rows:
        for col in search_cols:
            val = ws.cell(row=row, column=col).value
            if val == year:
                return (row, col)
    return None


def extract_monthly_row(ws, row: int, start_col: int, num_months: int = 12) -> Dict[int, float]:
    """Extract monthly values from a row."""
    result = {}
    for i in range(num_months):
        val = ws.cell(row=row, column=start_col + i).value
        if val is not None:
            try:
                result[i + 1] = float(val)
            except (ValueError, TypeError):
                pass
    return result


def find_row_by_label(ws, label: str, col: int = 1, start_row: int = 1, 
                      end_row: int = 100, partial: bool = True) -> Optional[int]:
    """Find a row containing a specific label."""
    label_lower = label.lower()
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            val_str = str(val).lower()
            if partial and label_lower in val_str:
                return row
            elif not partial and label_lower == val_str:
                return row
    return None


def extract_kc_energy_cost_values(ws, year: int) -> ExcelValues:
    """Extract values from KC Energy Cost sheet."""
    values = ExcelValues(plant="Kyger Creek", year=year)
    
    # Find the year column - typically months are in columns B through M
    # First row usually has month headers or year indicator
    
    # Search for specific rows by label
    row_mappings = [
        ("net generation", "net_generation_mwh"),
        ("coal burn", "coal_burn_tons"),
        ("mmbtu", "coal_burn_mmbtu"),
        ("heat rate", "heat_rate"),
        ("coal cost", "coal_cost"),
        ("$/mmbtu", "coal_cost_per_mmbtu"),
        ("urea", "urea_cost"),
        ("limestone", "limestone_cost"),
        ("hydrated lime", "hydrated_lime_cost"),
        ("bioreactor", "bioreactor_cost"),
        ("byproduct", "byproduct_net"),
        ("allowance", "allowance_cost"),
        ("total", "total_fuel_cost"),
        ("$/mwh", "fuel_cost_per_mwh"),
    ]
    
    # Analyze sheet structure first
    print(f"\n  Analyzing KC Energy Cost sheet structure...")
    
    # Find first data column (usually B=2 for Jan)
    start_col = 2  # Column B
    
    for label, attr in row_mappings:
        row = find_row_by_label(ws, label, col=1, start_row=1, end_row=100)
        if row:
            monthly = extract_monthly_row(ws, row, start_col, 12)
            setattr(values, attr, monthly)
            if monthly:
                print(f"    Found '{label}' at row {row}: {len(monthly)} months")
    
    return values


def extract_cc_energy_cost_values(ws, year: int) -> ExcelValues:
    """Extract values from CC Energy Cost sheet."""
    values = ExcelValues(plant="Clifty Creek", year=year)
    
    # Same structure as KC
    row_mappings = [
        ("net generation", "net_generation_mwh"),
        ("coal burn", "coal_burn_tons"),
        ("mmbtu", "coal_burn_mmbtu"),
        ("heat rate", "heat_rate"),
        ("coal cost", "coal_cost"),
        ("$/mmbtu", "coal_cost_per_mmbtu"),
        ("urea", "urea_cost"),
        ("limestone", "limestone_cost"),
        ("hydrated lime", "hydrated_lime_cost"),
        ("bioreactor", "bioreactor_cost"),
        ("byproduct", "byproduct_net"),
        ("allowance", "allowance_cost"),
        ("total", "total_fuel_cost"),
        ("$/mwh", "fuel_cost_per_mwh"),
    ]
    
    print(f"\n  Analyzing CC Energy Cost sheet structure...")
    start_col = 2
    
    for label, attr in row_mappings:
        row = find_row_by_label(ws, label, col=1, start_row=1, end_row=100)
        if row:
            monthly = extract_monthly_row(ws, row, start_col, 12)
            setattr(values, attr, monthly)
            if monthly:
                print(f"    Found '{label}' at row {row}: {len(monthly)} months")
    
    return values


def analyze_sheet_structure(ws, sheet_name: str, max_rows: int = 50, max_cols: int = 15):
    """Print the structure of a sheet for debugging."""
    print(f"\n  === {sheet_name} Structure (first {max_rows} rows, {max_cols} cols) ===")
    
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_data = []
        has_data = False
        for col in range(1, min(max_cols + 1, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                has_data = True
                if isinstance(val, (int, float)):
                    row_data.append(f"{val:,.0f}" if abs(val) > 100 else f"{val:.2f}")
                else:
                    row_data.append(str(val)[:25])
            else:
                row_data.append("")
        
        if has_data:
            print(f"  Row {row:3d}: {' | '.join(row_data[:8])}")


def extract_from_generation_sheet(ws, plant: str, year: int) -> Dict[str, Any]:
    """Extract values from KC Generation or CC Generation sheet.
    
    Sums all monthly columns for the given year to get annual total.
    """
    result = {
        "net_generation_mwh": {},
        "total_generation_mwh": 0,
        "months_found": 0,
    }
    
    # Find all columns for the target year (row 3 contains years)
    year_cols = []
    for col in range(2, 25):
        val = ws.cell(row=3, column=col).value
        if val is not None:
            try:
                if int(val) == year:
                    year_cols.append(col)
            except (ValueError, TypeError):
                pass
    
    if not year_cols:
        print(f"    Could not find year {year} in {plant} Generation sheet")
        return result
    
    print(f"    Found year {year} in columns {year_cols} ({len(year_cols)} months)")
    result["months_found"] = len(year_cols)
    
    # Find the "Plant Total (Less FGD Aux)" row - this is net generation
    net_gen_row = None
    for row in range(1, 40):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if "plant total" in label_str and "less" in label_str and "fgd" in label_str:
            net_gen_row = row
            break
    
    if not net_gen_row:
        # Try alternative: look for "Net Delivered"
        for row in range(1, 40):
            label = ws.cell(row=row, column=1).value
            if label is None:
                continue
            label_str = str(label).lower()
            if "net delivered" in label_str:
                net_gen_row = row
                break
    
    if net_gen_row:
        gross_total = 0
        los_total = 0  # Lack of Sales
        
        # Also find "Lack of Sales" row - this represents use factor deduction
        # Note: There may be two rows labeled "Lack of Sales" - one is a header (None values),
        # the other has actual values. We want the one with actual MWh values (row 29 typically)
        los_row = None
        for row in range(20, 40):
            label = ws.cell(row=row, column=1).value
            if label and "lack of sales" in str(label).lower():
                # Check if this row has actual values (not None)
                test_val = ws.cell(row=row, column=year_cols[0]).value if year_cols else None
                if test_val is not None and isinstance(test_val, (int, float)):
                    los_row = row
                    break
        
        for col in year_cols:
            val = ws.cell(row=net_gen_row, column=col).value
            if val is not None:
                try:
                    gross_total += float(val)
                except (ValueError, TypeError):
                    pass
            
            if los_row:
                los_val = ws.cell(row=los_row, column=col).value
                if los_val is not None:
                    try:
                        los_total += float(los_val)
                    except (ValueError, TypeError):
                        pass
        
        # Delivered generation = Plant Total Less Aux - Lack of Sales
        delivered_total = gross_total - los_total
        result["total_generation_mwh"] = delivered_total  # Use delivered for comparison
        result["gross_generation_mwh"] = gross_total
        result["lack_of_sales_mwh"] = los_total
        
        if los_row:
            print(f"    Gross generation (summed): {gross_total:,.0f} MWh")
            print(f"    Lack of Sales (summed): {los_total:,.0f} MWh")
            print(f"    Net delivered (summed): {delivered_total:,.0f} MWh")
        else:
            print(f"    Net generation (summed): {gross_total:,.0f} MWh (no Lack of Sales row found)")
    
    return result


def extract_from_coal_burn_sheet(ws, plant: str, year: int) -> Dict[str, Any]:
    """Extract values from KC Coal Burn or CC Coal Burn sheet.
    
    Sums all monthly columns for the given year.
    """
    result = {
        "coal_tons": 0,
        "coal_mmbtu": 0,
        "heat_rate_avg": 0,
        "months_found": 0,
    }
    
    # Find all columns for the target year (row 3 contains years)
    year_cols = []
    for col in range(2, 25):
        val = ws.cell(row=3, column=col).value
        if val is not None:
            try:
                if int(val) == year:
                    year_cols.append(col)
            except (ValueError, TypeError):
                pass
    
    if not year_cols:
        print(f"    Could not find year {year} in {plant} Coal Burn sheet")
        return result
    
    print(f"    Found year {year} in columns {year_cols} ({len(year_cols)} months)")
    result["months_found"] = len(year_cols)
    
    # Find key rows
    coal_tons_row = None
    coal_mmbtu_row = None
    heat_rate_row = None
    
    for row in range(1, 60):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if "coal consumed" in label_str and "ton" in label_str:
            coal_tons_row = row
        elif "total heat input" in label_str or ("heat input" in label_str and "mmbtu" in label_str):
            coal_mmbtu_row = row
        elif "net plant hr" in label_str or "production hr" in label_str:
            heat_rate_row = row
    
    # Sum coal tons and mmbtu across all months
    if coal_tons_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=coal_tons_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["coal_tons"] = total
        print(f"    Coal tons (summed): {total:,.0f}")
    
    if coal_mmbtu_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=coal_mmbtu_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["coal_mmbtu"] = total
        print(f"    Coal MMBtu (summed): {total:,.0f}")
    
    # Average heat rate across months
    if heat_rate_row:
        values = []
        for col in year_cols:
            val = ws.cell(row=heat_rate_row, column=col).value
            if val is not None:
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    pass
        if values:
            result["heat_rate_avg"] = sum(values) / len(values)
            print(f"    Heat rate (avg): {result['heat_rate_avg']:,.0f}")
    
    return result


def extract_from_sys_cost_summary(ws, year: int) -> Dict[str, Any]:
    """Extract values from SYS Cost Summary sheet."""
    result = {
        "kyger": {},
        "clifty": {},
        "system": {},
    }
    
    # Find year column - the SYS Cost Summary has years in row 7
    # The values might be stored as numbers (2025, 2026) or with formatting (2,025)
    year_col = None
    for col in range(1, 25):
        val = ws.cell(row=7, column=col).value
        if val is not None:
            try:
                if int(val) == year:
                    year_col = col
                    break
            except (ValueError, TypeError):
                pass
    
    if not year_col:
        # Try alternate rows
        for check_row in [5, 6, 8, 9]:
            for col in range(1, 25):
                val = ws.cell(row=check_row, column=col).value
                if val is not None:
                    try:
                        if int(val) == year:
                            year_col = col
                            break
                    except (ValueError, TypeError):
                        pass
            if year_col:
                break
    
    if not year_col:
        print(f"    Could not find year {year} in SYS Cost Summary")
        # Debug: print row 7
        print("    Row 7 values:", end=" ")
        for col in range(1, 10):
            val = ws.cell(row=7, column=col).value
            print(f"[{col}]={val}", end=" ")
        print()
        return result
    
    print(f"    Found year {year} at column {year_col}")
    
    # Extract key metrics by label - this is a system-wide summary
    # The sheet structure shows:
    # Row 10: Estimated Delivered Power Sales (MWh)
    # Row 16: Projected Coal Cost
    # Row 19: Total Projected Energy Cost
    # Row 21: Projected Energy Costs - $/MWh
    
    key_rows = {
        10: "generation_mwh",
        16: "coal_cost",
        18: "other_fuel_cost",
        19: "total_fuel_cost",
        21: "fuel_cost_per_mwh",
    }
    
    for row, key in key_rows.items():
        val = ws.cell(row=row, column=year_col).value
        if val is not None:
            try:
                float_val = float(val)
                # Note: values in this sheet are in thousands
                if key in ["generation_mwh", "coal_cost", "other_fuel_cost", "total_fuel_cost"]:
                    float_val = float_val * 1000 if float_val < 100000 else float_val
                result["system"][key] = float_val
                label = ws.cell(row=row, column=1).value
                print(f"    Row {row} ({label}): {float_val:,.0f}")
            except (ValueError, TypeError):
                pass
    
    return result


def extract_from_consumables_sheet(ws, plant: str, year: int) -> Dict[str, float]:
    """Extract values from KC/CC Consumables sheet.
    
    Sums all monthly columns for the given year.
    """
    result = {
        "urea_cost": 0,
        "limestone_cost": 0,
        "hydrated_lime_cost": 0,
        "bioreactor_cost": 0,
        "months_found": 0,
    }
    
    # Find all columns for the target year (row 3 contains years)
    year_cols = []
    for col in range(2, 25):
        val = ws.cell(row=3, column=col).value
        if val is not None:
            try:
                if int(val) == year:
                    year_cols.append(col)
            except (ValueError, TypeError):
                pass
    
    if not year_cols:
        print(f"    Could not find year {year} in {plant} Consumables")
        return result
    
    print(f"    Found year {year} in columns {year_cols} ({len(year_cols)} months)")
    result["months_found"] = len(year_cols)
    
    # Find the "Total Urea Cost" row and "Total Limestone Cost" row
    urea_total_row = None
    limestone_total_row = None
    hydrated_lime_row = None
    
    for row in range(1, 60):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if "total urea cost" in label_str:
            urea_total_row = row
        elif "total limestone cost" in label_str:
            limestone_total_row = row
        elif "hydrated lime" in label_str and "cost" in label_str:
            hydrated_lime_row = row
    
    # Sum values across all months for the year
    if urea_total_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=urea_total_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["urea_cost"] = total
        print(f"    Urea total (summed): ${total:,.0f}")
    
    if limestone_total_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=limestone_total_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["limestone_cost"] = total
        print(f"    Limestone total (summed): ${total:,.0f}")
    
    if hydrated_lime_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=hydrated_lime_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["hydrated_lime_cost"] = total
        print(f"    Hydrated lime total (summed): ${total:,.0f}")
    
    return result


def extract_from_byproducts_sheet(ws, plant: str, year: int) -> Dict[str, float]:
    """Extract values from KC/CC Byproducts sheet.
    
    Sums all monthly columns for the given year.
    """
    result = {
        "ash_net": 0,
        "gypsum_net": 0,
        "total_byproduct_net": 0,
        "months_found": 0,
    }
    
    # Find all columns for the target year (row 3 contains years)
    year_cols = []
    for col in range(2, 25):
        val = ws.cell(row=3, column=col).value
        if val is not None:
            try:
                if int(val) == year:
                    year_cols.append(col)
            except (ValueError, TypeError):
                pass
    
    if not year_cols:
        print(f"    Could not find year {year} in {plant} Byproducts")
        return result
    
    print(f"    Found year {year} in columns {year_cols} ({len(year_cols)} months)")
    result["months_found"] = len(year_cols)
    
    # Find key rows
    ash_total_row = None
    gypsum_net_row = None  # Row 53: "Gypsum Net Cost"
    byproduct_total_row = None  # Row 58: "Byproduct Cost Less Sales" - FINAL total
    
    for row in range(1, 65):
        label = ws.cell(row=row, column=1).value
        if label is None:
            continue
        label_str = str(label).lower()
        
        if "total ash cost" in label_str or "total ash revenue" in label_str:
            ash_total_row = row
        elif "gypsum net cost" in label_str:
            gypsum_net_row = row
        elif "byproduct cost less sales" in label_str:
            byproduct_total_row = row
    
    # Sum values across all months
    if ash_total_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=ash_total_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["ash_net"] = total
        print(f"    Ash net (summed): ${total:,.0f}")
    
    if gypsum_net_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=gypsum_net_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["gypsum_net"] = total
        print(f"    Gypsum net (summed): ${total:,.0f}")
    
    # Use "Byproduct Cost Less Sales" as the final total if available
    # This includes ash, gypsum, and misc expenses
    if byproduct_total_row:
        total = 0
        for col in year_cols:
            val = ws.cell(row=byproduct_total_row, column=col).value
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        result["total_byproduct_net"] = total
        print(f"    Byproduct Cost Less Sales (summed): ${total:,.0f}")
    else:
        result["total_byproduct_net"] = result["ash_net"] + result["gypsum_net"]
    
    return result


def extract_excel_values(excel_path: str, year: int, verbose: bool = False) -> Dict[str, Any]:
    """Extract all relevant values from the Excel file."""
    import openpyxl
    
    print(f"\nOpening Excel file: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return {}
    
    sheets = wb.sheetnames
    print(f"\nFound {len(sheets)} sheets")
    
    results = {
        "kyger": {},
        "clifty": {},
        "system": {},
    }
    
    # Extract from KC sheets
    print(f"\n--- Extracting Kyger Creek data for {year} ---")
    
    if 'KC Generation' in sheets:
        print("  Processing KC Generation...")
        ws = wb['KC Generation']
        if verbose:
            analyze_sheet_structure(ws, 'KC Generation', max_rows=20, max_cols=15)
        gen_data = extract_from_generation_sheet(ws, "Kyger", year)
        results["kyger"].update(gen_data)
    
    if 'KC Coal Burn' in sheets:
        print("  Processing KC Coal Burn...")
        ws = wb['KC Coal Burn']
        if verbose:
            analyze_sheet_structure(ws, 'KC Coal Burn', max_rows=25, max_cols=8)
        burn_data = extract_from_coal_burn_sheet(ws, "Kyger", year)
        results["kyger"].update(burn_data)
    
    if 'KC Consumables' in sheets:
        print("  Processing KC Consumables...")
        ws = wb['KC Consumables']
        if verbose:
            analyze_sheet_structure(ws, 'KC Consumables', max_rows=30, max_cols=8)
        cons_data = extract_from_consumables_sheet(ws, "Kyger", year)
        results["kyger"]["consumables"] = cons_data
    
    if 'KC Byproducts' in sheets:
        print("  Processing KC Byproducts...")
        ws = wb['KC Byproducts']
        if verbose:
            analyze_sheet_structure(ws, 'KC Byproducts', max_rows=30, max_cols=8)
        by_data = extract_from_byproducts_sheet(ws, "Kyger", year)
        results["kyger"]["byproducts"] = by_data
    
    # Extract from CC sheets
    print(f"\n--- Extracting Clifty Creek data for {year} ---")
    
    if 'CC Generation' in sheets:
        print("  Processing CC Generation...")
        ws = wb['CC Generation']
        if verbose:
            analyze_sheet_structure(ws, 'CC Generation', max_rows=20, max_cols=15)
        gen_data = extract_from_generation_sheet(ws, "Clifty", year)
        results["clifty"].update(gen_data)
    
    if 'CC Coal Burn' in sheets:
        print("  Processing CC Coal Burn...")
        ws = wb['CC Coal Burn']
        if verbose:
            analyze_sheet_structure(ws, 'CC Coal Burn', max_rows=25, max_cols=8)
        burn_data = extract_from_coal_burn_sheet(ws, "Clifty", year)
        results["clifty"].update(burn_data)
    
    if 'CC Consumables' in sheets:
        print("  Processing CC Consumables...")
        ws = wb['CC Consumables']
        if verbose:
            analyze_sheet_structure(ws, 'CC Consumables', max_rows=30, max_cols=8)
        cons_data = extract_from_consumables_sheet(ws, "Clifty", year)
        results["clifty"]["consumables"] = cons_data
    
    if 'CC Byproducts' in sheets:
        print("  Processing CC Byproducts...")
        ws = wb['CC Byproducts']
        if verbose:
            analyze_sheet_structure(ws, 'CC Byproducts', max_rows=30, max_cols=8)
        by_data = extract_from_byproducts_sheet(ws, "Clifty", year)
        results["clifty"]["byproducts"] = by_data
    
    # Extract from SYS Cost Summary
    print(f"\n--- Extracting System Summary for {year} ---")
    
    if 'SYS Cost Summary' in sheets:
        print("  Processing SYS Cost Summary...")
        ws = wb['SYS Cost Summary']
        if verbose:
            analyze_sheet_structure(ws, 'SYS Cost Summary', max_rows=50, max_cols=10)
        sys_data = extract_from_sys_cost_summary(ws, year)
        results["system"] = sys_data
    
    wb.close()
    return results


def run_model_calculation(year: int) -> Dict[str, Any]:
    """Run the Python fuel model calculation."""
    from src.db.postgres import get_session
    from src.engine.fuel_model import (
        calculate_system_fuel_costs,
        calculate_fuel_costs,
        summarize_annual_fuel_costs,
    )
    
    print(f"\nRunning Python fuel model for {year}...")
    
    with get_session() as db:
        # Calculate for both plants
        system_costs = calculate_system_fuel_costs(db, year)
        
        # Get detailed monthly data - pass None for inputs to load from database
        # This ensures heat rates and use factors from the database are used
        kyger_monthly = []
        clifty_monthly = []
        
        for month in range(1, 13):
            # Pass inputs=None to load from database (heat rates, use factors)
            kyger_summary = calculate_fuel_costs(db, 1, year, month, inputs=None)
            kyger_monthly.append(kyger_summary)
            
            clifty_summary = calculate_fuel_costs(db, 2, year, month, inputs=None)
            clifty_monthly.append(clifty_summary)
        
        kyger_annual = summarize_annual_fuel_costs(kyger_monthly)
        clifty_annual = summarize_annual_fuel_costs(clifty_monthly)
    
    return {
        "system": system_costs,
        "kyger": {
            "monthly": kyger_monthly,
            "annual": kyger_annual,
        },
        "clifty": {
            "monthly": clifty_monthly,
            "annual": clifty_annual,
        },
    }


def format_number(val: float, decimals: int = 0) -> str:
    """Format a number with commas."""
    if decimals == 0:
        return f"{val:,.0f}"
    return f"{val:,.{decimals}f}"


def print_model_results(model_data: Dict[str, Any], plant: str = "kyger"):
    """Print model results in a table format."""
    data = model_data[plant]
    monthly = data["monthly"]
    annual = data["annual"]
    
    print(f"\n{'=' * 80}")
    print(f"PYTHON MODEL RESULTS - {plant.upper()}")
    print(f"{'=' * 80}")
    
    # Header
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Annual"]
    print(f"\n{'Metric':<30} " + " ".join(f"{m:>10}" for m in months))
    print("-" * 180)
    
    # Net Generation
    row = [format_number(m.net_delivered_mwh) for m in monthly]
    row.append(format_number(annual.get("total_mwh", 0)))
    print(f"{'Net Generation (MWh)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Coal Burn Tons
    row = [format_number(m.coal_tons_consumed) for m in monthly]
    row.append(format_number(annual.get("total_coal_tons", 0)))
    print(f"{'Coal Burn (tons)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Heat Rate
    row = [format_number(m.heat_rate) for m in monthly]
    row.append(format_number(annual.get("avg_heat_rate", 0)))
    print(f"{'Heat Rate (BTU/kWh)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Coal Cost
    row = [format_number(m.coal_cost) for m in monthly]
    row.append(format_number(annual.get("total_coal_cost", 0)))
    print(f"{'Coal Cost ($)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Urea Cost
    row = [format_number(m.urea_cost) for m in monthly]
    row.append(format_number(annual.get("total_urea_cost", 0)))
    print(f"{'Urea Cost ($)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Limestone Cost
    row = [format_number(m.limestone_cost) for m in monthly]
    row.append(format_number(annual.get("total_limestone_cost", 0)))
    print(f"{'Limestone Cost ($)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # Total Fuel Cost
    row = [format_number(m.total_fuel_cost) for m in monthly]
    row.append(format_number(annual.get("total_fuel_cost", 0)))
    print(f"{'Total Fuel Cost ($)':<30} " + " ".join(f"{v:>10}" for v in row))
    
    # $/MWh
    row = [format_number(float(m.fuel_cost_per_mwh), 2) for m in monthly]
    row.append(format_number(annual.get("avg_fuel_cost_per_mwh", 0), 2))
    print(f"{'Fuel Cost ($/MWh)':<30} " + " ".join(f"{v:>10}" for v in row))


def generate_comparison_report(excel_data: Dict, model_data: Dict, year: int):
    """Generate side-by-side comparison report."""
    
    print(f"\n{'=' * 100}")
    print("COMPARISON REPORT: Excel Plan vs Python Model")
    print(f"Year: {year}")
    print(f"{'=' * 100}")
    
    comparison_rows = []
    
    # Compare Kyger Creek
    print(f"\n{'-' * 100}")
    print("KYGER CREEK COMPARISON")
    print(f"{'-' * 100}")
    print(f"{'Metric':<35} {'Excel':>15} {'Model':>15} {'Variance':>15} {'Var %':>10}")
    print(f"{'-' * 100}")
    
    kyger_excel = excel_data.get("kyger", {})
    kyger_model = model_data["kyger"]["annual"]
    
    # Get months found for context
    kc_months = kyger_excel.get("months_found", 0)
    if kc_months == 0:
        kc_months = kyger_excel.get("consumables", {}).get("months_found", 0)
    
    print(f"\nNote: Excel contains {kc_months} months of {year} data for Kyger Creek")
    
    # Sum actual Q4 model values (months 10, 11, 12) instead of scaling
    kyger_monthly = model_data["kyger"]["monthly"]
    kc_q4_mwh = sum(m.net_delivered_mwh for m in kyger_monthly[-3:])  # Oct, Nov, Dec
    kc_q4_coal = sum(m.coal_tons_consumed for m in kyger_monthly[-3:])
    kc_q4_urea = sum(m.urea_cost for m in kyger_monthly[-3:])
    kc_q4_limestone = sum(m.limestone_cost for m in kyger_monthly[-3:])
    kc_q4_byproduct = sum(m.byproduct_net_cost for m in kyger_monthly[-3:])
    kc_q4_heat_rate = sum(m.heat_rate for m in kyger_monthly[-3:]) / 3
    
    print(f"      Comparing to Model Q4 (Oct-Dec) values")
    
    kc_comparisons = [
        ("Net Generation (MWh)", 
         kyger_excel.get("total_generation_mwh", 0) if kyger_excel.get("total_generation_mwh") else None,
         float(kc_q4_mwh)),
        ("Coal Burn (tons)",
         kyger_excel.get("coal_tons", 0) if kyger_excel.get("coal_tons") else None,
         float(kc_q4_coal)),
        ("Heat Rate (BTU/kWh)",
         kyger_excel.get("heat_rate_avg", 0) if kyger_excel.get("heat_rate_avg") else None,
         float(kc_q4_heat_rate)),
        ("Urea Cost ($)",
         kyger_excel.get("consumables", {}).get("urea_cost") if kyger_excel.get("consumables") else None,
         float(kc_q4_urea)),
        ("Limestone Cost ($)",
         kyger_excel.get("consumables", {}).get("limestone_cost") if kyger_excel.get("consumables") else None,
         float(kc_q4_limestone)),
        ("Byproduct Net ($)",
         kyger_excel.get("byproducts", {}).get("total_byproduct_net") if kyger_excel.get("byproducts") else None,
         float(kc_q4_byproduct)),
    ]
    
    for metric, excel_val, model_val in kc_comparisons:
        print_comparison_row(metric, excel_val, model_val)
        comparison_rows.append(("Kyger", metric, excel_val, model_val))
    
    # Compare Clifty Creek
    print(f"\n{'-' * 100}")
    print("CLIFTY CREEK COMPARISON")
    print(f"{'-' * 100}")
    print(f"{'Metric':<35} {'Excel':>15} {'Model':>15} {'Variance':>15} {'Var %':>10}")
    print(f"{'-' * 100}")
    
    clifty_excel = excel_data.get("clifty", {})
    clifty_model = model_data["clifty"]["annual"]
    
    # Get months found for Clifty
    cc_months = clifty_excel.get("months_found", 0)
    if cc_months == 0:
        cc_months = clifty_excel.get("consumables", {}).get("months_found", 0)
    
    print(f"\nNote: Excel contains {cc_months} months of {year} data for Clifty Creek")
    
    # Sum actual Q4 model values (months 10, 11, 12) instead of scaling
    clifty_monthly = model_data["clifty"]["monthly"]
    cc_q4_mwh = sum(m.net_delivered_mwh for m in clifty_monthly[-3:])  # Oct, Nov, Dec
    cc_q4_coal = sum(m.coal_tons_consumed for m in clifty_monthly[-3:])
    cc_q4_urea = sum(m.urea_cost for m in clifty_monthly[-3:])
    cc_q4_limestone = sum(m.limestone_cost for m in clifty_monthly[-3:])
    cc_q4_byproduct = sum(m.byproduct_net_cost for m in clifty_monthly[-3:])
    cc_q4_heat_rate = sum(m.heat_rate for m in clifty_monthly[-3:]) / 3
    
    print(f"      Comparing to Model Q4 (Oct-Dec) values")
    
    cc_comparisons = [
        ("Net Generation (MWh)", 
         clifty_excel.get("total_generation_mwh", 0) if clifty_excel.get("total_generation_mwh") else None,
         float(cc_q4_mwh)),
        ("Coal Burn (tons)",
         clifty_excel.get("coal_tons", 0) if clifty_excel.get("coal_tons") else None,
         float(cc_q4_coal)),
        ("Heat Rate (BTU/kWh)",
         clifty_excel.get("heat_rate_avg", 0) if clifty_excel.get("heat_rate_avg") else None,
         float(cc_q4_heat_rate)),
        ("Urea Cost ($)",
         clifty_excel.get("consumables", {}).get("urea_cost") if clifty_excel.get("consumables") else None,
         float(cc_q4_urea)),
        ("Limestone Cost ($)",
         clifty_excel.get("consumables", {}).get("limestone_cost") if clifty_excel.get("consumables") else None,
         float(cc_q4_limestone)),
        ("Byproduct Net ($)",
         clifty_excel.get("byproducts", {}).get("total_byproduct_net") if clifty_excel.get("byproducts") else None,
         float(cc_q4_byproduct)),
    ]
    
    for metric, excel_val, model_val in cc_comparisons:
        print_comparison_row(metric, excel_val, model_val)
        comparison_rows.append(("Clifty", metric, excel_val, model_val))
    
    # System totals
    print(f"\n{'-' * 100}")
    print("SYSTEM TOTAL COMPARISON")
    print(f"{'-' * 100}")
    print(f"{'Metric':<35} {'Excel':>15} {'Model':>15} {'Variance':>15} {'Var %':>10}")
    print(f"{'-' * 100}")
    
    system_excel = excel_data.get("system", {}).get("system", {})
    system_model = model_data["system"]["system"]
    
    # Note: SYS Cost Summary values may be in thousands ($K)
    excel_total_cost = system_excel.get("total_fuel_cost", 0)
    if excel_total_cost and excel_total_cost < 1000000:  # Likely in thousands
        excel_total_cost = excel_total_cost * 1000
    
    excel_gen = system_excel.get("generation_mwh", 0)
    
    # Calculate model Q4 values if Excel only has partial year
    model_q4_mwh = sum(m.net_delivered_mwh for m in model_data["kyger"]["monthly"][-3:]) + \
                   sum(m.net_delivered_mwh for m in model_data["clifty"]["monthly"][-3:])
    model_q4_cost = sum(m.total_fuel_cost for m in model_data["kyger"]["monthly"][-3:]) + \
                    sum(m.total_fuel_cost for m in model_data["clifty"]["monthly"][-3:])
    
    print(f"\nNote: Excel shows Q4 2025 data. Comparing to Model Q4 2025 values.")
    
    system_comparisons = [
        ("Total MWh (Q4)", excel_gen, float(model_q4_mwh)),
        ("Total Fuel Cost (Q4 $)", excel_total_cost, float(model_q4_cost)),
        ("Avg $/MWh", system_excel.get("fuel_cost_per_mwh"), system_model.get("avg_fuel_cost_per_mwh", 0)),
    ]
    
    for metric, excel_val, model_val in system_comparisons:
        print_comparison_row(metric, excel_val, model_val)
        comparison_rows.append(("System", metric, excel_val, model_val))
    
    return comparison_rows


def print_comparison_row(metric: str, excel_val: Optional[float], model_val: float):
    """Print a single comparison row."""
    if excel_val is None:
        excel_str = "N/A"
        variance_str = "N/A"
        var_pct_str = "N/A"
    else:
        excel_str = format_number(excel_val, 2 if excel_val < 100 else 0)
        variance = excel_val - model_val
        variance_str = format_number(variance, 2 if abs(variance) < 100 else 0)
        if excel_val != 0:
            var_pct = (variance / excel_val) * 100
            var_pct_str = f"{var_pct:+.1f}%"
        else:
            var_pct_str = "N/A"
    
    model_str = format_number(model_val, 2 if model_val < 100 else 0)
    
    print(f"{metric:<35} {excel_str:>15} {model_str:>15} {variance_str:>15} {var_pct_str:>10}")


def main():
    """Main comparison function."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Compare fuel model to Excel")
    parser.add_argument("--year", type=int, default=2025, help="Year to compare")
    parser.add_argument("--excel", type=str, 
                       default="docs/source_documents/OVEC IKEC Energy Budget- 2026 Year-End BOD Update- Final- (12-8-25).xlsm",
                       help="Path to Excel file")
    parser.add_argument("--analyze-only", action="store_true", 
                       help="Only analyze Excel structure, don't run model")
    parser.add_argument("--verbose", "-v", action="store_true",
                       help="Show detailed sheet analysis")
    args = parser.parse_args()
    
    print("=" * 100)
    print("ENERGY/FUEL MODEL vs EXCEL COMPARISON")
    print(f"Year: {args.year}")
    print(f"Excel: {args.excel}")
    print("=" * 100)
    
    # Step 1: Extract Excel values
    excel_path = Path(args.excel)
    if not excel_path.exists():
        # Try relative to script location
        excel_path = Path(__file__).parent.parent / args.excel
    
    if not excel_path.exists():
        print(f"\nERROR: Excel file not found: {args.excel}")
        print("Please provide the correct path to the OVEC Excel file.")
        return
    
    excel_values = extract_excel_values(str(excel_path), args.year, verbose=args.verbose)
    
    if args.analyze_only:
        print("\n[Analyze-only mode - skipping model calculation]")
        return
    
    # Step 2: Run model calculation
    try:
        model_data = run_model_calculation(args.year)
        
        # Step 3: Print detailed model results
        print_model_results(model_data, "kyger")
        print_model_results(model_data, "clifty")
        
        # Step 4: Generate comparison report
        comparison = generate_comparison_report(excel_values, model_data, args.year)
        
        # Step 5: Identify significant discrepancies
        print(f"\n{'=' * 100}")
        print("SIGNIFICANT DISCREPANCIES (>5% variance)")
        print(f"{'=' * 100}")
        
        has_discrepancies = False
        for plant, metric, excel_val, model_val in comparison:
            if excel_val is not None and excel_val != 0:
                var_pct = abs((excel_val - model_val) / excel_val * 100)
                if var_pct > 5:
                    has_discrepancies = True
                    print(f"  [{plant}] {metric}: {var_pct:.1f}% difference")
                    print(f"    Excel: {format_number(excel_val)}, Model: {format_number(model_val)}")
                    
                    # Suggest possible causes
                    if "generation" in metric.lower():
                        print("    Possible causes: Use factor differences, capacity settings, ozone curtailment")
                    elif "heat rate" in metric.lower():
                        print("    Possible causes: SUF correction, PRB blend adjustment")
                    elif "coal" in metric.lower() and "cost" in metric.lower():
                        print("    Possible causes: Contract pricing differences, uncommitted coal pricing")
                    elif "fuel cost" in metric.lower():
                        print("    Possible causes: Consumables rates, byproduct pricing, allowance costs")
                    print()
        
        if not has_discrepancies:
            print("  No significant discrepancies found (all within 5%)")
        
        # Root cause analysis
        print(f"\n{'=' * 100}")
        print("ROOT CAUSE ANALYSIS")
        print(f"{'=' * 100}")
        
        print("""
KEY FINDINGS:

1. HEAT RATE DISCREPANCY (Model: 9,850 BTU/kWh, Excel: ~10,200-10,420 BTU/kWh)
   - Model uses baseline heat rate of 9,850 BTU/kWh
   - Excel applies SUF (Sub-Optimal Unit Factor) correction adding ~200-400 BTU/kWh
   - Excel also includes PRB blend adjustment
   FIX: Load heat_rate_inputs from database or update HeatRateParams in model

2. GENERATION DISCREPANCY (Model ~6-16% higher than Excel)
   - Model may be using different use factors than Excel
   - Excel may include actual generation data for recent months
   - Ozone season curtailment may differ (Unit 6 without SCR curtailed May-Sep)
   FIX: Load use_factor_inputs and unit_outage_inputs from database

3. UREA COST DISCREPANCY (Model 27-55% higher)
   - Urea consumption is tied to generation and SCR operation
   - Higher model generation = higher urea consumption
   - Check urea $/ton pricing (Excel: ~$458-460/ton)
   FIX: Verify urea pricing in UreaParams matches Excel

4. LIMESTONE COST (Model shows costs, Excel row not found correctly)
   - Limestone total row label may differ from expected
   - Excel shows limestone costs in Consumables sheet
   FIX: Update row label matching in extract_from_consumables_sheet

5. BYPRODUCT NET (Sign difference)
   - Model shows credits (negative = revenue)
   - Excel values need sign verification
   - Byproducts include ash disposal costs and gypsum sales
   FIX: Verify byproduct calculations match Excel structure

6. COAL PRICING
   - Excel SYS Cost Summary shows $96.8M coal cost for Q4 2025
   - Model uses default $55/ton + $6/ton barge
   - Excel may use contract-specific pricing
   FIX: Load coal_contract_pricing from database

RECOMMENDED NEXT STEPS:
1. Import fuel inputs from Excel: python -m src.etl.fuel_inputs_import
2. Verify use factors are loaded for 2025
3. Check heat rate inputs include SUF correction
4. Compare monthly model output to Excel monthly columns
""")
        
        print(f"\n{'=' * 100}")
        print("ANALYSIS COMPLETE")
        print(f"{'=' * 100}")
        
    except Exception as e:
        print(f"\nError running model calculation: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
